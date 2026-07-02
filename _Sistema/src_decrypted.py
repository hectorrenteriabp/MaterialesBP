"""
╔══════════════════════════════════════════════════════════════════════════════╗
║    BERRIES PARADISE — Actualizador Dashboard Unificado (Inventarios + MRP)  ║
║                        ACTUALIZAR_BP_UNIFICADO.py                           ║
╠══════════════════════════════════════════════════════════════════════════════╣
║  FUENTE ÚNICA:                                                               ║
║    Referencia_Inventario_BP.xlsx                                             ║
║    Hojas: Bitacora, Inventario, Historico Movs, Capacidad, Master, Costo,   ║
║           Plan (Min/Max por SKU+Almacén), GLM (proveedores)                 ║
║                                                                              ║
║  Output: BP_Dashboard_Unificado.html (HTML autónomo, sin servidor)          ║
║                                                                              ║
║  Uso:                                                                        ║
║    python ACTUALIZAR_BP_UNIFICADO.py                                         ║
║    o doble clic en ACTUALIZAR_BP_UNIFICADO.bat                               ║
╚══════════════════════════════════════════════════════════════════════════════╝
"""

import os, sys, json, re, warnings, traceback, datetime
import pandas as pd
import numpy as np
from pathlib import Path
from collections import defaultdict

warnings.filterwarnings('ignore')

# ══════════════════════════════════════════════════════════════════════════════
# RUTAS — ajusta si cambias la estructura de carpetas
# ══════════════════════════════════════════════════════════════════════════════

if getattr(sys, 'frozen', False):
    BASE_DIR = Path(sys.executable).parent
else:
    BASE_DIR = Path(__file__).parent

# ── ORGANIZACIÓN DE CARPETAS ─────────────────────────────────────────────────
# La RAÍZ (compartida con usuarios) solo contiene lo esencial: el ejecutable y su
# .bat, el dashboard HTML, y los 2 Excel que se editan (inventario y cambios de
# plan). Todo lo demás se reparte en subcarpetas para mantener limpia la vista:
#   /Reportes  → archivos de seguimiento y capacidad (generados / editados)
#   /_Sistema  → estado interno, plantilla, logs, código (no tocar)
REPORTES_DIR = BASE_DIR / "Reportes"
SISTEMA_DIR  = BASE_DIR / "_Sistema"
for _d in (REPORTES_DIR, SISTEMA_DIR):
    try:
        _d.mkdir(exist_ok=True)
    except Exception:
        pass

# ── SUPABASE CONFIG & SYNC ──
SUPABASE_URL = None
SUPABASE_KEY = None

def load_supabase_config():
    global SUPABASE_URL, SUPABASE_KEY
    config_file = SISTEMA_DIR / "supabase_config.json"
    if config_file.exists():
        try:
            with open(config_file, "r", encoding="utf-8") as f:
                cfg = json.load(f)
                SUPABASE_URL = cfg.get("SUPABASE_URL")
                SUPABASE_KEY = cfg.get("SUPABASE_KEY")
            if SUPABASE_URL and SUPABASE_KEY and "your-project-id" not in SUPABASE_URL:
                print(f"  [+] Configuración de Supabase cargada: {SUPABASE_URL}")
                return True
        except Exception as e:
            print(f"  [!] Error cargando supabase_config.json: {e}")
    return False

import urllib.request

def supabase_get(table_name):
    if not SUPABASE_URL or not SUPABASE_KEY or "your-project-id" in SUPABASE_URL:
        return None
    url = f"{SUPABASE_URL.rstrip('/')}/rest/v1/{table_name}"
    req = urllib.request.Request(
        url,
        headers={
            "apikey": SUPABASE_KEY,
            "Authorization": f"Bearer {SUPABASE_KEY}"
        }
    )
    try:
        with urllib.request.urlopen(req, timeout=10) as response:
            return json.loads(response.read().decode('utf-8'))
    except Exception as e:
        print(f"  [!] Error en Supabase GET {table_name}: {e}")
        return None

def supabase_upsert(table_name, data):
    if not SUPABASE_URL or not SUPABASE_KEY or "your-project-id" in SUPABASE_URL:
        return False
    url = f"{SUPABASE_URL.rstrip('/')}/rest/v1/{table_name}"
    payload_bytes = json.dumps(data).encode('utf-8')
    req = urllib.request.Request(
        url,
        data=payload_bytes,
        headers={
            "apikey": SUPABASE_KEY,
            "Authorization": f"Bearer {SUPABASE_KEY}",
            "Content-Type": "application/json",
            "Prefer": "resolution=merge-duplicates"
        },
        method="POST"
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as response:
            return True
    except Exception as e:
        print(f"  [!] Error en Supabase UPSERT {table_name}: {e}")
        return False

load_supabase_config()

# FUENTE ÚNICA — todos los datos vienen de este archivo (RAÍZ, visible)
INVENTARIO_MAESTRO = BASE_DIR / "Referencia_Inventario_BP.xlsx"

TEMPLATE   = SISTEMA_DIR / "_bp_unificado_template.html"
DASHBOARD  = BASE_DIR / "BP_Dashboard_Unificado.html"

# CAMBIOS DE PLAN — archivo compartido con Planeación (RAÍZ, visible)
# Planeación llena la hoja "Cambios Plan" y guarda. El actualizador lo lee en cada run.
CAMBIOS_PLAN_FILE = BASE_DIR / "Cambios_Plan_BP.xlsx"

# ARCHIVOS DE SEGUIMIENTO — generados por el actualizador, llenados por Compras/Almacén
TRACKING_CORTOS_FILE    = REPORTES_DIR / "Tracking_Cortos_BP.xlsx"
TRACKING_TRASLADOS_FILE = REPORTES_DIR / "Tracking_Traslados_BP.xlsx"
TRACKING_GLM_FILE       = REPORTES_DIR / "Tracking_GLM_BP.xlsx"
CAPACIDAD_FILE          = REPORTES_DIR / "Capacidad_Almacenes_BP.xlsx"

# ══════════════════════════════════════════════════════════════════════════════
# CONSTANTES
# ══════════════════════════════════════════════════════════════════════════════

FECHA_CORTE      = None          # None = hoy; ej: "2026-03-19"
DIAS_PENDIENTES  = 30
# MESES_VALIDOS se calcula dinámicamente al cargar los datos
# Se define después de leer Historico Movs (ver más abajo)
MESES_VALIDOS    = []   # placeholder — se rellena automáticamente
SEMANAS_DEMANDA  = 8             # semanas para calcular demanda semanal MRP

PREFIJOS_TIPO = {
    'BX': 'BOXES (BX)',
    'CL': 'CLAMSHELLS (CL)',
    'CV': 'COVERS (CV)',
    'PA': 'PALLETS (PA)',
    'LE': 'LABELS (LE)',
}

# ── ALMACENES OPERATIVOS (solo estos se usan para ABC, capacidad, veracidad) ──
# Fuente: tabla de zonas validada por Gerente Nacional Inventarios
# OTAL se fusiona a MTAL antes de cualquier procesamiento
ALMACENES_OPERATIVOS = [
    'MABS','MIXT','MLAJ','MMAZ','MPAZ','MPER',
    'MQUE','MQUI','MREY','MSAY','MTAC','MTAL',
    'MTAN','MTUX','MVIL','MZIR','6007','5322',
    'MBAJ',
]

# ── MAPA CÓDIGO SAP → CÓDIGO INTERNO (para normalizar hoja Plan) ──────────────
# '5322' (Tacambaro/Berry Wonder) tiene stock en SAP bajo ese mismo código,
# así que NO se convierte — se usa tal cual. El mapa queda vacío por ahora.
ALM_SAP_TO_INTERNO = {
    # Ejemplo: '5999': 'MXXX',
}

# ── ZONAS POR ALMACÉN (para priorizar traslados) ──
ZONA_MAP = {
    'MABS': 'Michoacan', 'MLAJ': 'Michoacan', 'MPAZ': 'Michoacan',
    'MPER': 'Michoacan', 'MREY': 'Michoacan', 'MSAY': 'Michoacan',
    'MTAC': 'Michoacan', 'MZIR': 'Michoacan', '5322': 'Michoacan',
    'MIXT': 'Jalisco',   'MMAZ': 'Jalisco',   'MQUE': 'Jalisco',
    'MQUI': 'Jalisco',   'MTAL': 'Jalisco',   'MTAN': 'Jalisco',
    'MTUX': 'Jalisco',   'MVIL': 'Jalisco',
    '6007': 'Sinaloa',   # Muy lejano — excluir de traslados
    'MBAJ': 'Baja California',
}
# Sinaloa se excluye de traslados por distancia
EXCLUIR_TRASLADO = {'6007', 'MBAJ'}

CAP_RESPALDO = {
    'MABS': 30,  'MREY': 240, 'MMAZ': 55,  'MPAZ': 120,
    'MPER': 140, 'MQUE': 24,  'MQUI': 50,  'MSAY': 211,
    'MTAL': 1784,'MTAN': 36,  'MTUX': 370, 'MIXT': 10,  'MLAJ': 200,
    'MVIL': 50,  'MZIR': 30,  'MTAC': 60,  '5322': 2500,
    'MBAJ': 100,
}
CAP_PREV = {**CAP_RESPALDO}

# ══════════════════════════════════════════════════════════════════════════════
# UTILIDADES
# ══════════════════════════════════════════════════════════════════════════════

def col(df, *candidates):
    """Devuelve el primer nombre de columna que exista, ignorando acentos/mayúsculas."""
    norm = lambda s: (
        s.lower()
        .replace('á','a').replace('é','e').replace('í','i')
        .replace('ó','o').replace('ú','u').replace('ñ','n').strip()
    )
    mapping = {norm(c): c for c in df.columns}
    for cand in candidates:
        if norm(cand) in mapping:
            return mapping[norm(cand)]
    raise ValueError(
        f"No se encontró ninguna de las columnas {candidates}.\n"
        f"Columnas disponibles: {list(df.columns)}"
    )

def safe_numeric(series, fill=0):
    return pd.to_numeric(series, errors='coerce').fillna(fill)

def parse_date(x):
    """
    Parsea fechas de la hoja Bitacora.
    Regla BP: Excel almacena YYYY-DD-MM.
    Si ts.day == 3 y ts.month > 3 → el mes real es 3, el día es ts.month.
    """
    if pd.isna(x):
        return pd.NaT
    if isinstance(x, (pd.Timestamp, datetime.datetime, datetime.date)):
        ts = pd.Timestamp(x)
        if ts.day == 3 and ts.month > 3:
            try:
                return pd.Timestamp(ts.year, 3, ts.month)
            except Exception:
                return ts
        return ts
    s = str(x).strip()
    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})$', s)
    if m:
        a, b, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        return pd.Timestamp(y, b, a) if a > 12 else pd.Timestamp(y, a, b)
    try:
        return pd.to_datetime(s)
    except Exception:
        return pd.NaT

def parse_hist_date(x):
    """
    Parsea fechas del Histórico de Movimientos. Formato real SAP: DD/MM/YYYY.

    PROBLEMA conocido (export SAP→Excel): las fechas con día <= 12 se
    auto-convierten a datetime interpretándolas como MM/DD (invierten día y mes),
    mientras que las de día > 12 quedan como texto DD/MM correcto. Esta función
    revierte esa inversión para recuperar la fecha real.

    Maneja: datetime/Timestamp (revierte inversión si día<=12), Excel serial
    (int 40000-60000) y strings DD/MM/YYYY.
    """
    if pd.isna(x):
        return pd.NaT
    if isinstance(x, (pd.Timestamp, datetime.datetime)):
        ts = pd.Timestamp(x)
        # día y mes <= 12 => Excel pudo leerlo como MM/DD. El real es DD/MM => invertir.
        # Excepción: si invertir cae en el futuro y el valor directo es válido (pasado),
        # Excel ya lo había leído bien (las cargas SAP mezclan config DD/MM y MM/DD).
        if ts.day <= 12 and ts.month <= 12 and ts.day != ts.month:
            try:
                inv = pd.Timestamp(year=ts.year, month=ts.day, day=ts.month)
            except Exception:
                return ts
            try:
                _hoy = pd.Timestamp(TODAY).normalize()
            except Exception:
                _hoy = pd.Timestamp.today().normalize()
            if inv > _hoy and ts <= _hoy:
                return ts
            return inv
        return ts
    if isinstance(x, (int, float)) and 40000 < x < 60000:
        return pd.Timestamp('1899-12-30') + pd.Timedelta(days=int(x))
    try:
        return pd.to_datetime(str(x), dayfirst=True)
    except Exception:
        return pd.NaT

def fmt_date(v):
    if v is None:
        return None
    if hasattr(v, 'strftime'):
        return v.strftime('%Y-%m-%d')
    return str(v)[:10]

def step(msg): print(f"\n{'-'*60}\n  {msg}...")
def ok(msg):   print(f"  OK: {msg}")
def warn(msg): print(f"  AVISO: {msg}")


# ══════════════════════════════════════════════════════════════════════════════
# INICIO
# ══════════════════════════════════════════════════════════════════════════════

TODAY = pd.Timestamp(FECHA_CORTE) if FECHA_CORTE else pd.Timestamp.today().normalize()

print("=" * 62)
print("  BERRIES PARADISE - Dashboard Unificado (Inventarios + MRP)")
print("=" * 62)
print(f"  Corte:    {TODAY.strftime('%d/%m/%Y')}")
print(f"  Fuente:   {INVENTARIO_MAESTRO}")

# ── VERIFICAR ARCHIVOS ────────────────────────────────────────────────────────
if not INVENTARIO_MAESTRO.exists():
    print(f"\nERROR: No se encontró el archivo maestro:")
    print(f"  {INVENTARIO_MAESTRO}")
    sys.exit(1)

if not TEMPLATE.exists():
    print(f"\nERROR: No se encontró el template HTML:")
    print(f"  {TEMPLATE}")
    sys.exit(1)

# ══════════════════════════════════════════════════════════════════════════════
#  LEER EXCEL — FUENTE ÚNICA
# ══════════════════════════════════════════════════════════════════════════════

step("Leyendo Referencia_Inventario_BP.xlsx")

# Intentar abrir el Excel con engine explícito (más robusto que auto-detect)
_engines_to_try = [('openpyxl', None), ('xlrd', None)]
xl = None
_last_err = None
for _eng, _ in _engines_to_try:
    try:
        xl = pd.ExcelFile(str(INVENTARIO_MAESTRO), engine=_eng)
        ok(f"Engine utilizado: {_eng}")
        break
    except Exception as _e:
        _last_err = _e
        continue

if xl is None:
    # Diagnóstico adicional
    try:
        _fsize = INVENTARIO_MAESTRO.stat().st_size
        with open(str(INVENTARIO_MAESTRO), 'rb') as _fcheck:
            _magic = _fcheck.read(4)
        _nonzero = sum(1 for b in _magic if b != 0)
        if _nonzero == 0:
            print("\n" + "="*62)
            print("  [!]  El archivo parece estar vacío o dañado")
            print(f"  Tamaño: {_fsize:,} bytes  |  Magic: {_magic.hex()}")
            print("  Cierra Excel si está abierto y vuelve a intentar.")
            print("="*62 + "\n")
        else:
            print(f"\n[!]  No se pudo leer el archivo Excel.")
            print(f"  Error: {_last_err}")
            print("  Asegúrate de que no esté abierto en Excel.\n")
    except Exception:
        pass
    sys.exit(1)
sheets_available = xl.sheet_names
ok(f"Hojas disponibles: {sheets_available}")

def load_sheet(name, required=True):
    if name not in sheets_available:
        if required:
            raise ValueError(f"Hoja '{name}' no encontrada. Hojas: {sheets_available}")
        return None
    df = xl.parse(name)
    df.columns = [str(c).strip() for c in df.columns]
    return df

# ── RECONSTRUCCIÓN DE 'Historico Movs' DESDE 'Raw Movs' ──────────────────────
# Reemplaza el Power Query externo (borrado) que transformaba el export crudo de
# SAP ('Raw Movs', estructura jerárquica artículo→detalle) al formato tabular
# 'Historico Movs'. Se hace EN MEMORIA: NO se modifica el Excel maestro.
RAW_MOVS_SHEET = 'Raw Movs'
HIST_COLS_STD = ['Número de artículo', 'Descripción', 'Fecha del sistema',
                 'Fecha de contabilización', 'Documento', 'Almacén', 'Cantidad',
                 'Costos', 'Valor trans.', 'Cantidad acumulada', 'Valor acumulado']
LOG_FECHAS_AMBIGUAS = SISTEMA_DIR / '_log_fechas_ambiguas.csv'

def _cargar_historico_desde_raw(xl_file):
    """Transforma la hoja 'Raw Movs' (export crudo SAP) al formato 'Historico Movs':
      1. Detecta la fila de encabezados y descarta la fila vacía superior.
      2. Rellena hacia abajo (ffill) Código y Descripción — en SAP el código solo
         aparece en la fila cabecera de cada artículo; el detalle va sin código.
      3. Filtra las filas que NO son movimientos reales: cabeceras de artículo
         (sin Documento/Cantidad) y los renglones 'Saldo inicial'.
    Devuelve un DataFrame con columnas estándar, o None si la hoja no existe."""
    if RAW_MOVS_SHEET not in xl_file.sheet_names:
        return None
    raw = xl_file.parse(RAW_MOVS_SHEET, header=None)
    if raw.empty or raw.shape[1] < len(HIST_COLS_STD):
        return None
    # 1. Detectar fila de encabezados (contiene 'núm…artículo' en la primera col)
    hdr_idx = None
    for _i in range(min(15, len(raw))):
        _v0 = str(raw.iloc[_i, 0]).lower()
        if 'art' in _v0 and ('num' in _v0 or 'mero' in _v0):
            hdr_idx = _i
            break
    _start = (hdr_idx + 1) if hdr_idx is not None else 0
    df = raw.iloc[_start:, :len(HIST_COLS_STD)].copy()
    df.columns = HIST_COLS_STD
    # 2. Fill-down de código y descripción (estructura jerárquica SAP)
    df['Número de artículo'] = df['Número de artículo'].ffill()
    df['Descripción']        = df['Descripción'].ffill()
    # 3. Mantener solo movimientos reales
    _doc  = df['Documento'].astype(str).str.strip()
    _cant = pd.to_numeric(df['Cantidad'], errors='coerce')
    _keep = (df['Documento'].notna()
             & (_doc != '') & (_doc.str.lower() != 'nan')
             & (_doc.str.lower() != 'saldo inicial')
             & _cant.notna())
    df = df[_keep].copy()
    df['Cantidad'] = _cant[_keep]
    return df.reset_index(drop=True)

def _log_fechas_dudosas(df):
    """Registra en CSV las filas con fecha de contabilización dudosa: datetime
    ambiguo (día y mes <= 12 con día != mes, que se invierte por heurística) y las
    que quedan en el futuro tras corregir (señal de inversión incorrecta)."""
    try:
        _fc = 'Fecha de contabilización'
        if _fc not in df.columns:
            return
        _hoy = pd.Timestamp(TODAY).normalize()
        _orig = df[_fc]
        _corr = _orig.apply(parse_hist_date)
        _amb = _orig.apply(lambda v: isinstance(v, (pd.Timestamp, datetime.datetime))
                           and v.day <= 12 and v.month <= 12 and v.day != v.month)
        _fut = _corr.notna() & (_corr > _hoy)
        _mask = _amb | _fut
        if bool(_mask.any()):
            out = pd.DataFrame({
                'codigo':          df.loc[_mask, 'Número de artículo'],
                'almacen':         df.loc[_mask, 'Almacén'],
                'documento':       df.loc[_mask, 'Documento'],
                'valor_original':  _orig[_mask].astype(str),
                'fecha_corregida': _corr[_mask].dt.strftime('%Y-%m-%d'),
                'futura':          _fut[_mask],
            })
            out.to_csv(str(LOG_FECHAS_AMBIGUAS), index=False, encoding='utf-8-sig')
            warn(f"Fechas dudosas: {int(_mask.sum())} filas (de {len(df)}) -> "
                 f"revisar {LOG_FECHAS_AMBIGUAS.name}")
    except Exception as _e:
        warn(f"No se pudo generar log de fechas dudosas: {_e}")

df_bit  = load_sheet('Bitacora')
df_inv  = load_sheet('Inventario')
# Histórico de Movimientos: reconstruido EN MEMORIA desde 'Raw Movs' (reemplaza el
# Power Query borrado). La hoja estática 'Historico Movs' es solo un RESPALDO
# opcional: si no existe, el dashboard funciona igual con 'Raw Movs'.
df_hist = _cargar_historico_desde_raw(xl)
if df_hist is None or df_hist.empty:
    warn("Hoja 'Raw Movs' no disponible o vacía — intentando respaldo 'Historico Movs'")
    df_hist = load_sheet('Historico Movs', required=False)
    if df_hist is None or df_hist.empty:
        print("\n" + "="*62)
        print("  [!]  No hay datos de movimientos.")
        print("  Falta la hoja 'Raw Movs' (export de SAP) y tampoco existe el")
        print("  respaldo 'Historico Movs'. Pega el export de SAP en 'Raw Movs'")
        print("  y vuelve a ejecutar.")
        print("="*62 + "\n")
        sys.exit(1)
    ok(f"Histórico desde respaldo 'Historico Movs': {len(df_hist):,} filas")
else:
    ok(f"Histórico reconstruido desde 'Raw Movs': {len(df_hist):,} movimientos | "
       f"{df_hist['Número de artículo'].nunique()} SKUs")
    _log_fechas_dudosas(df_hist)
df_mas  = load_sheet('Master')

df_cos = load_sheet('Costo', required=False)
if df_cos is None:
    df_cos = load_sheet('Cost', required=False)
if df_cos is None:
    raise ValueError("No se encontró hoja 'Costo' ni 'Cost'.")

df_plan    = load_sheet('Plan',      required=False)
df_glm_inv = load_sheet('GLM',      required=False)
df_cap_raw = load_sheet('Capacidad', required=False)
df_emb     = load_sheet('Embalajes', required=False)
df_six     = load_sheet('Six',       required=False)

ok(f"Bitacora: {len(df_bit)} | Inventario: {len(df_inv)} | Historico: {len(df_hist)}")

# ══════════════════════════════════════════════════════════════════════════════
#  MÓDULO 1 — INVENTARIOS
# ══════════════════════════════════════════════════════════════════════════════

# ── CAPACIDAD ─────────────────────────────────────────────────────────────────
step("Leyendo capacidades")
CAP_NEW = {}
if df_cap_raw is not None:
    try:
        df_cap_data = df_cap_raw.iloc[2:].copy()
        df_cap_data.columns = ['_', 'Zona', 'Almacen', 'Cap'] + list(df_cap_data.columns[4:])
        for _, r in df_cap_data.iterrows():
            alm = str(r['Almacen']).strip() if pd.notna(r['Almacen']) else ''
            cap = pd.to_numeric(r['Cap'], errors='coerce')
            if alm.startswith('M') and not pd.isna(cap):
                CAP_NEW[alm] = int(cap)
        ok(f"Capacidad del sheet: {len(CAP_NEW)} almacenes")
    except Exception as e:
        warn(f"Sheet Capacidad error ({e}). Usando respaldo.")

if not CAP_NEW:
    CAP_NEW = CAP_RESPALDO.copy()
    warn("Usando capacidades de respaldo.")

# ── Aplicar overrides manuales desde pestaña _Track_Capacidad del Excel ─────────
def _apply_cap_overrides(cap_dict_source):
    for _alm, _v in cap_dict_source.items():
        try:
            _nc = int(_v.get('new_cap', 0)) if isinstance(_v, dict) else int(float(str(_v)))
            if _nc > 0:
                CAP_NEW[_alm] = _nc
        except Exception:
            pass
    return len(cap_dict_source)

_cap_overrides = 0
_TRACKING_FILE_EARLY = BASE_DIR / '_Tracking_BP.xlsx' if (BASE_DIR / '_Tracking_BP.xlsx').exists() else REPORTES_DIR / '_Tracking_BP.xlsx'
try:
    import openpyxl as _xl_cap
    if _TRACKING_FILE_EARLY.exists():
        _wb_cap = _xl_cap.load_workbook(str(_TRACKING_FILE_EARLY), read_only=True, data_only=True)
        if '_Track_Capacidad' in _wb_cap.sheetnames:
            _ws_cap = _wb_cap['_Track_Capacidad']
            _cap_tab_data = {}
            _cap_rows = list(_ws_cap.values)
            for _r in _cap_rows[1:]:
                if _r and _r[0] and _r[1] is not None:
                    _cap_tab_data[str(_r[0]).strip()] = {'new_cap': _r[1]}
            _wb_cap.close()
            _cap_overrides = _apply_cap_overrides(_cap_tab_data)
            if _cap_overrides:
                ok(f"Capacidades editadas (_Tracking_BP.xlsx): {_cap_overrides} almacenes")
        else:
            _wb_cap.close()
except Exception as _ce:
    warn(f"No se pudo leer _Track_Capacidad de _Tracking_BP.xlsx: {_ce}")

try:
    if INVENTARIO_MAESTRO.exists() and '_Track_Capacidad' in _xl_cap.load_workbook(
            str(INVENTARIO_MAESTRO), read_only=True).sheetnames:
        _wb_cap2 = _xl_cap.load_workbook(str(INVENTARIO_MAESTRO), read_only=True, data_only=True)
        _ws_cap2 = _wb_cap2['_Track_Capacidad']
        _cap_tab_data2 = {}
        for _r in list(_ws_cap2.values)[1:]:
            if _r and _r[0] and _r[1] is not None:
                _cap_tab_data2[str(_r[0]).strip()] = {'new_cap': _r[1]}
        _wb_cap2.close()
        _n = _apply_cap_overrides(_cap_tab_data2)
        if _n: ok(f"Capacidades editadas (Excel principal, retrocompat): {_n} almacenes")
except Exception:
    pass

try:
    _tc_path = SISTEMA_DIR / '_tracking_capacidad.json'
    if _tc_path.exists():
        with open(str(_tc_path), 'r', encoding='utf-8') as _f:
            _tracking_cap_json = json.load(_f)
        if _tracking_cap_json:
            _apply_cap_overrides(_tracking_cap_json)
            ok(f"Capacidades editadas (JSON fallback): {len(_tracking_cap_json)} almacenes")
except Exception as _ce2:
    pass

# ── Leer capacidades desde Supabase (o de Capacidad_Almacenes_BP.xlsx como fallback) ──
_cap_loaded_from_supabase = False
if SUPABASE_URL and SUPABASE_KEY:
    try:
        step("Descargando capacidades de almacenes desde Supabase...")
        _sup_caps = supabase_get("capacidades")
        if _sup_caps is not None:
            _cap_ext_overrides = 0
            for _row in _sup_caps:
                _alm_ext = str(_row.get('alm', '')).strip()
                _nc_ext = _row.get('nueva_capacidad')
                _ca_ext = _row.get('capacidad_actual')
                _use_val = _nc_ext if (_nc_ext is not None and str(_nc_ext).strip() not in ('', 'None', 'nan')) else _ca_ext
                if _use_val is not None:
                    try:
                        _v = int(float(str(_use_val)))
                        if _v > 0:
                            CAP_NEW[_alm_ext] = _v
                            _cap_ext_overrides += 1
                    except Exception:
                        pass
            if _cap_ext_overrides > 0:
                ok(f"Capacidades descargadas desde Supabase: {_cap_ext_overrides} almacenes")
                _cap_loaded_from_supabase = True
    except Exception as _esup:
        warn(f"Error descargando capacidades de Supabase: {_esup}")

if not _cap_loaded_from_supabase:
    _cap_ext_path = REPORTES_DIR / "Capacidad_Almacenes_BP.xlsx"
    try:
        if _cap_ext_path.exists():
            import openpyxl as _xl_cap_ext
            _wb_cap_ext = _xl_cap_ext.load_workbook(str(_cap_ext_path), read_only=True, data_only=True)
            if 'Capacidad' in _wb_cap_ext.sheetnames:
                _ws_cap_ext = _wb_cap_ext['Capacidad']
                _rows_cap_ext = list(_ws_cap_ext.values)
                _cap_ext_overrides = 0
                for _r in _rows_cap_ext[1:]:   # saltar encabezado
                    if _r and _r[0]:
                        _alm_ext  = str(_r[0]).strip()
                        _nc_ext   = _r[2] if len(_r) > 2 else None
                        _ca_ext   = _r[1] if len(_r) > 1 else None
                        _use_val  = _nc_ext if (_nc_ext is not None and str(_nc_ext).strip() not in ('', 'None', 'nan')) else _ca_ext
                        if _use_val is not None:
                            try:
                                _v = int(float(str(_use_val)))
                                if _v > 0:
                                    CAP_NEW[_alm_ext] = _v
                                    _cap_ext_overrides += 1
                            except Exception:
                                pass
            _wb_cap_ext.close()
            if _cap_ext_overrides:
                ok(f"Capacidades desde Capacidad_Almacenes_BP.xlsx: {_cap_ext_overrides} almacenes")
    except Exception as _ce3:
        warn(f"No se pudo leer Capacidad_Almacenes_BP.xlsx: {_ce3}")

M_ALMS = [a for a in ALMACENES_OPERATIVOS if a in CAP_NEW or a == '6007']
for _a in ALMACENES_OPERATIVOS:
    if _a not in CAP_NEW and _a in CAP_RESPALDO:
        CAP_NEW[_a] = CAP_RESPALDO[_a]
M_ALMS = list(ALMACENES_OPERATIVOS)
cap_changes = {
    k: {'prev': CAP_PREV.get(k, 0), 'new': CAP_NEW[k],
        'diff': CAP_NEW[k] - CAP_PREV.get(k, 0)}
    for k in CAP_NEW if CAP_PREV.get(k, CAP_NEW[k]) != CAP_NEW[k]
}

# ── CATÁLOGOS ──────────────────────────────────────────────────────────────────
step("Construyendo catálogos (costo, pallet, descripción)")

cos_map  = {}
desc_map = {}
c_sku  = col(df_cos, 'Número de artículo', 'Numero de articulo', 'SKU', 'Artículo', 'Articulo', 'Código', 'Codigo')
c_cost = col(df_cos, 'Costo del artículo', 'Costo del articulo', 'Costo', 'Precio')
c_desc = col(df_cos, 'Descripción del artículo', 'Descripcion del articulo', 'Descripción', 'Descripcion', 'Nombre')
for _, r in df_cos.iterrows():
    sku = str(r[c_sku]).strip()
    cos_map[sku]  = float(safe_numeric(pd.Series([r[c_cost]]))[0])
    desc_map[sku] = str(r[c_desc])[:45]
ok(f"Costos: {len(cos_map)} artículos")

pallet_map = {}
c_mas_sku = col(df_mas, 'SKU', 'Número de artículo', 'Numero de articulo', 'Artículo', 'Codigo')
c_mas_qp  = col(df_mas, 'Qty x Pallet', 'Qty Pallet', 'QtyxPallet', 'Pallet', 'Cantidad por pallet')
for _, r in df_mas.iterrows():
    sku = str(r[c_mas_sku]).strip()
    qp  = float(safe_numeric(pd.Series([r[c_mas_qp]]), fill=500)[0])
    pallet_map[sku] = qp if qp > 0 else 500.0
ok(f"Pallets: {len(pallet_map)} artículos")

# ── MERGE OTAL → MTAL ─────────────────────────────────────────────────────────
step("Fusionando OTAL -> MTAL")

def merge_otal(df, *col_candidates):
    try:
        c = col(df, *col_candidates)
        before = (df[c].astype(str).str.strip().str.upper() == 'OTAL').sum()
        df[c] = df[c].astype(str).str.strip().str.replace(r'^OTAL$', 'MTAL', regex=True, case=False)
        return df, c, before
    except Exception:
        return df, None, 0

df_bit,  c_bit_alm,  n1 = merge_otal(df_bit,  'Almacén', 'Almacen', 'Column1')
df_hist, c_hist_alm, n3 = merge_otal(df_hist, 'Almacén', 'Almacen')
ok(f"Renombrados -> Bitacora: {n1} | Historico: {n3}")

# ── BITÁCORA — MAXif ──────────────────────────────────────────────────────────
step("Procesando Bitácora (MAXif por Codigo+Almacen)")

c_cod   = col(df_bit, 'Codigo', 'Código', 'Cod', 'Artículo', 'Articulo')
c_fecha = col(df_bit, 'Fecha')
c_delta = col(df_bit, 'Delta')
c_sap   = col(df_bit, 'Stock SAP', 'SAP', 'StockSAP')
c_fis   = col(df_bit, 'Conteo Fisico', 'Conteo Físico', 'ConFisico', 'Conteo')

df_bit['_Fecha'] = df_bit[c_fecha].apply(parse_date)
df_bit = df_bit.dropna(subset=['_Fecha'])
df_bit['_Cod']   = df_bit[c_cod].astype(str).str.strip()
df_bit['_Alm']   = df_bit[c_bit_alm].astype(str).str.strip() if c_bit_alm else ''
df_bit['_Delta'] = safe_numeric(df_bit[c_delta])
df_bit['_SAP']   = safe_numeric(df_bit[c_sap])
df_bit['_Fis']   = safe_numeric(df_bit[c_fis])

df_bit = df_bit[df_bit['_Alm'].isin(M_ALMS) & (df_bit['_Cod'].str.len() == 6)]
df_bit['_r'] = df_bit.groupby(['_Cod', '_Alm'])['_Fecha'].rank(method='first', ascending=False)
df_max = df_bit[df_bit['_r'] == 1].copy()

df_max['costo_unit']  = df_max['_Cod'].map(cos_map).fillna(0)
df_max['impacto_abs'] = (df_max['_Delta'].abs() * df_max['costo_unit']).round(2)
df_max['mxn_signed']  = (df_max['_Delta'] * df_max['costo_unit']).round(2)
df_max['desc']        = df_max['_Cod'].map(desc_map).fillna('N/D')

ok(f"Registros MAXif: {len(df_max)} | "
   f"Fecha min: {df_max['_Fecha'].min().date()} | "
   f"max: {df_max['_Fecha'].max().date()}")

total_sap        = df_max['_SAP'].sum()
total_delta_abs  = df_max['_Delta'].abs().sum()
ver_piezas       = round((1 - total_delta_abs / total_sap) * 100, 2) if total_sap > 0 else 0.0
total_conteos    = len(df_max)
total_piezas_dif = int(total_delta_abs)

ver_alm = df_max.groupby('_Alm').agg(
    sap      =('_SAP',   'sum'),
    delta_abs=('_Delta', lambda x: x.abs().sum()),
    n_conteos=('_Cod',   'count'),
    exactos  =('_Delta', lambda x: (x == 0).sum())
).reset_index().rename(columns={'_Alm': 'Almacén'})
ver_alm['ver_pct'] = (
    (1 - ver_alm['delta_abs'] / ver_alm['sap'].replace(0, 1)) * 100
).clip(0, 100).round(2)
ver_alm['cob_pct'] = (
    ver_alm['exactos'] / ver_alm['n_conteos'] * 100
).round(1)
ver_alm = ver_alm.sort_values('ver_pct', ascending=True)

df_neg  = df_max[df_max['_Delta'] < 0]
df_pos  = df_max[df_max['_Delta'] > 0]
imp_neg = df_neg.groupby('_Alm')['mxn_signed'].sum().reset_index()
imp_neg.columns = ['alm', 'mxn']
imp_pos = df_pos.groupby('_Alm')['mxn_signed'].sum().reset_index()
imp_pos.columns = ['alm', 'mxn']
imp_neg = imp_neg.sort_values('mxn', ascending=True)
imp_pos = imp_pos.sort_values('mxn', ascending=False)

total_faltante    = float(round(df_neg['mxn_signed'].sum(), 0))
total_sobrante    = float(round(df_pos['mxn_signed'].sum(), 0))
total_neto        = round(total_sobrante + total_faltante, 0)
total_dif_count   = int((df_max['_Delta'] != 0).sum())
total_dif_pzs_neg = int(df_neg['_Delta'].abs().sum())
total_dif_pzs_pos = int(df_pos['_Delta'].sum())

diffs_all = df_max[df_max['_Delta'] != 0].copy()
diffs_all = diffs_all.sort_values('impacto_abs', ascending=False).head(500)
diffs_out = diffs_all[
    ['_Cod','desc','_Alm','_Fecha','_Delta','_SAP','_Fis','mxn_signed']
].copy()
diffs_out.columns = ['Codigo','desc','Almacén','Fecha','Delta','Stock SAP','Conteo Fisico','mxn_signed']
diffs_out['Fecha']         = diffs_out['Fecha'].dt.strftime('%d/%m/%y')
diffs_out['Stock SAP']     = diffs_out['Stock SAP'].astype(int)
diffs_out['Conteo Fisico'] = diffs_out['Conteo Fisico'].astype(int)

net_pos = df_max[df_max['_Delta'] > 0][['_Cod','desc','_Alm','_Delta','mxn_signed']].copy()
net_neg = df_max[df_max['_Delta'] < 0][['_Cod','_Alm','_Delta','mxn_signed']].copy()
net_pos.columns = ['Codigo','desc','alm_sobre','delta_sobre','mxn_sobre']
net_neg.columns = ['Codigo','alm_falta','delta_falta','mxn_falta']
net_merged = net_pos.merge(net_neg, on='Codigo')
net_merged['neto_pzs'] = net_merged['delta_sobre'] + net_merged['delta_falta']
net_merged['neto_mxn'] = net_merged['mxn_sobre']   + net_merged['mxn_falta']
neteos_list = (
    net_merged.sort_values('mxn_sobre', ascending=False)
    .head(20).round(0).to_dict(orient='records')
)
ok(f"Veracidad: {ver_piezas}% | Diferencias: {total_dif_count} | Neteos: {len(neteos_list)}")

# ── INVENTARIO — Stock por SKU ─────────────────────────────────────────────────
step("Procesando Inventario")

c_inv_cod = col(df_inv, 'Número de artículo', 'Numero de articulo', 'Código', 'Codigo', 'SKU')
c_inv_stk = col(df_inv, 'En stock', 'Enstock', 'Stock', 'Cantidad')
c_inv_alm = None
for candidate in ['Código de almacén', 'Codigo de almacen', 'Almacén', 'Almacen', 'Nombre de almacén']:
    try:
        c_inv_alm = col(df_inv, candidate)
        break
    except ValueError:
        pass

df_inv['_Cod'] = df_inv[c_inv_cod].astype(str).str.strip()
df_inv['_Stk'] = safe_numeric(df_inv[c_inv_stk])

df_inv_all = df_inv[df_inv['_Cod'].str.len() == 6].copy()

ALM_EXCLUIR_ABC = {'M90', 'MATE2122', 'MTAL2'}
ALM_EXCLUIR_NAC = {'M90', 'MATE2122', 'MTAL2'}
ALMACENES_NAC = [a for a in ALMACENES_OPERATIVOS if a.startswith('M') and a not in ALM_EXCLUIR_NAC]

if c_inv_alm:
    df_inv_all['_Alm_code'] = df_inv_all[c_inv_alm].astype(str).str.strip()
    df_inv_all['_Alm_code'] = df_inv_all['_Alm_code'].replace('OTAL', 'MTAL')
else:
    df_inv_all['_Alm_code'] = ''

df_inv_m_only = df_inv_all[
    df_inv_all['_Alm_code'].isin(ALMACENES_NAC) & (df_inv_all['_Stk'] > 0)
].copy()
inv_total_by_sku = df_inv_m_only.groupby('_Cod')['_Stk'].sum().reset_index()
inv_total_by_sku.columns = ['Código', 'En stock']
total_inv_codes = int(inv_total_by_sku['Código'].nunique())

inv_by_alm = (df_inv_all[df_inv_all['_Stk'] > 0]
              .groupby(['_Cod', '_Alm_code'])['_Stk']
              .sum().reset_index())
inv_by_alm.columns = ['Codigo', 'Almacen_code', 'Stock']

inv_pos = df_max[df_max['_SAP'] > 0][['_Cod', '_Alm', '_SAP']].copy()
inv_pos.columns = ['Código', 'alm', 'En stock']
inv_pos_desc = inv_pos.copy()
inv_pos_desc['desc'] = inv_pos_desc['Código'].map(desc_map).fillna('N/D')

ok(f"SKUs únicos con stock (total nacional): {total_inv_codes}")

cutoff_30  = TODAY - pd.Timedelta(days=DIAS_PENDIENTES)
counted_30 = df_max[df_max['_Fecha'] >= cutoff_30][['_Cod','_Alm']].drop_duplicates()
counted_30.columns = ['Código','alm']

# FIX #3: Base de pendientes = inventario REAL por almacén (inv_by_alm), no bitácora (inv_pos)
# Excluye 6007 (Sinaloa) — no aplica conteos físicos en ese almacén
EXCLUIR_CONTEOS = {'6007'}
ALMS_CONTEO = [a for a in M_ALMS if a not in EXCLUIR_CONTEOS]

inv_real_alm = inv_by_alm[
    inv_by_alm['Almacen_code'].isin(ALMS_CONTEO) &
    (inv_by_alm['Stock'] > 0) &
    (inv_by_alm['Codigo'].str.len() == 6)
].copy()
inv_real_alm.columns = ['Código', 'alm', 'En stock']
inv_real_alm['desc'] = inv_real_alm['Código'].map(desc_map).fillna('N/D')

inv_30   = inv_real_alm.groupby('alm')['Código'].nunique().reset_index()
inv_30.columns = ['alm','en_inv']
# counted_30 se restringe a los mismos almacenes de conteo para no inflar el numerador
cov_30   = counted_30[counted_30['alm'].isin(ALMS_CONTEO)].groupby('alm')['Código'].nunique().reset_index()
cov_30.columns = ['alm','contados']
cov_merged = inv_30.merge(cov_30, on='alm', how='left').fillna(0)
# % por almacén: clipeado a 100 (no puede haber más contados que en inventario)
cov_merged['pct'] = (cov_merged['contados'] / cov_merged['en_inv'] * 100).clip(upper=100).round(1)
cov_merged = cov_merged.sort_values('alm')
# Cobertura global: pares únicos (SKU×Almacén) contados ÷ pares únicos (SKU×Almacén) con stock
# Calculado a nivel de pares para evitar sobre-conteo cuando un almacén tiene más conteos que SKUs en inv
_pares_con_stock     = set(zip(inv_real_alm['Código'], inv_real_alm['alm']))
_counted_filtrado    = counted_30[counted_30['alm'].isin(ALMS_CONTEO)]
_pares_contados      = set(zip(_counted_filtrado['Código'], _counted_filtrado['alm']))
_pares_cubiertos     = len(_pares_con_stock & _pares_contados)
total_cob_pct = round(_pares_cubiertos / max(1, len(_pares_con_stock)) * 100, 1)

pen_merged = inv_real_alm.merge(counted_30, on=['Código','alm'], how='left', indicator=True)
pendientes_full = pen_merged[pen_merged['_merge'] == 'left_only'][
    ['Código','desc','alm','En stock']].copy()
pendientes_full['En stock'] = pendientes_full['En stock'].astype(int)
pendientes_full = pendientes_full.sort_values(['alm','Código'])
n_pendientes    = len(pendientes_full)
pendientes_list = pendientes_full.to_dict(orient='records')
ok(f"Pendientes de conteo: {n_pendientes}")

# ── CAPACIDAD ─────────────────────────────────────────────────────────────────
step("Calculando capacidad y ocupación")

inv_pallets = inv_by_alm[inv_by_alm['Almacen_code'].isin(ALMACENES_OPERATIVOS)].copy()
inv_pallets.columns = ['Código', 'alm', 'En stock']
inv_pallets['qty_pallet'] = inv_pallets['Código'].map(pallet_map).fillna(500)
inv_pallets['pal_used']   = inv_pallets['En stock'] / inv_pallets['qty_pallet']

cap_df = inv_pallets.groupby('alm')['pal_used'].sum().reset_index()
cap_df.columns = ['alm','pallets_used']
cap_df['pallets_used'] = cap_df['pallets_used'].round(1)
cap_df['cap']      = cap_df['alm'].map(CAP_NEW).fillna(50)
cap_df['pct_ocup'] = (cap_df['pallets_used'] / cap_df['cap'] * 100).clip(upper=150).round(1)

n_skus_map = inv_pos.groupby('alm')['Código'].nunique()
cap_df['n_skus'] = cap_df['alm'].map(n_skus_map).fillna(0).astype(int)
cap_df = cap_df.sort_values('pct_ocup', ascending=False)

total_cap_pct = round(cap_df['pallets_used'].sum() / cap_df['cap'].sum() * 100, 1)
ok(f"Capacidad global: {total_cap_pct}%")

step("Generando Excel detalle capacidad por ítem")
try:
    cap_detail = inv_pallets[['Código','alm','En stock','qty_pallet','pal_used']].copy()
    cap_detail['desc'] = cap_detail['Código'].map(desc_map).fillna('N/D')
    cap_detail.columns = ['Codigo','Almacen','Stock','Qty_x_Pallet','Tarimas','Descripcion']
    cap_detail = cap_detail[['Codigo','Descripcion','Almacen','Stock','Qty_x_Pallet','Tarimas']]
    cap_detail = cap_detail.sort_values(['Almacen','Tarimas'], ascending=[True, False])
    cap_detail['Stock']      = cap_detail['Stock'].astype(int)
    cap_detail['Tarimas']    = cap_detail['Tarimas'].round(2)

    excel_cap_path = REPORTES_DIR / 'Detalle_Capacidad_BP.xlsx'
    with pd.ExcelWriter(str(excel_cap_path), engine='openpyxl') as writer:
        cap_detail.to_excel(writer, sheet_name='Detalle_Items', index=False)
        ws1 = writer.sheets['Detalle_Items']
        for col_idx, col_w in enumerate([12,40,10,12,14,10], start=1):
            ws1.column_dimensions[
                __import__('openpyxl').utils.get_column_letter(col_idx)
            ].width = col_w
        cap_df_exp = cap_df[['alm','pallets_used','cap','pct_ocup','n_skus']].copy()
        cap_df_exp.columns = ['Almacen','Tarimas_Usadas','Capacidad','Pct_Ocupacion','N_SKUs']
        cap_df_exp.to_excel(writer, sheet_name='Resumen_Almacen', index=False)
    ok(f"Excel capacidad: {excel_cap_path.name} ({len(cap_detail)} ítems)")
    cap_detail_list = cap_detail.to_dict(orient='records')
except Exception as e:
    warn(f"Error generando Excel capacidad: {e}")
    cap_detail_list = []

# ── Generar/actualizar Capacidad_Almacenes_BP.xlsx ──────────────────────────
# Este archivo es la FUENTE para que Almacén actualice la capacidad de cada almacén.
# Si ya existe, preserva los valores de "Nueva Capacidad" que alguien haya llenado.
# Columnas: Almacen | Capacidad Actual | Nueva Capacidad | Pct Ocupación | Tarimas Usadas | N SKUs
step("Generando/actualizando Capacidad_Almacenes_BP.xlsx")
try:
    import openpyxl as _oxl_cap
    from openpyxl.styles import (Font as _CF, PatternFill as _CPF,
                                  Alignment as _CA, Border as _CB, Side as _CS)

    _cap_file = REPORTES_DIR / 'Capacidad_Almacenes_BP.xlsx'

    # Leer valores de "Nueva Capacidad" ya guardados para preservarlos
    _prev_nueva_cap = {}
    if _cap_file.exists():
        try:
            _wb_prev = _oxl_cap.load_workbook(str(_cap_file), read_only=True, data_only=True)
            if 'Capacidad' in _wb_prev.sheetnames:
                _prev_rows = list(_wb_prev['Capacidad'].values)
                for _pr in _prev_rows[1:]:
                    if _pr and _pr[0] and _pr[2] is not None and str(_pr[2]).strip() not in ('','None','nan'):
                        _prev_nueva_cap[str(_pr[0]).strip()] = _pr[2]
            _wb_prev.close()
        except Exception:
            pass

    _wb_cap2 = _oxl_cap.Workbook()
    _wb_cap2.remove(_wb_cap2.active)
    _ws_cap2 = _wb_cap2.create_sheet('Capacidad')
    _ws_cap2.sheet_view.showGridLines = False
    _ws_cap2.tab_color = '0EA5E9'

    # Encabezados
    _cap_hdrs = ['Almacen','Capacidad Actual','Nueva Capacidad','% Ocupación','Tarimas Usadas','N SKUs']
    _cap_widths = [14, 18, 18, 14, 16, 10]
    _s = _CS(style='thin', color='93C5FD')
    _border = _CB(left=_s, right=_s, top=_s, bottom=_s)
    _ws_cap2.row_dimensions[1].height = 26
    for _ci, (_h, _w) in enumerate(zip(_cap_hdrs, _cap_widths), 1):
        _c = _ws_cap2.cell(row=1, column=_ci, value=_h)
        _c.font      = _CF(name='Arial', bold=True, size=10, color='FFFFFF')
        _c.fill      = _CPF('solid', fgColor='1E3A5F')
        _c.alignment = _CA(horizontal='center', vertical='center')
        _c.border    = _border
        _ws_cap2.column_dimensions[_oxl_cap.utils.get_column_letter(_ci)].width = _w

    # Nota instructiva en fila 2
    _ws_cap2.merge_cells('A2:F2')
    _nota = _ws_cap2['A2']
    _nota.value = 'ℹ  Llena la columna "Nueva Capacidad" para actualizar la capacidad de un almacén. El dashboard la tomará en el siguiente run del actualizador.'
    _nota.font  = _CF(name='Arial', italic=True, size=8, color='64748B')
    _nota.fill  = _CPF('solid', fgColor='F8FAFC')
    _nota.alignment = _CA(horizontal='left', vertical='center')
    _ws_cap2.row_dimensions[2].height = 18

    # Datos de cada almacén
    _cap_dict_rows = cap_df.to_dict(orient='records')
    for _ri, _row in enumerate(_cap_dict_rows, 3):
        _alm    = str(_row.get('alm',''))
        _cap_ac = int(_row.get('cap', 0))
        _nueva  = _prev_nueva_cap.get(_alm, '')   # preservar valor previo
        _pct    = round(float(_row.get('pct_ocup', 0)), 1)
        _tar    = round(float(_row.get('pallets_used', 0)), 1)
        _nskus  = int(_row.get('n_skus', 0))

        _vals = [_alm, _cap_ac, _nueva, _pct, _tar, _nskus]
        for _ci, _v in enumerate(_vals, 1):
            _c = _ws_cap2.cell(row=_ri, column=_ci, value=_v)
            _c.font      = _CF(name='Arial', size=10)
            _c.alignment = _CA(horizontal='center', vertical='center')
            _c.border    = _border
            bg = 'EFF6FF' if _ri % 2 == 0 else 'FFFFFF'
            _c.fill = _CPF('solid', fgColor=bg)

        # Col C (Nueva Capacidad) — resaltar en amarillo para indicar que es editable
        _c_nueva = _ws_cap2.cell(row=_ri, column=3)
        _c_nueva.fill = _CPF('solid', fgColor='FEFCE8')
        _c_nueva.font = _CF(name='Arial', size=10, color='854D0E', bold=True)

    _ws_cap2.freeze_panes = 'A3'
    _wb_cap2.save(str(_cap_file))
    ok(f"Capacidad_Almacenes_BP.xlsx: {len(_cap_dict_rows)} almacenes")
except Exception as _ecap:
    warn(f"Error generando Capacidad_Almacenes_BP.xlsx: {_ecap}")

step("Extrayendo inventario MATE2122")
df_mate2122 = df_inv[
    df_inv[c_inv_alm].astype(str).str.strip() == 'MATE2122'
].copy() if c_inv_alm else pd.DataFrame()

mate2122_list = []
if len(df_mate2122) > 0:
    df_mate2122['_Cod'] = df_mate2122[c_inv_cod].astype(str).str.strip()
    df_mate2122['_Stk'] = safe_numeric(df_mate2122[c_inv_stk])
    df_mate2122 = df_mate2122[df_mate2122['_Stk'] > 0].copy()
    df_mate2122['desc']      = df_mate2122['_Cod'].map(desc_map).fillna('N/D')
    df_mate2122['costo_u']   = df_mate2122['_Cod'].map(cos_map).fillna(0)
    df_mate2122['monto_mxn'] = (df_mate2122['_Stk'] * df_mate2122['costo_u']).round(0)
    df_mate2122['qty_pal']   = df_mate2122['_Cod'].map(pallet_map).fillna(500)
    df_mate2122['tarimas']   = (df_mate2122['_Stk'] / df_mate2122['qty_pal']).round(1)
    df_mate2122['prefix']    = df_mate2122['_Cod'].str[:2]
    df_mate2122 = df_mate2122.sort_values('monto_mxn', ascending=False)
    mate2122_list = df_mate2122[
        ['_Cod','desc','prefix','_Stk','monto_mxn','tarimas','costo_u']
    ].rename(columns={
        '_Cod':'Codigo','_Stk':'Stock','costo_u':'Costo_unit'
    }).to_dict(orient='records')
    total_mate2122_mxn = round(float(df_mate2122['monto_mxn'].sum()))
    total_mate2122_pzs = int(df_mate2122['_Stk'].sum())
    ok(f"MATE2122: {len(mate2122_list)} SKUs | {total_mate2122_pzs:,} pzs | ${total_mate2122_mxn:,.0f} MXN")
else:
    total_mate2122_mxn = 0
    total_mate2122_pzs = 0
    warn("Almacén MATE2122 no encontrado o sin stock")

step("Calculando costo de inventario")
inv_total_desc = inv_total_by_sku.copy()
inv_total_desc['costo_unit'] = inv_total_desc['Código'].map(cos_map).fillna(0)
inv_total_desc['valor']      = (inv_total_desc['En stock'] * inv_total_desc['costo_unit']).round(2)

_ica = inv_by_alm[inv_by_alm['Almacen_code'].isin(ALMACENES_OPERATIVOS)].copy()
_ica = _ica.rename(columns={'Almacen_code': 'alm', 'Stock': 'En stock', 'Codigo': 'Código'})
_ica['costo_unit'] = _ica['Código'].map(cos_map).fillna(0)
_ica['valor']      = (_ica['En stock'] * _ica['costo_unit']).round(2)
costo_alm = _ica.groupby('alm')['valor'].sum().reset_index()
costo_alm.columns = ['alm','valor_total']
costo_alm = costo_alm[costo_alm['valor_total'] > 0].sort_values('valor_total', ascending=False)

total_costo = float(inv_total_desc['valor'].sum())
sku_cost    = inv_total_desc[inv_total_desc['valor'] > 0].copy()
sku_cost['desc'] = sku_cost['Código'].map(desc_map).fillna('N/D')
sku_cost    = sku_cost.nlargest(20, 'valor')
ok(f"Costo total inventario: ${total_costo:,.0f} MXN")

_inv_nac = inv_by_alm[
    inv_by_alm['Almacen_code'].str.startswith('M') &
    ~inv_by_alm['Almacen_code'].isin(ALM_EXCLUIR_NAC)
].copy()
_inv_nac = _inv_nac.rename(columns={'Almacen_code': 'Almacen', 'Stock': 'Stock', 'Codigo': 'Codigo'})
_inv_nac['Descripcion'] = _inv_nac['Codigo'].map(desc_map).fillna('N/D')
_inv_nac['Costo_unit']  = _inv_nac['Codigo'].map(cos_map).fillna(0)
_inv_nac['Monto_MXN']   = (_inv_nac['Stock'] * _inv_nac['Costo_unit']).round(0).astype(int)
_inv_nac['Stock']       = _inv_nac['Stock'].astype(int)
_inv_nac = _inv_nac[['Codigo','Descripcion','Almacen','Stock','Costo_unit','Monto_MXN']]
_inv_nac = _inv_nac[_inv_nac['Stock'] > 0].sort_values(['Almacen','Codigo'])
inv_nac_list = _inv_nac.to_dict(orient='records')
ok(f"Inventario Nac M*: {len(inv_nac_list)} registros en {_inv_nac['Almacen'].nunique()} almacenes")

step("Procesando Histórico de Movimientos")
c_hart = col(df_hist, 'Número de artículo', 'Numero de articulo', 'Artículo', 'Articulo', 'SKU', 'Código', 'Codigo')
c_hfec = col(df_hist, 'Fecha del sistema', 'Fecha Sistema', 'Fecha', 'FechaSistema')
c_hval = col(df_hist, 'Valor trans.', 'Valor Transaccion', 'Valor', 'ValorTrans', 'Valor transacción')
c_hcan = col(df_hist, 'Cantidad')

df_hist['art']       = df_hist[c_hart].astype(str).str.strip()
df_hist['_Alm']      = df_hist[c_hist_alm].astype(str).str.strip() if c_hist_alm else ''
df_hist['fecha_sys'] = df_hist[c_hfec].apply(parse_hist_date)
df_hist['Cantidad']  = safe_numeric(df_hist[c_hcan])
df_hist['Valor trans.'] = safe_numeric(df_hist[c_hval])

df_v = df_hist.dropna(subset=['art','_Alm','fecha_sys']).copy()
df_v = df_v[(df_v['art'].str.len() == 6) & (df_v['_Alm'].isin(M_ALMS))]

df_v['t2']         = df_v['Cantidad'].apply(lambda x: 'E' if x > 0 else 'S')
df_v['Mes']        = df_v['fecha_sys'].dt.to_period('M').astype(str)
df_v['costo']      = df_v['art'].map(cos_map).fillna(0)
df_v['valor_calc'] = (df_v['Cantidad'].abs() * df_v['costo']).round(2)
df_v['fecha_str']  = df_v['fecha_sys'].dt.strftime('%Y-%m-%d')
df_v['mxn_e_col']  = df_v['valor_calc'].where(df_v['t2'] == 'E', 0.0)
df_v['mxn_s_col']  = df_v['valor_calc'].where(df_v['t2'] == 'S', 0.0)

movs_by_alm_date = df_v.groupby(['fecha_str','_Alm']).size().reset_index(name='count')
movs_by_alm_date.columns = ['fecha_str','Almacén','count']
movs_by_alm_date = movs_by_alm_date.sort_values('fecha_str')
hist_date_min = str(df_v['fecha_str'].min())
hist_date_max = str(df_v['fecha_str'].max())
ok(f"Movimientos válidos: {len(df_v):,} | Rango: {hist_date_min} -> {hist_date_max}")

df_sal_all = df_v[df_v['t2'] == 'S'].copy()
sal_by_date_alm = (
    df_sal_all.groupby(['fecha_str', '_Alm'])
    .agg(
        cantidad=('Cantidad', lambda x: round(abs(x.sum()))),
        valor=('valor_calc', lambda x: round(x.sum()))
    ).reset_index()
)
sal_by_date_alm.columns = ['fecha', 'alm', 'cantidad', 'valor']
ok(f"Salidas agrupadas: {len(sal_by_date_alm)} fecha×alm rows")

step("Clasificación ABC")
abc_counts = df_v.groupby('art').size().reset_index(name='n_movs')
abc_counts  = abc_counts.sort_values('n_movs', ascending=False)
total_movs  = abc_counts['n_movs'].sum()
abc_counts['cum_pct'] = (abc_counts['n_movs'].cumsum() / total_movs * 100).round(2)
abc_counts['clase']   = abc_counts['cum_pct'].apply(
    lambda x: 'A' if x <= 80 else ('B' if x <= 95 else 'C'))
abc_counts['desc']    = abc_counts['art'].map(desc_map).fillna('N/D')

stk_m = inv_total_by_sku.copy()
stk_m.columns = ['art','stock']
abc_counts = abc_counts.merge(stk_m, on='art', how='left')
abc_counts['stock']      = abc_counts['stock'].fillna(0).astype(int)
abc_counts['costo_unit'] = abc_counts['art'].map(cos_map).fillna(0)
abc_counts['monto_mxn']  = (abc_counts['stock'] * abc_counts['costo_unit']).round(0)
abc_counts['qty_pallet'] = abc_counts['art'].map(pallet_map).fillna(500)
abc_counts['n_tarimas']  = (abc_counts['stock'] / abc_counts['qty_pallet']).round(1)

abc_summary = abc_counts.groupby('clase').agg(
    n_skus   =('art',      'count'),
    n_movs   =('n_movs',   'sum'),
    monto_mxn=('monto_mxn','sum'),
    n_tarimas=('n_tarimas','sum')
).reset_index()
abc_summary['pct_skus'] = (abc_summary['n_skus'] / len(abc_counts) * 100).round(1)
abc_summary['pct_movs'] = (abc_summary['n_movs'] / total_movs * 100).round(1)
abc_summary['monto_mxn'] = abc_summary['monto_mxn'].round(0)
abc_summary['n_tarimas']  = abc_summary['n_tarimas'].round(1)

# FIX #2: Agregar almacenes donde cada SKU tiene stock (para export Excel)
_abc_alms = (
    inv_by_alm[inv_by_alm['Almacen_code'].isin(ALMACENES_NAC) & (inv_by_alm['Stock'] > 0)]
    .groupby('Codigo')['Almacen_code']
    .apply(lambda x: ', '.join(sorted(x.astype(str).unique())))
    .reset_index()
)
_abc_alms.columns = ['art', 'alms']
abc_counts = abc_counts.merge(_abc_alms, on='art', how='left')
abc_counts['alms'] = abc_counts['alms'].fillna('—')

abc_full = abc_counts[['art','desc','n_movs','cum_pct','clase','stock','monto_mxn','n_tarimas','alms']].copy()

hist_skus   = set(abc_counts['art'].unique())
no_mov_skus = set(inv_total_by_sku['Código'].unique()) - hist_skus
no_mov_list = []
for sku in sorted(no_mov_skus):
    stock = int(inv_total_by_sku[inv_total_by_sku['Código'] == sku]['En stock'].sum())
    cu    = cos_map.get(sku, 0)
    qp    = pallet_map.get(sku, 500)
    alms_n = int(inv_pos[inv_pos['Código'] == sku]['alm'].nunique())
    alms_str = ', '.join(sorted(inv_pos[inv_pos['Código'] == sku]['alm'].tolist()))
    no_mov_list.append({
        'art': sku, 'desc': desc_map.get(sku, 'N/D'),
        'stock': stock, 'n_alm': alms_n, 'alms': alms_str,
        'status': 'Transferible' if alms_n > 1 else 'Obsoleto',
        'monto_mxn': round(stock * cu, 0),
        'n_tarimas': round(stock / qp, 1)
    })
ok(f"ABC: {len(abc_full)} SKUs activos | Sin movimiento: {len(no_mov_list)}")

step("Top 10 por tipo de producto")
df_v['tipo'] = df_v['art'].apply(
    lambda s: next((p for p in PREFIJOS_TIPO if s.startswith(p)), None))

top10 = {}
for tipo in PREFIJOS_TIPO:
    sub = df_v[df_v['tipo'] == tipo].copy()
    if sub.empty:
        top10[tipo] = []
        continue
    cnt = sub.groupby('art').agg(
        movs    =('art',       'count'),
        entradas=('Cantidad',  lambda x: float(x[x > 0].sum())),
        salidas =('Cantidad',  lambda x: float(x[x < 0].abs().sum())),
        mxn_e   =('mxn_e_col','sum'),
        mxn_s   =('mxn_s_col','sum'),
    ).reset_index()
    cnt['desc'] = cnt['art'].map(desc_map).fillna('N/D')
    stk = (
        inv_total_by_sku[inv_total_by_sku['Código'].str.startswith(tipo)]
        .rename(columns={'Código':'art','En stock':'stock'})
    )
    cnt = cnt.merge(stk, on='art', how='left').fillna(0)
    cnt = cnt.nlargest(10, 'movs')
    top10[tipo] = cnt[['art','desc','movs','entradas','salidas','mxn_e','mxn_s','stock']
                     ].round(0).to_dict(orient='records')
ok(f"Top10 calculado para {len(top10)} tipos")

step("Tendencia mensual")
# Calcular MESES_VALIDOS dinámicamente: todos los meses con datos,
# con corte al mes anterior al corriente (mes actual puede estar incompleto)
_hoy = TODAY
_mes_corte = (datetime.date(_hoy.year, _hoy.month, 1) - datetime.timedelta(days=1))
_mes_corte_str = _mes_corte.strftime('%Y-%m')
_meses_datos = sorted(df_v['Mes'].dropna().unique())
MESES_VALIDOS = [m for m in _meses_datos if m <= _mes_corte_str]
if not MESES_VALIDOS:
    MESES_VALIDOS = _meses_datos[-6:] if len(_meses_datos) >= 6 else list(_meses_datos)
ok(f"Meses disponibles: {MESES_VALIDOS}")
df_v_valid  = df_v[df_v['Mes'].isin(MESES_VALIDOS)]
alms_hist   = sorted(df_v_valid['_Alm'].unique())
trend_pivot = {}
for alm in alms_hist + ['TODOS']:
    sub_alm = (df_v_valid[df_v_valid['_Alm'] == alm]
               if alm != 'TODOS' else df_v_valid)
    e_vals, s_vals = [], []
    for mes in MESES_VALIDOS:
        sub_m = sub_alm[sub_alm['Mes'] == mes]
        e_vals.append(round(float(sub_m[sub_m['t2'] == 'E']['valor_calc'].sum())))
        s_vals.append(round(float(sub_m[sub_m['t2'] == 'S']['valor_calc'].sum())))
    trend_pivot[alm] = {'E': e_vals, 'S': s_vals}
ok(f"Tendencia: {len(MESES_VALIDOS)} meses | {len(alms_hist)} almacenes")

print(f"\n  OK Módulo Inventarios completo")

# ══════════════════════════════════════════════════════════════════════════════
#  MÓDULO 2 — MRP (desde Plan + Inventario + GLM + Historico Movs)
# ══════════════════════════════════════════════════════════════════════════════
step("Módulo MRP — construyendo desde Plan + Inventario + GLM + Historico Movs")

try:
    cutoff_8w = TODAY - pd.Timedelta(weeks=SEMANAS_DEMANDA)
    df_sal_mrp = df_v[(df_v['fecha_sys'] >= cutoff_8w) & (df_v['fecha_sys'] <= TODAY) & (df_v['Cantidad'] < 0)].copy()

    dem_wk_nac = (
        df_sal_mrp.groupby('art')['Cantidad']
        .apply(lambda x: abs(x.sum()) / SEMANAS_DEMANDA)
        .to_dict()
    )

    cutoff_2w = TODAY - pd.Timedelta(weeks=2)
    df_sal_2w = df_v[(df_v['fecha_sys'] >= cutoff_2w) & (df_v['fecha_sys'] <= TODAY) & (df_v['Cantidad'] < 0)].copy()
    sal_2wk_by_alm = (
        df_sal_2w.groupby(['art', '_Alm'])['Cantidad']
        .apply(lambda x: abs(x.sum()))
        .to_dict()
    )
    sal_2wk_nac = (
        df_sal_2w.groupby('art')['Cantidad']
        .apply(lambda x: abs(x.sum()))
        .to_dict()
    )

    _days_to_sat = (TODAY.weekday() - 5) % 7
    _cutoff_sat  = TODAY - pd.Timedelta(days=max(0, _days_to_sat))
    df_sal_sat   = df_v[
        (df_v['fecha_sys'] >= _cutoff_sat) &
        (df_v['fecha_sys'] <= TODAY) &
        (df_v['Cantidad'] < 0)
    ].copy()
    sal_sat_by_alm = (
        df_sal_sat.groupby(['art', '_Alm'])['Cantidad']
        .apply(lambda x: abs(x.sum()))
        .to_dict()
    )
    sal_sat_nac = (
        df_sal_sat.groupby('art')['Cantidad']
        .apply(lambda x: abs(x.sum()))
        .to_dict()
    )

    ok(f"Demanda semanal nacional: {len(dem_wk_nac)} SKUs | Salidas 2w: {len(sal_2wk_nac)} SKUs")

    glm_map = defaultdict(float)
    glm_prov = []
    if df_glm_inv is not None and len(df_glm_inv) > 0:
        c_glm_part = col(df_glm_inv, '# de Parte BP', 'Parte BP', 'Codigo', 'SKU',
                         'Número de artículo', 'Numero de articulo')
        c_glm_disp = col(df_glm_inv, 'Qty disponible', 'Qty Total', 'Disponible',
                         'Qty Disponible')
        for _, r in df_glm_inv.iterrows():
            k = str(r[c_glm_part]).strip()
            v = float(safe_numeric(pd.Series([r[c_glm_disp]]))[0])
            glm_map[k] += v
            row_d = {}
            for col_name, val in r.items():
                if pd.isna(val):
                    row_d[col_name] = None
                elif isinstance(val, (int, float, np.integer, np.floating)):
                    row_d[col_name] = float(val)
                else:
                    row_d[col_name] = str(val)
            glm_prov.append(row_d)
        ok(f"GLM mapa: {sum(1 for v in glm_map.values() if v > 0)} SKUs con stock "
           f"| {len(glm_prov)} registros totales")
    else:
        warn("Hoja GLM no encontrada o vacía.")

    inv_nac_dict = inv_total_by_sku.set_index('Código')['En stock'].to_dict()

    if df_plan is None or len(df_plan) == 0:
        raise ValueError("Hoja 'Plan' no encontrada o vacía.")

    c_plan_cod = col(df_plan, 'Codigo', 'Código', 'SKU', 'Número de artículo')
    c_plan_alm = col(df_plan, 'Almacen', 'Almacén')
    c_plan_min = col(df_plan, 'Min', 'Minimo', 'Mínimo')
    c_plan_max = col(df_plan, 'Max', 'Maximo', 'Máximo')
    c_plan_des = None
    for cand in ['Descripcion', 'Descripción', 'Desc', 'Nombre']:
        try:
            c_plan_des = col(df_plan, cand)
            break
        except ValueError:
            pass

    inv_alm_lookup = (
        inv_by_alm.set_index(['Codigo', 'Almacen_code'])['Stock']
        .to_dict()
    )

    plan_raw = []
    for _, r in df_plan.iterrows():
        codigo  = str(r[c_plan_cod]).strip()
        almacen = str(r[c_plan_alm]).strip()
        almacen = ALM_SAP_TO_INTERNO.get(almacen, almacen)  # normalizar SAP→interno
        if not codigo or codigo in ('nan', 'None') or len(codigo) != 6:
            continue
        p_min = float(safe_numeric(pd.Series([r[c_plan_min]]))[0])
        p_max = float(safe_numeric(pd.Series([r[c_plan_max]]))[0])
        stock_local = float(inv_alm_lookup.get((codigo, almacen), 0.0))
        plan_raw.append({
            'Codigo': codigo, 'Almacen': almacen,
            'p_min': p_min, 'p_max': p_max, 'stock_local': stock_local,
        })

    plan_by_sku = defaultdict(list)
    for rec in plan_raw:
        plan_by_sku[rec['Codigo']].append(rec)

    mrp_plan = []

    for rec in plan_raw:
        codigo  = rec['Codigo']
        almacen = rec['Almacen']
        p_min   = rec['p_min']
        p_max   = rec['p_max']
        stock_local_raw = rec['stock_local']

        sal_desde_sat_local = float(sal_sat_by_alm.get((codigo, almacen), 0.0))
        sal_desde_sat_nac   = float(sal_sat_nac.get(codigo, 0.0))
        stock_local = stock_local_raw

        stock_nac_raw = float(inv_nac_dict.get(codigo, 0.0))
        stock_nac = stock_nac_raw

        p_min_effective = max(0.0, p_min - sal_desde_sat_local)

        pct_inv = round(stock_local / stock_nac * 100, 1) if stock_nac > 0 else 0.0

        dw_nac = float(dem_wk_nac.get(codigo, 0.0))

        dem_wk_plan = p_min
        sems_local = round(stock_local / dem_wk_plan, 2) if dem_wk_plan > 0 else (
            99.0 if stock_local > 0 else 0.0)

        sems_nac = round(stock_nac / dw_nac, 2) if dw_nac > 0 else (
            99.0 if stock_nac > 0 else 0.0)

        qty_2wk = round(p_max)

        delta_local = round(stock_local - p_max)

        deficit = round(max(0.0, p_min_effective - stock_local))
        exceso_qty = round(max(0.0, stock_local - p_max)) if p_max > 0 else 0

        pct_cob = round(stock_local / p_min * 100, 1) if p_min > 0 else (
            100.0 if stock_local > 0 else 0.0)

        sal_2wk = round(float(sal_2wk_by_alm.get((codigo, almacen), 0.0)))
        sal_2wk_total = round(float(sal_2wk_nac.get(codigo, 0.0)))
        delta_pend = round(max(0.0, qty_2wk - sal_2wk))

        inv_glm = float(glm_map.get(codigo, 0.0))

        zona_dest = ZONA_MAP.get(almacen, 'Otra')
        otros = [o for o in plan_by_sku[codigo]
                 if o['Almacen'] != almacen and o['Almacen'] not in EXCLUIR_TRASLADO]
        otros_surplus = []
        for o in otros:
            surplus = max(0.0, o['stock_local'] - o['p_min'])
            zona_src = ZONA_MAP.get(o['Almacen'], 'Otra')
            prioridad = 2.0 if zona_src == zona_dest else 1.0
            otros_surplus.append((o['Almacen'], surplus, prioridad))
        transferible = round(sum(s for _, s, _ in otros_surplus))
        best_src = max(otros_surplus, key=lambda x: x[1] * x[2], default=('', 0, 1))
        alm_fuente = best_src[0] if best_src[1] > 0 else ''

        deficit_to_max = round(max(0.0, p_max - stock_local)) if p_max > 0 else 0

        glm_urgente = (sems_nac < 4.0) and (inv_glm > 0)

        if p_max > 0 and stock_local > p_max:
            dec_int = 'Sin acción'
            dec_ext = 'Cubierto'
        elif p_max > 0 and p_min > 0 and stock_local >= p_min and stock_local < p_max:
            if glm_urgente and inv_glm >= deficit_to_max:
                dec_int = 'Sin acción'
                dec_ext = 'GLM'
            elif glm_urgente and inv_glm > 0:
                dec_int = 'Sin acción'
                dec_ext = 'GLM'
            elif transferible >= deficit_to_max:
                dec_int = 'Traslado'
                dec_ext = 'Traslado'
            elif inv_glm >= deficit_to_max:
                dec_int = 'Sin acción'
                dec_ext = 'GLM'
            elif inv_glm > 0:
                dec_int = 'Sin acción'
                dec_ext = 'GLM'
            else:
                dec_int = 'Sin acción'
                dec_ext = 'Cubierto'
        elif p_min > 0 and stock_local >= p_min:
            dec_int = 'Sin acción'
            dec_ext = 'Cubierto'
        elif p_min > 0 and stock_local < p_min:
            if deficit == 0:
                dec_int = 'Sin acción'
                dec_ext = 'Cubierto'
            elif glm_urgente and inv_glm >= deficit:
                dec_int = 'Sin acción'
                dec_ext = 'GLM'
            elif glm_urgente and inv_glm > 0:
                dec_int = 'Sin acción'
                dec_ext = 'GLM'
            elif transferible >= deficit:
                dec_int = 'Traslado'
                dec_ext = 'Traslado'
            elif inv_glm >= deficit:
                dec_int = 'Sin acción'
                dec_ext = 'GLM'
            elif inv_glm > 0:
                dec_int = 'Sin acción'
                dec_ext = 'GLM'
            else:
                dec_int = 'Sin acción'
                dec_ext = 'Corto'
        else:
            if stock_local > 0:
                dec_int = 'Sin acción'
                dec_ext = 'Cubierto'
            else:
                dec_int = 'Sin acción'
                dec_ext = 'Sin stock'

        costo_u  = cos_map.get(codigo, 0.0)
        desc_val = desc_map.get(codigo, 'N/D')

        # REFILL: condición correcta usando salidas de la semana corriente (sáb-hoy)
        # Criterios:
        #   1. stock actual < mínimo  (está bajo)
        #   2. hubo salidas ESTA semana (sal_desde_sat_local > 0)
        #   3. al inicio de la semana sí cubría el mínimo (stock + salidas_semana >= mínimo)
        # → Si hoy es sábado y aún no hay salidas, no es Refill aunque esté bajo el mínimo
        # → La decisión (Corto/Traslado/GLM) se calcula normal pero NO cuenta en summary
        is_refill = bool(
            stock_local < p_min and
            sal_desde_sat_local > 0 and
            (stock_local + sal_desde_sat_local) >= p_min
        )

        mrp_plan.append({
            'Codigo':       codigo,
            'Descripcion':  desc_val,
            'Almacen':      almacen,
            'Stock':        round(stock_local),
            'Inv_nac':      round(stock_nac),
            'Pct_inv':      pct_inv,
            'Min':          round(p_min),
            'Max':          round(p_max),
            'Qty_2wk':      qty_2wk,
            'Deficit':      deficit,
            'Exceso_qty':   exceso_qty,
            'Delta_local':  delta_local,
            'Sems_local':   sems_local,
            'Sems_nac':     sems_nac,
            'Pct_cob':      pct_cob,
            'Dem_wk_plan':  round(dem_wk_plan, 1),
            'Dem_wk_hist':  round(dw_nac, 1),
            'Sal_2wk':      sal_2wk,
            'Sal_2wk_nac':  sal_2wk_total,
            'Delta_pend':   delta_pend,
            'Inv_glm':      round(inv_glm),
            'Transferible': transferible,
            'Alm_fuente':   alm_fuente,
            'Deficit_to_max': deficit_to_max,
            'Zona':         zona_dest,
            'Dec_interna':  dec_int,
            'Dec_externa':  dec_ext,
            'Decision':     dec_ext,
            'Costo_unit':   costo_u,
            'Valor_stock':  round(stock_local * costo_u),
            'Valor_deficit':round(deficit * costo_u),
            'Is_refill':    is_refill,
        })

    ok(f"Plan procesado: {len(mrp_plan)} registros")

    # ══════════════════════════════════════════════════════════════════════════
    # CAMBIOS DE PLAN — leer Cambios_Plan_BP.xlsx (capturado por Planeación)
    # Lee la hoja "Cambios Plan" (header en fila 3, datos desde fila 4).
    # Convierte Lugar de Cosecha → Almacén SAP usando hoja "Conversion".
    # Lee TODAS las filas con Embalaje + Lugar + Cajas válidos (sin filtro de estado).
    # Marca los registros del mrp_plan con Is_cambio_plan=True.
    # ══════════════════════════════════════════════════════════════════════════
    cambios_plan_list = []
    _cambios_aplicados = 0

    if CAMBIOS_PLAN_FILE.exists():
        try:
            _df_cp = pd.read_excel(str(CAMBIOS_PLAN_FILE), sheet_name='Cambios Plan',
                                   header=2, engine='openpyxl')
            _df_cv = pd.read_excel(str(CAMBIOS_PLAN_FILE), sheet_name='Conversion',
                                   header=0, engine='openpyxl')

            # Mapa Lugar de Cosecha → Almacén SAP desde hoja Conversion
            _lugar_col = next((c for c in _df_cv.columns if 'nombre' in str(c).lower() or 'lugar' in str(c).lower()), None)
            _alm_col   = next((c for c in _df_cv.columns if 'almac' in str(c).lower() or 'sap' in str(c).lower()), None)
            _lugar_to_alm = {}
            if _lugar_col and _alm_col:
                _lugar_to_alm = dict(zip(
                    _df_cv[_lugar_col].astype(str).str.strip(),
                    _df_cv[_alm_col].astype(str).str.strip()
                ))

            # Detectar columnas en "Cambios Plan"
            _col_emb = next((c for c in _df_cp.columns if 'embalaje' in str(c).lower() or 'c\u00f3digo' in str(c).lower() or 'codigo' in str(c).lower()), None)
            _col_lug = next((c for c in _df_cp.columns if 'lugar' in str(c).lower() or 'cosecha' in str(c).lower()), None)
            _col_caj = next((c for c in _df_cp.columns if 'caja' in str(c).lower() or 'total' in str(c).lower()), None)
            _col_cli = next((c for c in _df_cp.columns if 'cliente' in str(c).lower()), None)
            _col_com = next((c for c in _df_cp.columns if 'comentario' in str(c).lower()), None)
            _col_freq= next((c for c in _df_cp.columns if 'requer' in str(c).lower()), None)
            _col_falm= next((c for c in _df_cp.columns if 'alm' in str(c).lower() and 'fecha' in str(c).lower()), None)

            # ── Mapa PT → {BX, qty_bx, CL, qty_cl} usando hoja Embalajes de Referencia_Inventario_BP ──
            # Columnas: Codigo PT | Codigo Clam | Cantidad Clam | Codigo Caja | Cantidad Caja2
            # Un PT = 1 caja BX × qty_bx (siempre 1) + qty_clam × clamshells CL/CV
            _pt_map = {}   # pt_code → {'bx': 'BX0005', 'qty_bx': 1, 'cl': 'CL0263', 'qty_cl': 12}
            try:
                _df_emb = load_sheet('Embalajes', required=False)
                if _df_emb is not None:
                    _c_pt   = next((c for c in _df_emb.columns if str(c).strip().lower() == 'codigo pt'), None)
                    _c_cl   = next((c for c in _df_emb.columns if 'clam' in str(c).lower() and 'codigo' in str(c).lower()), None)
                    _c_qcl  = next((c for c in _df_emb.columns if 'cantidad' in str(c).lower() and 'clam' in str(c).lower()), None)
                    _c_bx   = next((c for c in _df_emb.columns if 'caja' in str(c).lower() and 'codigo' in str(c).lower()), None)
                    _c_qbx  = next((c for c in _df_emb.columns if 'cantidad' in str(c).lower() and 'caja' in str(c).lower()), None)
                    if _c_pt and _c_bx:
                        for _, er in _df_emb.iterrows():
                            pt_val  = str(er[_c_pt] or '').strip()
                            bx_val  = str(er[_c_bx] or '').strip()
                            cl_val  = str(er[_c_cl]  or '').strip() if _c_cl  else ''
                            try:    qty_cl = int(float(er[_c_qcl])) if _c_qcl and er[_c_qcl] else 0
                            except: qty_cl = 0
                            try:    qty_bx = int(float(er[_c_qbx])) if _c_qbx and er[_c_qbx] else 1
                            except: qty_bx = 1
                            if pt_val and pt_val not in ('nan','None') and bx_val and bx_val not in ('nan','None'):
                                _pt_map[pt_val] = {
                                    'bx':     bx_val,
                                    'qty_bx': qty_bx if qty_bx > 0 else 1,
                                    'cl':     cl_val if cl_val not in ('nan','None','') else '',
                                    'qty_cl': qty_cl,
                                }
                ok(f"Embalajes PT map: {len(_pt_map)} códigos PT cargados")
            except Exception as _ee:
                warn(f"No se pudo cargar mapa Embalajes: {_ee}")

            # ── Helper: recalcular Dec_interna / Dec_externa para una fila de mrp_plan ──
            def _recalc_decision(row, plan_index):
                """Recalcula deficit, sems y decisión de una fila ya en mrp_plan."""
                cod = row['Codigo']; alm = row['Almacen']
                stk  = row['Stock'];  p_min = row['Min']; p_max = row['Max']
                dw   = row.get('Dem_wk_hist', row.get('Dem_wk_plan', 0)) or 0
                sal_sat = float(sal_sat_by_alm.get((cod, alm), 0.0))
                p_min_eff = max(0.0, p_min - sal_sat)
                deficit  = round(max(0.0, p_min_eff - stk))
                exceso   = round(max(0.0, stk - p_max)) if p_max > 0 else 0
                sems_l   = round(stk / p_min, 2) if p_min > 0 else (99.0 if stk > 0 else 0.0)
                stk_nac  = float(inv_nac_dict.get(cod, 0.0))
                dw_nac   = float(dem_wk_nac.get(cod, 0.0))
                sems_n   = round(stk_nac / dw_nac, 2) if dw_nac > 0 else (99.0 if stk_nac > 0 else 0.0)
                inv_glm  = float(glm_map.get(cod, 0.0))
                # transferible: excedente en otros almacenes del mismo SKU
                otros = [(r['Almacen'], max(0.0, r['Stock'] - r['Min']))
                         for r in plan_index.get(cod, [])
                         if r['Almacen'] != alm and r['Almacen'] not in EXCLUIR_TRASLADO]
                transferible = round(sum(s for _, s in otros))
                alm_fuente   = max(otros, key=lambda x: x[1], default=('', 0))[0] if otros else ''
                deficit_to_max = round(max(0.0, p_max - stk)) if p_max > 0 else 0
                glm_urgente  = (sems_n < 4.0) and (inv_glm > 0)

                if p_max > 0 and stk > p_max:
                    di, de = 'Sin acción', 'Cubierto'
                elif p_max > 0 and p_min > 0 and stk >= p_min:
                    if glm_urgente:            di, de = 'Sin acción', 'GLM'
                    elif transferible >= deficit_to_max: di, de = 'Traslado', 'Traslado'
                    elif inv_glm > 0:          di, de = 'Sin acción', 'GLM'
                    else:                      di, de = 'Sin acción', 'Cubierto'
                elif p_min > 0 and stk < p_min:
                    if deficit == 0:           di, de = 'Sin acción', 'Cubierto'
                    elif glm_urgente and inv_glm >= deficit: di, de = 'Sin acción', 'GLM'
                    elif glm_urgente and inv_glm > 0:        di, de = 'Sin acción', 'GLM'
                    elif transferible >= deficit:             di, de = 'Traslado', 'Traslado'
                    elif inv_glm >= deficit:   di, de = 'Sin acción', 'GLM'
                    elif inv_glm > 0:          di, de = 'Sin acción', 'GLM'
                    else:                      di, de = 'Sin acción', 'Corto'
                else:
                    di, de = ('Sin acción', 'Cubierto') if stk > 0 else ('Sin acción', 'Sin stock')

                costo_u = row.get('Costo_unit', 0.0)
                row.update({
                    'Deficit': deficit, 'Exceso_qty': exceso,
                    'Sems_local': sems_l, 'Sems_nac': sems_n,
                    'Inv_glm': round(inv_glm), 'Transferible': transferible,
                    'Alm_fuente': alm_fuente,
                    'Dec_interna': di, 'Dec_externa': de, 'Decision': de,
                    'Valor_deficit': round(deficit * costo_u),
                })

            # Índice mutable de mrp_plan para lookups rápidos y creación de filas
            # Se actualiza al agregar filas nuevas de cambio plan
            _plan_idx_mut = defaultdict(list)
            for _rr in mrp_plan:
                _plan_idx_mut[_rr['Codigo']].append(_rr)

            # ── Helper: crear fila nueva en mrp_plan para un SKU/Almacen sin plan ──
            def _nueva_fila_mrp(codigo, almacen, p_min, p_max, cliente='', pt='', comentario=''):
                stk  = float(inv_alm_lookup.get((codigo, almacen), 0.0))
                stk_nac = float(inv_nac_dict.get(codigo, 0.0))
                costo_u = cos_map.get(codigo, 0.0)
                desc_v  = desc_map.get(codigo, codigo)
                dw_nac  = float(dem_wk_nac.get(codigo, 0.0))
                pct_inv = round(stk / stk_nac * 100, 1) if stk_nac > 0 else 0.0
                sems_l  = round(stk / p_min, 2) if p_min > 0 else (99.0 if stk > 0 else 0.0)
                sems_n  = round(stk_nac / dw_nac, 2) if dw_nac > 0 else (99.0 if stk_nac > 0 else 0.0)
                sal_2wk_v = round(float(sal_2wk_by_alm.get((codigo, almacen), 0.0)))
                new_row = {
                    'Codigo': codigo, 'Descripcion': desc_v, 'Almacen': almacen,
                    'Stock': round(stk), 'Inv_nac': round(stk_nac),
                    'Pct_inv': pct_inv, 'Min': round(p_min), 'Max': round(p_max),
                    'Qty_2wk': round(p_max), 'Deficit': 0, 'Exceso_qty': 0,
                    'Delta_local': round(stk - p_max), 'Sems_local': sems_l,
                    'Sems_nac': sems_n, 'Pct_cob': round(stk / p_min * 100, 1) if p_min > 0 else 100.0,
                    'Dem_wk_plan': p_min, 'Dem_wk_hist': round(dw_nac, 1),
                    'Sal_2wk': sal_2wk_v, 'Sal_2wk_nac': round(float(sal_2wk_nac.get(codigo, 0.0))),
                    'Delta_pend': round(max(0.0, p_max - sal_2wk_v)),
                    'Inv_glm': round(float(glm_map.get(codigo, 0.0))),
                    'Transferible': 0, 'Alm_fuente': '',
                    'Deficit_to_max': round(max(0.0, p_max - stk)),
                    'Zona': ZONA_MAP.get(almacen, 'Otra'),
                    'Dec_interna': 'Sin acción', 'Dec_externa': 'Cubierto', 'Decision': 'Cubierto',
                    'Costo_unit': costo_u, 'Valor_stock': round(stk * costo_u),
                    'Valor_deficit': 0, 'Is_refill': False,
                    'Is_cambio_plan': True, 'Cambio_cliente': cliente,
                    'Cambio_pt': pt, 'Cambio_comentario': comentario,
                }
                return new_row

            if _col_emb and _col_lug and _col_caj:
                for _, cr in _df_cp.iterrows():
                    embalaje  = str(cr.get(_col_emb, '') or '').strip()
                    lugar     = str(cr.get(_col_lug, '') or '').strip()
                    cajas_val = cr.get(_col_caj, None)
                    cliente   = str(cr.get(_col_cli, '') or '').strip() if _col_cli else ''
                    comentario= str(cr.get(_col_com, '') or '').strip() if _col_com else ''

                    if not embalaje or embalaje in ('nan', 'None', ''):
                        continue
                    if not lugar or lugar in ('nan', 'None', ''):
                        continue
                    try:
                        qty_piezas = float(cajas_val)
                        if qty_piezas <= 0:
                            continue
                    except (TypeError, ValueError):
                        continue

                    alm_sap = _lugar_to_alm.get(lugar, '')
                    if not alm_sap or alm_sap in ('nan', 'None', ''):
                        warn(f"Cambio Plan: lugar '{lugar}' sin conversión a Almacén SAP -> omitido")
                        continue

                    # Resolver PT → BX + CL/CV
                    _emb_info = _pt_map.get(embalaje)
                    if _emb_info:
                        cod_bx = _emb_info['bx']
                        qty_bx = round(qty_piezas * max(1, _emb_info['qty_bx']))
                        cod_cl = _emb_info['cl']
                        qty_cl = round(qty_piezas * _emb_info['qty_cl']) if _emb_info['qty_cl'] > 0 else 0
                    else:
                        cod_bx = embalaje[:6] if len(embalaje) >= 6 else embalaje
                        qty_bx = round(qty_piezas)
                        cod_cl = ''; qty_cl = 0

                    qty_bx_max = round(qty_bx * 1.5)
                    qty_cl_max = round(qty_cl * 1.5) if qty_cl > 0 else 0

                    cambios_plan_list.append({
                        'embalaje': embalaje, 'lugar': lugar, 'alm_sap': alm_sap,
                        'qty_piezas': round(qty_piezas), 'cliente': cliente, 'comentario': comentario,
                        'cod_bx': cod_bx, 'qty_bx': qty_bx,
                        'qty_bx_min': qty_bx, 'qty_bx_max': qty_bx_max,
                        'cod_cl': cod_cl, 'qty_cl': qty_cl,
                        'qty_cl_min': qty_cl, 'qty_cl_max': qty_cl_max,
                    })

                    # ── Aplicar BX ────────────────────────────────────────────────────
                    _bx_row = next((r for r in mrp_plan
                                    if r['Codigo'] == cod_bx and r['Almacen'] == alm_sap), None)
                    if _bx_row:
                        # Sumar al Min/Max existente
                        _bx_row['Min'] = round(_bx_row['Min'] + qty_bx)
                        _bx_row['Max'] = round(_bx_row['Max'] + qty_bx_max)
                    else:
                        # Crear fila nueva para BX en este almacén
                        _bx_row = _nueva_fila_mrp(cod_bx, alm_sap, qty_bx, qty_bx_max,
                                                   cliente, embalaje, comentario)
                        mrp_plan.append(_bx_row)
                        _plan_idx_mut[cod_bx].append(_bx_row)
                        ok(f"Cambio Plan: fila nueva creada -> {cod_bx} en {alm_sap}")

                    # Marcar y guardar metadatos del cambio
                    _bx_row['Is_cambio_plan'] = True
                    _bx_row['Cambio_qty']     = qty_bx
                    _bx_row['Cambio_qty_min'] = qty_bx
                    _bx_row['Cambio_qty_max'] = qty_bx_max
                    _bx_row['Cambio_cliente'] = cliente
                    _bx_row['Cambio_pt']      = embalaje
                    _bx_row['Cambio_cl']      = cod_cl
                    _bx_row['Cambio_cl_qty']  = qty_cl
                    # Recalcular decisión con el nuevo Min
                    _recalc_decision(_bx_row, _plan_idx_mut)
                    _cambios_aplicados += 1

                    # ── Aplicar CL/CV ──────────────────────────────────────────────
                    if cod_cl and qty_cl > 0:
                        _cl_row = next((r for r in mrp_plan
                                        if r['Codigo'] == cod_cl and r['Almacen'] == alm_sap), None)
                        if _cl_row:
                            _cl_row['Min'] = round(_cl_row['Min'] + qty_cl)
                            _cl_row['Max'] = round(_cl_row['Max'] + qty_cl_max)
                        else:
                            _cl_row = _nueva_fila_mrp(cod_cl, alm_sap, qty_cl, qty_cl_max,
                                                       cliente, embalaje, comentario)
                            mrp_plan.append(_cl_row)
                            _plan_idx_mut[cod_cl].append(_cl_row)
                            ok(f"Cambio Plan: fila nueva creada -> {cod_cl} en {alm_sap}")

                        _cl_row['Is_cambio_plan'] = True
                        _cl_row['Cambio_qty']     = qty_cl
                        _cl_row['Cambio_qty_min'] = qty_cl
                        _cl_row['Cambio_qty_max'] = qty_cl_max
                        _cl_row['Cambio_cliente'] = cliente
                        _cl_row['Cambio_pt']      = embalaje
                        _recalc_decision(_cl_row, _plan_idx_mut)
                        _cambios_aplicados += 1

            ok(f"Cambios Plan: {len(cambios_plan_list)} PT leídos | {_cambios_aplicados} filas BX+CL aplicadas en mrp_plan")
        except Exception as _e_cp:
            warn(f"No se pudo leer Cambios_Plan_BP.xlsx: {_e_cp}")
    else:
        ok("Cambios Plan: archivo no encontrado en carpeta (se omite)")

    n_traslado = sum(1 for r in mrp_plan if r['Dec_interna'] == 'Traslado' and not r.get('Is_refill'))
    n_exceso   = sum(1 for r in mrp_plan if (r.get('Exceso_qty') or 0) > 0)
    # Cubiertos = Dec_externa=='Cubierto' (incluye sobre-inventario)
    n_cubierto_total = sum(1 for r in mrp_plan if r['Dec_externa'] == 'Cubierto')
    # REFILL: stock < mín y ya hubo salidas en el período — se reportan aparte, NO cuentan en Cortos/GLM/Traslados
    n_refill = sum(1 for r in mrp_plan if r.get('Is_refill'))
    n_cambio_plan = sum(1 for r in mrp_plan if r.get('Is_cambio_plan'))
    mrp_resumen = {
        'n_total':     len(mrp_plan),
        # Cortos reales = Corto SIN ser Refill
        'n_corto':     sum(1 for r in mrp_plan if r['Dec_externa'] == 'Corto'    and not r.get('Is_refill')),
        'n_cubierto':  n_cubierto_total,
        'n_exceso':    n_exceso,
        # GLM y Traslados también excluyen Refill del conteo del summary
        'n_glm':       sum(1 for r in mrp_plan if r['Dec_externa'] == 'GLM'      and not r.get('Is_refill')),
        'n_traslado':  n_traslado,
        'n_sin_stock': sum(1 for r in mrp_plan if r['Dec_externa'] == 'Sin stock'),
        'n_refill':    n_refill,
        'n_cambio_plan': n_cambio_plan,
        'valor_deficit_total': round(sum(r['Valor_deficit'] for r in mrp_plan)),
    }

    alm_mrp_dict = defaultdict(lambda: {
        'total':0,'cubierto':0,'glm':0,'corto':0,'exceso':0,
        'traslado':0,'sin_stock':0,'refill':0,'valor_deficit':0.0
    })
    for r in mrp_plan:
        a = r['Almacen']
        alm_mrp_dict[a]['total'] += 1
        alm_mrp_dict[a]['valor_deficit'] += r['Valor_deficit']
        ext      = r['Dec_externa']
        refill   = r.get('Is_refill', False)
        if ext == 'Cubierto':   alm_mrp_dict[a]['cubierto']  += 1
        elif refill:            alm_mrp_dict[a]['refill']    += 1   # Refill: no suma en su categoría, va aparte
        elif ext == 'GLM':      alm_mrp_dict[a]['glm']       += 1
        elif ext == 'Corto':    alm_mrp_dict[a]['corto']     += 1
        elif ext == 'Traslado': alm_mrp_dict[a]['traslado']  += 1
        elif ext == 'Sin stock':alm_mrp_dict[a]['sin_stock'] += 1
        if (r.get('Exceso_qty') or 0) > 0:
            alm_mrp_dict[a]['exceso'] += 1
    mrp_por_alm = []
    for a, v in sorted(alm_mrp_dict.items()):
        mrp_por_alm.append({'alm': a, **v,
                            'valor_deficit': round(v['valor_deficit'])})

    prefijos_embalaje = ('CL', 'CV', 'BX')
    codigos_con_plan = set(r['Codigo'] for r in plan_raw)
    inv_sin_demanda_list = []
    df_inv_m = inv_by_alm[inv_by_alm['Almacen_code'].isin(M_ALMS)].copy()
    df_inv_m = df_inv_m[df_inv_m['Codigo'].str[:2].isin([p[:2] for p in prefijos_embalaje])]
    df_inv_m = df_inv_m[~df_inv_m['Codigo'].isin(codigos_con_plan)]
    df_inv_m = df_inv_m[df_inv_m['Stock'] > 0]
    if len(df_inv_m) > 0:
        inv_sin_dem_grp = df_inv_m.groupby('Codigo').agg(
            stock_total=('Stock','sum'),
            n_alm=('Almacen_code','nunique')
        ).reset_index()
        inv_sin_dem_grp['desc']      = inv_sin_dem_grp['Codigo'].map(desc_map).fillna('N/D')
        inv_sin_dem_grp['costo_u']   = inv_sin_dem_grp['Codigo'].map(cos_map).fillna(0)
        inv_sin_dem_grp['monto_mxn'] = (inv_sin_dem_grp['stock_total'] * inv_sin_dem_grp['costo_u']).round(0)
        inv_sin_dem_grp['qty_pal']   = inv_sin_dem_grp['Codigo'].map(pallet_map).fillna(500)
        inv_sin_dem_grp['tarimas']   = (inv_sin_dem_grp['stock_total'] / inv_sin_dem_grp['qty_pal']).round(1)
        inv_sin_dem_grp['prefix']    = inv_sin_dem_grp['Codigo'].str[:2]
        inv_sin_dem_grp = inv_sin_dem_grp.sort_values('monto_mxn', ascending=False)
        inv_sin_demanda_list = inv_sin_dem_grp[
            ['Codigo','desc','prefix','stock_total','n_alm','monto_mxn','tarimas','costo_u']
        ].to_dict(orient='records')
    ok(f"Inv sin demanda (CL/CV/BX sin plan): {len(inv_sin_demanda_list)} SKUs")

    # OPCIÓN A: Cobertura ponderada real = Σstock_nac / Σdem_wk (por SKU único, sin doble conteo por almacén)
    # Evita el sesgo de contar el mismo SKU 6 veces si está en 6 almacenes
    _skus_vistos = {}
    for r in mrp_plan:
        cod = r['Codigo']
        dw  = r.get('Dem_wk_hist', 0.0)
        if dw > 0 and cod not in _skus_vistos:
            _skus_vistos[cod] = {'inv': r.get('Inv_nac', 0.0), 'dw': dw}
    _tot_inv_nac = sum(v['inv'] for v in _skus_vistos.values())
    _tot_dem_nac = sum(v['dw']  for v in _skus_vistos.values())
    sems_cob_general = round(_tot_inv_nac / _tot_dem_nac, 1) if _tot_dem_nac > 0 else 0.0
    # Solo registros con demanda real (excluye almacenes cerrados/sin movimientos donde dw=0 y sems=99)
    _con_dem = [r for r in mrp_plan if r.get('Dem_wk_hist', 0.0) > 0]
    n_con_max = len(_con_dem)
    if _con_dem:
        # % de cobertura vs objetivo 2 semanas:
        # Σ(min(Sems_nac, 2)) / (n × 2) × 100  → 100% = todos cubren 2 semanas o más
        _suma_cob = sum(min(r['Sems_nac'], 2.0) for r in _con_dem)
        pct_cob_2sem = round(_suma_cob / (n_con_max * 2.0) * 100, 1)
    else:
        pct_cob_2sem = 0.0
    _STOCK_MINIMO_BUFFER = 100
    n_cortos_plan      = sum(1 for r in mrp_plan if r.get('Dec_externa') == 'Corto')
    n_cortos_criticos  = sum(1 for r in mrp_plan
                             if r.get('Dec_externa') == 'Corto'
                             and r.get('Stock', 0) <= _STOCK_MINIMO_BUFFER)
    n_cortos_con_stock = n_cortos_plan - n_cortos_criticos
    pct_cob_1sem = round((len(mrp_plan) - n_cortos_criticos) / len(mrp_plan) * 100, 1) if mrp_plan else 100.0
    mrp_resumen['sems_cob_general']  = sems_cob_general
    mrp_resumen['pct_cob_2sem']      = pct_cob_2sem
    mrp_resumen['pct_cob_1sem']      = pct_cob_1sem
    mrp_resumen['n_cortos_plan']     = n_cortos_plan
    mrp_resumen['n_cortos_criticos'] = n_cortos_criticos
    mrp_resumen['n_cortos_con_stock']= n_cortos_con_stock

    from collections import defaultdict as _dd
    _cby = _dd(lambda: {'alms': [], 'desc': '', 'deficit': 0, 'max_stock': 0})
    for _r in mrp_plan:
        if _r.get('Dec_externa') == 'Corto' and _r.get('Stock', 0) <= _STOCK_MINIMO_BUFFER:
            _k = _r['Codigo']
            _cby[_k]['alms'].append(_r['Almacen'])
            _cby[_k]['desc']    = _r.get('Descripcion', '')
            _cby[_k]['deficit'] += _r.get('Deficit', 0)
            _cby[_k]['max_stock'] = max(_cby[_k]['max_stock'], _r.get('Stock', 0))
    cortos_resumen = [
        {'sku': k, 'desc': v['desc'],
         'alms': ', '.join(sorted(set(v['alms']))),
         'n_alms': len(set(v['alms'])),
         'deficit': round(v['deficit']),
         'max_stock': v['max_stock']}
        for k, v in sorted(_cby.items())
    ]
    ok(f"Cobertura ponderada (Sum stk/Sum dem SKUs únicos): {sems_cob_general}sem | Sin cortos críticos: {pct_cob_1sem}% | >=2sem: {pct_cob_2sem}% | Cortos totales: {n_cortos_plan} (críticos sin stock: {n_cortos_criticos}, con buffer: {n_cortos_con_stock})")

    cutoff_30d = TODAY - pd.Timedelta(days=30)
    df_sal_30d = df_v[
        (df_v['fecha_sys'] >= cutoff_30d) &
        (df_v['fecha_sys'] <= TODAY) &
        (df_v['Cantidad'] < 0)
    ].copy()
    mrp_sal_diario = {}
    if not df_sal_30d.empty:
        df_sal_30d['fecha_str'] = df_sal_30d['fecha_sys'].dt.strftime('%Y-%m-%d')
        for (art, alm, fecha), grp in df_sal_30d.groupby(['art', '_Alm', 'fecha_str']):
            key = f"{art}_{alm}"
            qty = float(abs(grp['Cantidad'].sum()))
            if qty > 0:
                if key not in mrp_sal_diario:
                    mrp_sal_diario[key] = []
                mrp_sal_diario[key].append({'f': fecha, 'q': qty})
    ok(f"MRP Sal Diario: {len(mrp_sal_diario)} combinaciones SKU×Alm con salidas en 30d")

    df_var_movs = df_v[
        (df_v['fecha_sys'] >= cutoff_30d) &
        (df_v['fecha_sys'] <= TODAY)
    ].copy()
    var_movs = []
    if not df_var_movs.empty:
        for _, row in df_var_movs.iterrows():
            art = str(row['art']) if pd.notna(row.get('art')) else ''
            alm = str(row.get('_Alm', '')) if pd.notna(row.get('_Alm')) else ''
            qty = float(row['Cantidad']) if pd.notna(row.get('Cantidad')) else 0.0
            fecha = str(row.get('fecha_str', ''))
            desc = desc_map.get(art, 'N/D')
            mxn = float(row.get('valor_calc', 0) or 0)
            if not art or qty == 0:
                continue
            var_movs.append({'art': art, 'alm': alm, 'fecha': fecha,
                             'cantidad': qty, 'desc': desc,
                             'mxn': round(mxn * (1 if qty > 0 else -1), 2)})
    ok(f"Var movimientos (30d): {len(var_movs)} registros")

    print(f"\n  OK Módulo MRP completo")
    print(f"     Total: {mrp_resumen['n_total']} | "
          f"Cubiertos: {mrp_resumen['n_cubierto']} | "
          f"Traslados: {mrp_resumen['n_traslado']} | "
          f"GLM: {mrp_resumen['n_glm']} | "
          f"Cortos: {mrp_resumen['n_corto']} | "
          f"Excesos: {mrp_resumen['n_exceso']}")
    print(f"     Valor déficit total: ${mrp_resumen['valor_deficit_total']:,.0f} MXN")

    # ══════════════════════════════════════════════════════════════════════════
    # GENERAR EXCEL DE TRACKING — Cortos, Traslados, GLM
    # Cada run regenera el archivo con datos actuales.
    # Si ya existe, preserva los estatus/comentarios que Compras/Almacén llenó.
    # ══════════════════════════════════════════════════════════════════════════
    import openpyxl as _oxl
    from openpyxl.styles import Font as _OFont, PatternFill as _OFill, Alignment as _OAlign, Border as _OBorder, Side as _OSide

    def _hdr_style(ws, row, cols, bg='1E3A5F'):
        _s = _OSide(style='thin', color='93C5FD')
        _b = _OBorder(left=_s, right=_s, top=_s, bottom=_s)
        for c in cols:
            cell = ws.cell(row=row, column=c)
            cell.font      = _OFont(name='Arial', bold=True, size=9, color='FFFFFF')
            cell.fill      = _OFill('solid', fgColor=bg)
            cell.alignment = _OAlign(horizontal='center', vertical='center', wrap_text=True)
            cell.border    = _b

    def _row_style(ws, row, n_cols, is_alt=False):
        _s = _OSide(style='thin', color='BFDBFE')
        _b = _OBorder(left=_s, right=_s, top=_s, bottom=_s)
        bg = 'EFF6FF' if is_alt else 'FFFFFF'
        for c in range(1, n_cols+1):
            cell = ws.cell(row=row, column=c)
            if not cell.fill or cell.fill.fgColor.rgb in ('00000000','FFFFFFFF','FFFFFF'):
                cell.fill = _OFill('solid', fgColor=bg)
            cell.font   = _OFont(name='Arial', size=9)
            cell.border = _b
            cell.alignment = _OAlign(vertical='center', wrap_text=True)

    # ── CAMPOS EDITABLES — dict maestro cargado ANTES de sobreescribir cualquier Excel ──
    # Clave: (Codigo, Almacen) → {Folio_OC, Estatus, Fecha_Carga, Owner, Comentarios}
    # Se construye leyendo los 3 archivos de tracking en su estado ACTUAL (con datos del usuario).
    # Luego se usa tanto para repoblar los Excels regenerados como para inyectar en el JSON.
    _EDIT_FIELDS = ('Folio_OC','Estatus','Fecha_Carga','Owner','Comentarios')

    def _load_prev_tracking(path, key_cols):
        """Lee el archivo previo y devuelve dict {tuple(keys): {col: val}} de columnas de estatus."""
        prev = {}
        if not path.exists():
            return prev
        try:
            _wb = _oxl.load_workbook(str(path), read_only=True, data_only=True)
            _ws = _wb.active
            rows = list(_ws.values)
            if not rows:
                return prev
            hdrs = [str(h).strip() if h else '' for h in rows[0]]
            key_idxs  = [hdrs.index(k) for k in key_cols if k in hdrs]
            stat_idxs = {h: i for i, h in enumerate(hdrs)
                         if h in _EDIT_FIELDS or h in ('Notas',)}
            for r in rows[1:]:
                if not r or all(v is None for v in r):
                    continue
                key = tuple(str(r[i] or '').strip() for i in key_idxs)
                entry = {}
                for col, idx in stat_idxs.items():
                    v = r[idx]
                    if hasattr(v, 'strftime'):
                        v = v.strftime('%Y-%m-%d')
                    entry[col] = str(v).strip() if v is not None else ''
                # Solo guardar si tiene al menos un campo con dato real
                if any(entry.values()):
                    prev[key] = entry
            _wb.close()
        except Exception:
            pass
        return prev

    # Cargar datos editables de los 3 archivos ANTES de regenerarlos
    # Se leen tanto de la RAÍZ (BASE_DIR) como de /Reportes (REPORTES_DIR) para asegurar que no se pierdan datos del usuario
    _master_track = {}   # (Codigo, Almacen) → {campo: valor}
    _tracking_loaded_from_supabase = False
    
    if SUPABASE_URL and SUPABASE_KEY:
        try:
            print("  [*] Descargando tracking de compras y MRP desde Supabase...")
            _sup_tracks = supabase_get("mrp_tracking")
            if _sup_tracks is not None:
                for _row in _sup_tracks:
                    _sku = str(_row.get('sku', '')).strip()
                    _alm = str(_row.get('alm', '')).strip()
                    _status = _row.get('status')
                    _comments = _row.get('comentarios')
                    if _sku and _alm:
                        _master_track[(_sku, _alm)] = {
                            'Codigo': _sku,
                            'Almacen': _alm,
                            'Estatus': _status or '',
                            'Comentarios': _comments or '',
                            'Folio_OC': '',
                            'Fecha_Carga': '',
                            'Owner': ''
                        }
                _n_con_datos = sum(1 for v in _master_track.values() if any(v.values()))
                ok(f"Tracking previo cargado desde Supabase: {_n_con_datos} registros con datos editables")
                _tracking_loaded_from_supabase = True
        except Exception as _esup_tr:
            warn(f"Error descargando tracking de Supabase: {_esup_tr}")

    if not _tracking_loaded_from_supabase:
        for _folder in (BASE_DIR, REPORTES_DIR):
            for _filename, _file_path in [
                ('Tracking_Cortos_BP.xlsx', _folder / 'Tracking_Cortos_BP.xlsx'),
                ('Tracking_Traslados_BP.xlsx', _folder / 'Tracking_Traslados_BP.xlsx'),
                ('Tracking_GLM_BP.xlsx', _folder / 'Tracking_GLM_BP.xlsx')
            ]:
                if _file_path.exists():
                    for _key, _vals in _load_prev_tracking(_file_path, ['Codigo','Almacen']).items():
                        if _key not in _master_track:
                            _master_track[_key] = _vals
                        else:
                            for _f, _v in _vals.items():
                                if _v and not _master_track[_key].get(_f):
                                    _master_track[_key][_f] = _v
        _n_con_datos = sum(1 for v in _master_track.values() if any(v.values()))
        ok(f"Tracking previo cargado (local): {_n_con_datos} registros con datos editables")

    ESTATUS_COMPRAS   = ['','Pendiente','En proceso','Ordenado','En tránsito','Entregado','Cancelado']
    ESTATUS_TRASLADOS = ['','Pendiente','En proceso','En tránsito','Recibido','Cancelado']
    ESTATUS_GLM       = ['','Pendiente','Cotizado','Ordenado','En carga','Cargado','Enviado','Recibido']

    def _add_dv(ws, col_letter, first_row, last_row, options):
        from openpyxl.worksheet.datavalidation import DataValidation as _DV
        formula = '"' + ','.join(o for o in options if o) + '"'
        dv = _DV(type='list', formula1=formula, allow_blank=True,
                 showErrorMessage=False)
        dv.sqref = f'{col_letter}{first_row}:{col_letter}{last_row}'
        ws.add_data_validation(dv)

    # ── CORTOS ────────────────────────────────────────────────────────────────
    step("Generando Tracking_Cortos_BP.xlsx")
    try:
        _cortos = [r for r in mrp_plan if r.get('Dec_externa') == 'Corto']
        _prev_c = _master_track  # usa el dict maestro cargado antes de regenerar
        _wb_c = _oxl.Workbook(); _wb_c.remove(_wb_c.active)
        _ws_c = _wb_c.create_sheet('Cortos')
        _ws_c.sheet_view.showGridLines = False
        _ws_c.tab_color = 'EF4444'
        _hdrs_c = ['Codigo','Descripcion','Almacen','Stock','Min','Deficit',
                   'Sems_local','Sems_nac','Dem_wk_hist','Valor_deficit','Costo_unit',
                   'Refill','Cambio_Plan','Cambio_Qty','Cambio_Cliente',
                   'Folio_OC','Estatus','Fecha_Carga','Owner','Comentarios']
        _col_w_c = [10,36,10,10,10,10,10,10,12,14,12,8,8,10,18,14,16,14,14,30]
        _ws_c.row_dimensions[1].height = 26
        for _ci, (_h, _w) in enumerate(zip(_hdrs_c, _col_w_c), 1):
            _ws_c.cell(row=1, column=_ci, value=_h)
            _ws_c.column_dimensions[_oxl.utils.get_column_letter(_ci)].width = _w
        _hdr_style(_ws_c, 1, range(1, len(_hdrs_c)+1), bg='7F1D1D')
        _ws_c.freeze_panes = 'A2'
        for _ri, _r in enumerate(_cortos, 2):
            _key = (_r.get('Codigo',''), _r.get('Almacen',''))
            _prev = _prev_c.get(_key, {})
            _vals = [
                _r.get('Codigo',''), _r.get('Descripcion',''), _r.get('Almacen',''),
                _r.get('Stock',0), _r.get('Min',0), _r.get('Deficit',0),
                round(_r.get('Sems_local',0),2), round(_r.get('Sems_nac',0),2),
                round(_r.get('Dem_wk_hist',0),1), round(_r.get('Valor_deficit',0)),
                round(_r.get('Costo_unit',0),2),
                'SÍ' if _r.get('Is_refill') else '',
                'SÍ' if _r.get('Is_cambio_plan') else '',
                _r.get('Cambio_qty',''), _r.get('Cambio_cliente',''),
                _prev.get('Folio_OC',''), _prev.get('Estatus',''),
                _prev.get('Fecha_Carga',''), _prev.get('Owner',''), _prev.get('Comentarios',''),
            ]
            for _ci, _v in enumerate(_vals, 1):
                _ws_c.cell(row=_ri, column=_ci, value=_v)
            _row_style(_ws_c, _ri, len(_hdrs_c), is_alt=_ri%2==0)
            # Colorear Estatus y Deficit; amarillo en cols editables (P=16,Q=17,R=18)
            _ws_c.cell(row=_ri, column=6).font  = _OFont(name='Arial', size=9, color='EF4444', bold=True)
            for _ec in (16, 17, 18):
                _ws_c.cell(row=_ri, column=_ec).fill = _OFill('solid', fgColor='FFFBEB')
        _add_dv(_ws_c, 'Q', 2, len(_cortos)+1, ESTATUS_COMPRAS)
        _wb_c.save(str(TRACKING_CORTOS_FILE))
        ok(f"Tracking_Cortos_BP.xlsx: {len(_cortos)} cortos")
    except Exception as _ec:
        warn(f"Error generando Tracking_Cortos_BP.xlsx: {_ec}")

    # ── TRASLADOS ─────────────────────────────────────────────────────────────
    step("Generando Tracking_Traslados_BP.xlsx")
    try:
        _traslados = [r for r in mrp_plan if r.get('Dec_interna') == 'Traslado']
        _prev_t = _master_track  # usa el dict maestro cargado antes de regenerar
        _wb_t = _oxl.Workbook(); _wb_t.remove(_wb_t.active)
        _ws_t = _wb_t.create_sheet('Traslados')
        _ws_t.sheet_view.showGridLines = False
        _ws_t.tab_color = '38BDF8'
        _hdrs_t = ['Codigo','Descripcion','Almacen','Alm_fuente','Stock','Min','Deficit',
                   'Transferible','Sems_local','Sems_nac','Pct_inv',
                   'Refill','Cambio_Plan','Cambio_Qty','Cambio_Cliente',
                   'Folio_OC','Estatus','Fecha_Carga','Owner','Comentarios']
        _col_w_t = [10,36,10,12,10,10,10,12,10,10,10,8,8,10,18,14,16,14,14,30]
        _ws_t.row_dimensions[1].height = 26
        for _ci, (_h, _w) in enumerate(zip(_hdrs_t, _col_w_t), 1):
            _ws_t.cell(row=1, column=_ci, value=_h)
            _ws_t.column_dimensions[_oxl.utils.get_column_letter(_ci)].width = _w
        _hdr_style(_ws_t, 1, range(1, len(_hdrs_t)+1), bg='0C4A6E')
        _ws_t.freeze_panes = 'A2'
        for _ri, _r in enumerate(_traslados, 2):
            _key = (_r.get('Codigo',''), _r.get('Almacen',''))
            _prev = _prev_t.get(_key, {})
            _vals = [
                _r.get('Codigo',''), _r.get('Descripcion',''), _r.get('Almacen',''),
                _r.get('Alm_fuente',''), _r.get('Stock',0), _r.get('Min',0), _r.get('Deficit',0),
                _r.get('Transferible',0), round(_r.get('Sems_local',0),2),
                round(_r.get('Sems_nac',0),2), round(_r.get('Pct_inv',0),1),
                'SÍ' if _r.get('Is_refill') else '',
                'SÍ' if _r.get('Is_cambio_plan') else '',
                _r.get('Cambio_qty',''), _r.get('Cambio_cliente',''),
                _prev.get('Folio_OC',''), _prev.get('Estatus',''),
                _prev.get('Fecha_Carga',''), _prev.get('Owner',''), _prev.get('Comentarios',''),
            ]
            for _ci, _v in enumerate(_vals, 1):
                _ws_t.cell(row=_ri, column=_ci, value=_v)
            _row_style(_ws_t, _ri, len(_hdrs_t), is_alt=_ri%2==0)
            for _ec in (16, 17, 18):
                _ws_t.cell(row=_ri, column=_ec).fill = _OFill('solid', fgColor='F0F9FF')
        _add_dv(_ws_t, 'Q', 2, len(_traslados)+1, ESTATUS_TRASLADOS)
        _wb_t.save(str(TRACKING_TRASLADOS_FILE))
        ok(f"Tracking_Traslados_BP.xlsx: {len(_traslados)} traslados")
    except Exception as _et:
        warn(f"Error generando Tracking_Traslados_BP.xlsx: {_et}")

    # ── GLM ───────────────────────────────────────────────────────────────────
    step("Generando Tracking_GLM_BP.xlsx")
    try:
        _glm = [r for r in mrp_plan if r.get('Dec_externa') == 'GLM']
        _prev_g = _master_track  # usa el dict maestro cargado antes de regenerar
        _wb_g = _oxl.Workbook(); _wb_g.remove(_wb_g.active)
        _ws_g = _wb_g.create_sheet('GLM')
        _ws_g.sheet_view.showGridLines = False
        _ws_g.tab_color = 'F59E0B'
        _hdrs_g = ['Codigo','Descripcion','Almacen','Stock','Min','Deficit',
                   'Inv_glm','Sems_local','Sems_nac','Dem_wk_hist','Valor_deficit','Costo_unit',
                   'Refill','Cambio_Plan','Cambio_Qty','Cambio_Cliente',
                   'Folio_OC','Estatus','Fecha_Carga','Owner','Comentarios']
        _col_w_g = [10,36,10,10,10,10,12,10,10,12,14,12,8,8,10,18,14,16,14,14,30]
        _ws_g.row_dimensions[1].height = 26
        for _ci, (_h, _w) in enumerate(zip(_hdrs_g, _col_w_g), 1):
            _ws_g.cell(row=1, column=_ci, value=_h)
            _ws_g.column_dimensions[_oxl.utils.get_column_letter(_ci)].width = _w
        _hdr_style(_ws_g, 1, range(1, len(_hdrs_g)+1), bg='78350F')
        _ws_g.freeze_panes = 'A2'
        for _ri, _r in enumerate(_glm, 2):
            _key = (_r.get('Codigo',''), _r.get('Almacen',''))
            _prev = _prev_g.get(_key, {})
            _vals = [
                _r.get('Codigo',''), _r.get('Descripcion',''), _r.get('Almacen',''),
                _r.get('Stock',0), _r.get('Min',0), _r.get('Deficit',0),
                round(_r.get('Inv_glm',0)), round(_r.get('Sems_local',0),2),
                round(_r.get('Sems_nac',0),2), round(_r.get('Dem_wk_hist',0),1),
                round(_r.get('Valor_deficit',0)), round(_r.get('Costo_unit',0),2),
                'SÍ' if _r.get('Is_refill') else '',
                'SÍ' if _r.get('Is_cambio_plan') else '',
                _r.get('Cambio_qty',''), _r.get('Cambio_cliente',''),
                _prev.get('Folio_OC',''), _prev.get('Estatus',''),
                _prev.get('Fecha_Carga',''), _prev.get('Owner',''), _prev.get('Comentarios',''),
            ]
            for _ci, _v in enumerate(_vals, 1):
                _ws_g.cell(row=_ri, column=_ci, value=_v)
            _row_style(_ws_g, _ri, len(_hdrs_g), is_alt=_ri%2==0)
            for _ec in (17, 18, 19):
                _ws_g.cell(row=_ri, column=_ec).fill = _OFill('solid', fgColor='FFFBEB')
        _add_dv(_ws_g, 'R', 2, len(_glm)+1, ESTATUS_GLM)
        _wb_g.save(str(TRACKING_GLM_FILE))
        ok(f"Tracking_GLM_BP.xlsx: {len(_glm)} GLM")
    except Exception as _eg:
        warn(f"Error generando Tracking_GLM_BP.xlsx: {_eg}")

except Exception as e:
    warn(f"Error en módulo MRP: {e}")
    traceback.print_exc()
    mrp_plan = []
    mrp_resumen = {
        'n_total':0,'n_corto':0,'n_cubierto':0,'n_exceso':0,
        'n_glm':0,'n_traslado':0,'n_sin_stock':0,'n_refill':0,
        'n_cambio_plan':0,'valor_deficit_total':0,
        'sems_cob_general':0,'pct_cob_2sem':0
    }
    mrp_por_alm = []
    glm_prov = []
    inv_sin_demanda_list = []
    mrp_sal_diario = {}
    var_movs = []
    cortos_resumen = []
    inv_nac_list = []
    cambios_plan_list = []
    _master_track = {}
    _n_con_datos  = 0


# ══════════════════════════════════════════════════════════════════════════════
#  MÓDULO PLANEACIÓN — BoM × Inventario sin plan por almacén
# ══════════════════════════════════════════════════════════════════════════════

step("Módulo Planeación — BoM × Inventario Sin Plan por almacén")
planeacion_list = []

try:
    if df_emb is not None and len(df_emb) > 0:
        emb = df_emb.copy()
        if 'Unnamed: 0' in str(emb.columns[0]):
            emb.columns = emb.iloc[0]
            emb = emb.iloc[1:].reset_index(drop=True)
        emb.columns = [str(c).strip() for c in emb.columns]

        c_pt   = next((c for c in emb.columns if 'pt' in c.lower()), None)
        c_clam = next((c for c in emb.columns if 'clam' in c.lower() and 'cod' in c.lower()), None)
        c_qcl  = next((c for c in emb.columns if 'clam' in c.lower() and 'cant' in c.lower()), None)
        c_caja = next((c for c in emb.columns if 'caja' in c.lower() and 'cod' in c.lower()), None)
        c_qcj  = next((c for c in emb.columns if 'caja' in c.lower() and 'cant' in c.lower()), None)

        if c_pt and c_clam and c_qcl and c_caja and c_qcj:
            emb = emb.dropna(subset=[c_pt])
            emb[c_qcl] = pd.to_numeric(emb[c_qcl], errors='coerce').fillna(0)
            emb[c_qcj] = pd.to_numeric(emb[c_qcj], errors='coerce').fillna(0)

            plan_alloc = {}
            if df_plan is not None:
                try:
                    _c_pc = next((c for c in df_plan.columns if 'codigo' in c.lower()), None)
                    _c_pa = next((c for c in df_plan.columns if 'almacen' in c.lower()), None)
                    _c_pm = next((c for c in df_plan.columns if c.lower().strip() == 'min'
                                  or c.lower().strip() == 'mínimo'
                                  or c.lower().strip() == 'minimo'), None)
                    if _c_pc and _c_pa and _c_pm:
                        for _, pr in df_plan.iterrows():
                            _alm_raw = ALM_SAP_TO_INTERNO.get(str(pr[_c_pa]).strip(), str(pr[_c_pa]).strip())
                            _k = (str(pr[_c_pc]).strip(), _alm_raw)
                            _min_val   = float(pr[_c_pm]) if pd.notna(pr[_c_pm]) else 0.0
                            _sal_week  = float(sal_sat_by_alm.get(_k, 0.0))
                            plan_alloc[_k] = max(0.0, _min_val - _sal_week)
                except Exception:
                    pass

            _prefijos_bom = ('CL', 'CV', 'BX')
            # Almacenes permitidos en Planeación: los operativos + cualquiera que aparezca en inventario M (que inicien con M y no excluidos)
            alms_planeacion = set(ALMACENES_OPERATIVOS)
            if 'Almacen_code' in inv_by_alm.columns:
                alms_inv_m = inv_by_alm[
                    inv_by_alm['Almacen_code'].str.startswith('M') &
                    ~inv_by_alm['Almacen_code'].isin(ALM_EXCLUIR_NAC)
                ]['Almacen_code'].unique()
                alms_planeacion.update(alms_inv_m)

            stock_alm = inv_by_alm[
                inv_by_alm['Almacen_code'].isin(alms_planeacion) &
                (inv_by_alm['Stock'] > 0) &
                inv_by_alm['Codigo'].str[:2].isin([p[:2] for p in _prefijos_bom])
            ].copy()
            stock_alm.columns = ['Codigo', 'Almacen', 'Stock']

            alms_with_stock = stock_alm['Almacen'].unique()

            for alm in sorted(alms_with_stock):
                alm_stock = {r.Codigo: r.Stock for r in stock_alm[stock_alm['Almacen'] == alm].itertuples()}

                for _, row in emb.iterrows():
                    pt_code    = str(row[c_pt]).strip()
                    clam_code  = str(row[c_clam]).strip() if pd.notna(row[c_clam]) else ''
                    caja_code  = str(row[c_caja]).strip() if pd.notna(row[c_caja]) else ''
                    qty_clam   = float(row[c_qcl])
                    qty_caja   = float(row[c_qcj])

                    if not clam_code or not caja_code:
                        continue

                    stk_clam_total = alm_stock.get(clam_code, 0)
                    stk_caja_total = alm_stock.get(caja_code, 0)

                    if stk_clam_total == 0 and stk_caja_total == 0:
                        continue

                    alloc_clam = min(plan_alloc.get((clam_code, alm), 0), stk_clam_total)
                    alloc_caja = min(plan_alloc.get((caja_code, alm), 0), stk_caja_total)

                    stk_clam = max(0, stk_clam_total - alloc_clam)
                    stk_caja = max(0, stk_caja_total - alloc_caja)

                    pt_por_clam = int(stk_clam / qty_clam) if qty_clam > 0 else 0
                    pt_por_caja = int(stk_caja / qty_caja) if qty_caja > 0 else 0
                    pt_max      = min(pt_por_clam, pt_por_caja)

                    if pt_max == 0 and not (stk_clam_total > 0 and stk_caja_total > 0):
                        continue

                    tipo_clam = clam_code[:2] if len(clam_code) >= 2 else '??'
                    desc_clam = desc_map.get(clam_code, '')

                    planeacion_list.append({
                        'pt':              pt_code,
                        'alm':             alm,
                        'clam':            clam_code,
                        'qty_clam':        int(qty_clam),
                        'stk_clam':        int(stk_clam),
                        'stk_clam_total':  int(stk_clam_total),
                        'alloc_clam':      int(alloc_clam),
                        'pt_x_clam':       pt_por_clam,
                        'caja':            caja_code,
                        'qty_caja':        int(qty_caja),
                        'stk_caja':        int(stk_caja),
                        'stk_caja_total':  int(stk_caja_total),
                        'alloc_caja':      int(alloc_caja),
                        'pt_x_caja':       pt_por_caja,
                        'pt_max':          pt_max,
                        'tipo_cl':         tipo_clam,
                        'desc_pt':         desc_clam or pt_code,
                    })

            planeacion_list.sort(key=lambda x: (x['alm'], -x['pt_max']))
            ok(f"Planeación: {len(planeacion_list)} combinaciones PT×Almacén posibles")
        else:
            warn(f"Embalajes: columnas no encontradas ({list(emb.columns)[:5]})")
    else:
        warn("Hoja Embalajes no disponible")
except Exception as e:
    warn(f"Error en módulo Planeación: {e}")
    import traceback; traceback.print_exc()
    planeacion_list = []

# ══════════════════════════════════════════════════════════════════════════════
#  HISTÓRICO PLAN SEMANAL
# ══════════════════════════════════════════════════════════════════════════════

step("Actualizando histórico semanal Plan MRP")

HISTORICO_PLAN = SISTEMA_DIR / '_historico_plan_semanal.json'

def load_historico():
    try:
        if HISTORICO_PLAN.exists():
            with open(str(HISTORICO_PLAN), 'r', encoding='utf-8') as f:
                return json.load(f)
    except Exception as e:
        warn(f"Error leyendo histórico semanal: {e}")
    return {}

def save_historico(hist_dict):
    try:
        with open(str(HISTORICO_PLAN), 'w', encoding='utf-8') as f:
            json.dump(hist_dict, f, ensure_ascii=False, indent=2, default=str)
    except Exception as e:
        warn(f"Error guardando histórico semanal: {e}")

_dow = TODAY.weekday()
_days_since_sat = (_dow - 5) % 7
_week_start_sat = TODAY - pd.Timedelta(days=_days_since_sat)
_week_end_fri   = _week_start_sat + pd.Timedelta(days=6)

semana_key   = _week_start_sat.strftime('%Y-%m-%d')
semana_label = (f"Sem {_week_start_sat.strftime('%d/%m')}–"
                f"{_week_end_fri.strftime('%d/%m/%Y')}")

plan_snap = []
for r in mrp_plan:
    plan_snap.append({
        'cod': r.get('Codigo',''),
        'alm': r.get('Almacen',''),
        'stk': round(r.get('Stock', 0), 0),
        'min': round(r.get('Min', 0), 0),
        'max': round(r.get('Max', 0), 0),
        'def': round(r.get('Deficit', 0), 0),
        'sn':  round(r.get('Sems_nac', 0), 2),
        'sl':  round(r.get('Sems_local', 0), 2),
        's2w': round(r.get('Sal_2wk', 0), 0),
        'dp':  round(r.get('Delta_pend', 0), 0),
        'di':  r.get('Dec_interna',''),
        'de':  r.get('Dec_externa',''),
        'vd':  round((r.get('Deficit', 0)) * (r.get('Costo_unit', 0)), 0),
    })

hist_data = load_historico()

prev_keys = sorted([k for k in hist_data if k != semana_key], reverse=True)
prev_snap  = hist_data[prev_keys[0]] if prev_keys else {}
prev_plan_map = {(r['cod'], r['alm']): r for r in prev_snap.get('plan', [])}
curr_plan_map = {(r['cod'], r['alm']): r for r in plan_snap}

diff_nuevos    = []
diff_salieron  = []
diff_empeoro   = []
diff_mejoro    = []
diff_dec_cambio= []
diff_qty       = []

_malos  = {'Corto', 'GLM'}
_buenos = {'Cubierto', 'Traslado'}

for key, cur in curr_plan_map.items():
    prv = prev_plan_map.get(key)
    if prv is None:
        diff_nuevos.append({'cod': cur['cod'], 'alm': cur['alm'],
                            'de': cur['de'], 'sn': cur['sn'], 'stk': cur['stk']})
        continue
    de_prv, de_cur = prv.get('de',''), cur.get('de','')
    if de_prv != de_cur:
        entry = {'cod': cur['cod'], 'alm': cur['alm'],
                 'de_ant': de_prv, 'de_new': de_cur,
                 'sn_ant': prv.get('sn', 0), 'sn_new': cur.get('sn', 0),
                 'stk': cur['stk']}
        if de_prv in _buenos and de_cur in _malos:
            diff_empeoro.append(entry)
        elif de_prv in _malos and de_cur in _buenos:
            diff_mejoro.append(entry)
        else:
            diff_dec_cambio.append(entry)
    delta_stk = abs((cur.get('stk', 0) or 0) - (prv.get('stk', 0) or 0))
    ref_stk   = max(abs(prv.get('stk', 0) or 0), 1)
    if delta_stk > 50 and (delta_stk / ref_stk) > 0.10:
        diff_qty.append({'cod': cur['cod'], 'alm': cur['alm'],
                         'stk_ant': prv.get('stk', 0), 'stk_new': cur.get('stk', 0),
                         'sn_ant': prv.get('sn', 0), 'sn_new': cur.get('sn', 0),
                         'de': cur.get('de','')})

for key, prv in prev_plan_map.items():
    if key not in curr_plan_map:
        diff_salieron.append({'cod': prv['cod'], 'alm': prv['alm'],
                              'de': prv.get('de',''), 'sn': prv.get('sn', 0)})

_hay_cambios = bool(diff_nuevos or diff_salieron or diff_empeoro or
                    diff_mejoro or diff_dec_cambio or diff_qty)

diff_resumen = {
    'n_nuevos':     len(diff_nuevos),
    'n_salieron':   len(diff_salieron),
    'n_empeoro':    len(diff_empeoro),
    'n_mejoro':     len(diff_mejoro),
    'n_dec_cambio': len(diff_dec_cambio),
    'n_qty':        len(diff_qty),
    'hay_cambios':  _hay_cambios,
    'nuevos':       diff_nuevos[:50],
    'salieron':     diff_salieron[:50],
    'empeoro':      diff_empeoro[:50],
    'mejoro':       diff_mejoro[:50],
    'dec_cambio':   diff_dec_cambio[:50],
    'qty':          diff_qty[:50],
    'semana_prev':  prev_keys[0] if prev_keys else None,
}

snap_kpis = {
    'semana':       semana_key,
    'label':        semana_label,
    'fecha':        TODAY.strftime('%d/%m/%Y'),
    'n_total':      mrp_resumen.get('n_total', 0),
    'n_cubierto':   mrp_resumen.get('n_cubierto', 0),
    'n_traslado':   mrp_resumen.get('n_traslado', 0),
    'n_glm':        mrp_resumen.get('n_glm', 0),
    'n_corto':      mrp_resumen.get('n_corto', 0),
    'n_exceso':     mrp_resumen.get('n_exceso', 0),
    'sems_cob':     mrp_resumen.get('sems_cob_general', 0),
    'pct_1sem':     mrp_resumen.get('pct_cob_1sem', 0),
    'pct_2sem':     mrp_resumen.get('pct_cob_2sem', 0),
    'val_def':      mrp_resumen.get('valor_deficit_total', 0),
    'mate2122_mxn': total_mate2122_mxn,
    'mate2122_pzs': total_mate2122_pzs,
    'diff':         diff_resumen,
    'plan':         plan_snap,
}

hist_data[semana_key] = snap_kpis
save_historico(hist_data)

hist_list = sorted(hist_data.values(),
                   key=lambda x: x.get('semana',''),
                   reverse=True)
hist_summary = [{k: v for k, v in s.items() if k != 'plan'} for s in hist_list]
hist_full    = hist_list[:20]

_cambios_txt = (f"{diff_resumen['n_empeoro']} deterioros, "
                f"{diff_resumen['n_mejoro']} mejoras, "
                f"{diff_resumen['n_nuevos']} nuevos, "
                f"{diff_resumen['n_salieron']} salieron"
                if _hay_cambios else "sin cambios vs semana anterior")
ok(f"Histórico BP: {len(hist_data)} semanas | Semana: {semana_key} ({semana_label}) | {_cambios_txt}")

# ══════════════════════════════════════════════════════════════════════════════
#  MÓDULO 3 — PLAN MRP MANUAL vs CONSUMOS REALES (semana sábado→viernes)
#  Lee hoja Plan_MRP del Excel maestro, calcula consumos por SKU×Alm×Semana
#  excluyendo traslados internos, congela snapshots marcados y produce dataset
#  para la pestaña "Plan vs Consumo".
# ══════════════════════════════════════════════════════════════════════════════

step("Plan MRP manual vs consumos reales")

HISTORICO_PLAN_MRP_DET = SISTEMA_DIR / '_historico_plan_mrp_detalle.json'

_PLAN_MRP_SHEET = 'Plan_MRP'
_PLAN_MRP_HDRS  = ['Codigo', 'Almacen', 'Semana_ID', 'Semana_Inicio',
                   'Cantidad_Plan', 'Notas', 'Congelar', 'Fecha_Congelado']

def _semana_sabado(fecha):
    """Devuelve (semana_id_str_YYYY-MM-DD, fecha_sabado) para la semana
    sábado→viernes que contiene 'fecha'."""
    ts = pd.Timestamp(fecha)
    if pd.isna(ts):
        return (None, None)
    dias = (ts.weekday() - 5) % 7   # 5 = sábado
    inicio = (ts - pd.Timedelta(days=dias)).normalize()
    return (inicio.strftime('%Y-%m-%d'), inicio)

def _wk_label(sem_id):
    """'2026-05-16' -> 'WK20'."""
    try:
        ts = pd.Timestamp(sem_id)
        # Para semana sábado→viernes, etiquetar por la semana ISO del DOMINGO siguiente
        # (más intuitivo: WK20 si el viernes cae en semana ISO 20)
        viernes = ts + pd.Timedelta(days=6)
        wk = viernes.isocalendar()[1]
        return f'WK{wk:02d}'
    except Exception:
        return str(sem_id) if sem_id else ''

def _semana_abierta(sem_id, hoy_ts):
    """True si la semana aún no cierra (hoy <= viernes de esa semana)."""
    try:
        ini = pd.Timestamp(sem_id)
        fin = ini + pd.Timedelta(days=6)
        return pd.Timestamp(hoy_ts).normalize() <= fin
    except Exception:
        return False

def _ensure_plan_mrp_sheet(xlsx_path):
    """Si la hoja Plan_MRP no existe en el Excel maestro, la crea con headers.
    Devuelve True si se creó (o ya existía), False si falló (archivo bloqueado)."""
    if _PLAN_MRP_SHEET in sheets_available:
        return True
    try:
        import openpyxl as _xl
        from openpyxl.styles import PatternFill, Font, Alignment
        _wb = _xl.load_workbook(str(xlsx_path))
        if _PLAN_MRP_SHEET in _wb.sheetnames:
            _wb.close()
            return True
        _ws = _wb.create_sheet(_PLAN_MRP_SHEET)
        _hdr_fill = PatternFill('solid', fgColor='6B1FA2')
        _hdr_font = Font(bold=True, color='FFFFFF', size=10)
        for _ci, _h in enumerate(_PLAN_MRP_HDRS, 1):
            _c = _ws.cell(row=1, column=_ci, value=_h)
            _c.fill = _hdr_fill
            _c.font = _hdr_font
            _c.alignment = Alignment(horizontal='center')
        for _ci, _w in enumerate([10, 10, 14, 14, 14, 30, 10, 22], 1):
            _ws.column_dimensions[_xl.utils.get_column_letter(_ci)].width = _w
        _ws.row_dimensions[1].height = 18
        _ws.sheet_properties.tabColor = '6B1FA2'
        _ws.freeze_panes = 'A2'
        _wb.save(str(xlsx_path))
        _wb.close()
        ok(f"Hoja {_PLAN_MRP_SHEET} creada en Excel maestro (vacía, lista para captura)")
        return True
    except PermissionError:
        warn(f"No se pudo crear hoja {_PLAN_MRP_SHEET} (Excel abierto). Ciérralo y reintenta.")
        return False
    except Exception as _e:
        warn(f"No se pudo crear hoja {_PLAN_MRP_SHEET}: {_e}")
        return False

def _identificar_traslados(df_hist_local):
    """Set de Documentos con suma neta ≈0 y al menos un + y un - (traslados internos)."""
    try:
        _doc_col = col(df_hist_local, 'Documento', 'documento')
        _qty_col = col(df_hist_local, 'Cantidad', 'cantidad')
        _g = df_hist_local.groupby(_doc_col)[_qty_col].agg(['sum', 'min', 'max'])
        return set(_g[(_g['sum'].abs() < 0.01) & (_g['min'] < 0) & (_g['max'] > 0)].index)
    except Exception:
        return set()

def _calcular_consumos_semanales(df_hist_local, semanas_objetivo=None, incluir_traslados=False):
    """Agrupa Σ|Cantidad| de salidas (Cant<0) por SKU×Almacén×Semana(sábado).
    Si incluir_traslados=False, excluye documentos identificados como traslados internos.
    Retorna dict: {(codigo, almacen, semana_id): cantidad_consumida}."""
    try:
        _sku_col = col(df_hist_local, 'Número de artículo', 'Numero de articulo', 'Codigo')
        _alm_col = col(df_hist_local, 'Almacén', 'Almacen')
        _fec_col = col(df_hist_local, 'Fecha de contabilización', 'Fecha de contabilizacion', 'Fecha')
        _qty_col = col(df_hist_local, 'Cantidad')
        _doc_col = col(df_hist_local, 'Documento')
    except ValueError as _ve:
        warn(f"Consumos: {_ve}")
        return {}

    if incluir_traslados:
        mask = (df_hist_local[_qty_col] < 0)
    else:
        traslados = _identificar_traslados(df_hist_local)
        mask = (df_hist_local[_qty_col] < 0) & (~df_hist_local[_doc_col].isin(traslados))
    salidas = df_hist_local.loc[mask, [_sku_col, _alm_col, _fec_col, _qty_col]].copy()
    salidas['_fec'] = salidas[_fec_col].apply(parse_hist_date)
    salidas = salidas.dropna(subset=['_fec'])
    if salidas.empty:
        return {}
    salidas['_consumo'] = -salidas[_qty_col]
    sem_info = salidas['_fec'].apply(lambda f: _semana_sabado(f))
    salidas['_sem_id'] = sem_info.apply(lambda t: t[0])
    if semanas_objetivo is not None:
        salidas = salidas[salidas['_sem_id'].isin(semanas_objetivo)]
        if salidas.empty:
            return {}
    grouped = (salidas.groupby([_sku_col, _alm_col, '_sem_id'])['_consumo']
                       .sum().reset_index())
    out = {}
    for _, r in grouped.iterrows():
        cod = str(r[_sku_col]).strip()
        alm = str(r[_alm_col]).strip()
        sem = r['_sem_id']
        if cod and alm and sem:
            out[(cod, alm, sem)] = float(r['_consumo'])
    return out

def _cargar_plan_mrp(xlsx_path):
    """Carga la hoja Plan_MRP. Normaliza tipos. Devuelve DataFrame o vacío."""
    if _PLAN_MRP_SHEET not in sheets_available:
        return pd.DataFrame(columns=_PLAN_MRP_HDRS)
    try:
        df = xl.parse(_PLAN_MRP_SHEET)
        df.columns = [str(c).strip() for c in df.columns]
        for h in _PLAN_MRP_HDRS:
            if h not in df.columns:
                df[h] = None
        df = df[_PLAN_MRP_HDRS].copy()
        df['Codigo']  = df['Codigo'].astype(str).str.strip()
        df['Almacen'] = df['Almacen'].astype(str).str.strip()
        # Normalizar Semana_ID: aceptar fecha o string YYYY-MM-DD
        def _norm_sem(row):
            sid = row['Semana_ID']
            sini = row['Semana_Inicio']
            ts = None
            if pd.notna(sini):
                ts = pd.Timestamp(sini) if not isinstance(sini, pd.Timestamp) else sini
            elif isinstance(sid, str) and sid:
                try:
                    ts = pd.Timestamp(sid)
                except Exception:
                    ts = None
            elif isinstance(sid, (pd.Timestamp, datetime.datetime)):
                ts = pd.Timestamp(sid)
            if ts is None or pd.isna(ts):
                return pd.Series([None, None])
            sem_id, sem_ini = _semana_sabado(ts)
            return pd.Series([sem_id, sem_ini])
        _se = df.apply(_norm_sem, axis=1)
        df['Semana_ID']     = _se[0]
        df['Semana_Inicio'] = _se[1]
        # Plan_MRP acepta filas con Cantidad_Plan vacía (semana en curso sin plan
        # aún capturado pero con consumo real visible) — se mostrarán con plan=–
        df['Cantidad_Plan'] = pd.to_numeric(df['Cantidad_Plan'], errors='coerce')
        df = df[(df['Codigo'] != '') & (df['Codigo'] != 'nan')
                & (df['Almacen'] != '') & (df['Almacen'] != 'nan')
                & df['Semana_ID'].notna()
                & (df['Cantidad_Plan'].isna() | (df['Cantidad_Plan'] > 0))].copy()
        return df.reset_index(drop=True)
    except Exception as _e:
        warn(f"Plan_MRP: error de lectura ({_e})")
        return pd.DataFrame(columns=_PLAN_MRP_HDRS)

def _load_hist_plan_mrp_det():
    if HISTORICO_PLAN_MRP_DET.exists():
        try:
            with open(str(HISTORICO_PLAN_MRP_DET), 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as _e:
            warn(f"Histórico plan MRP detalle: {_e}")
    return {'snapshots': {}}

def _save_hist_plan_mrp_det(d):
    try:
        with open(str(HISTORICO_PLAN_MRP_DET), 'w', encoding='utf-8') as f:
            json.dump(d, f, ensure_ascii=False, indent=2, default=str)
    except Exception as _e:
        warn(f"No se pudo guardar histórico plan MRP detalle: {_e}")

def _marcar_fecha_congelado_excel(xlsx_path, filas_a_marcar, fecha_ts):
    """Actualiza la columna Fecha_Congelado en la hoja Plan_MRP del Excel maestro
    para las filas indicadas (lista de índices 0-based dentro del df cargado).
    filas_a_marcar: lista de tuplas (Codigo, Almacen, Semana_ID) para localizar."""
    if not filas_a_marcar:
        return True
    try:
        import openpyxl as _xl
        _wb = _xl.load_workbook(str(xlsx_path))
        if _PLAN_MRP_SHEET not in _wb.sheetnames:
            _wb.close()
            return False
        _ws = _wb[_PLAN_MRP_SHEET]
        # Mapear headers → col index
        _hdr = {}
        for _ci in range(1, _ws.max_column + 1):
            _v = _ws.cell(row=1, column=_ci).value
            if _v: _hdr[str(_v).strip()] = _ci
        col_cod = _hdr.get('Codigo')
        col_alm = _hdr.get('Almacen')
        col_sem = _hdr.get('Semana_ID')
        col_ini = _hdr.get('Semana_Inicio')
        col_fec = _hdr.get('Fecha_Congelado')
        if not all([col_cod, col_alm, col_fec]):
            _wb.close()
            warn("Plan_MRP: headers incompletos, no se marcó Fecha_Congelado.")
            return False
        target = set((str(c).strip(), str(a).strip(), str(s).strip()) for c, a, s in filas_a_marcar)
        marcadas = 0
        for _ri in range(2, _ws.max_row + 1):
            _cod = _ws.cell(row=_ri, column=col_cod).value
            _alm = _ws.cell(row=_ri, column=col_alm).value
            _sem_v = _ws.cell(row=_ri, column=col_sem).value if col_sem else None
            _ini_v = _ws.cell(row=_ri, column=col_ini).value if col_ini else None
            # Normalizar semana de esta fila
            _ts = None
            if isinstance(_ini_v, (datetime.datetime, datetime.date)):
                _ts = pd.Timestamp(_ini_v)
            elif _sem_v:
                try: _ts = pd.Timestamp(_sem_v)
                except Exception: _ts = None
            elif _ini_v:
                try: _ts = pd.Timestamp(_ini_v)
                except Exception: _ts = None
            if _ts is None or pd.isna(_ts): continue
            _sid, _ = _semana_sabado(_ts)
            _k = (str(_cod).strip() if _cod else '', str(_alm).strip() if _alm else '', _sid or '')
            if _k in target:
                _ws.cell(row=_ri, column=col_fec, value=fecha_ts)
                marcadas += 1
        _wb.save(str(xlsx_path))
        _wb.close()
        ok(f"Plan_MRP: {marcadas} filas marcadas Fecha_Congelado={fecha_ts.strftime('%Y-%m-%d %H:%M')}")
        return True
    except PermissionError:
        warn("No se pudo escribir Fecha_Congelado en Excel (archivo abierto). Snapshot quedó en JSON.")
        return False
    except Exception as _e:
        warn(f"Error escribiendo Fecha_Congelado: {_e}")
        return False

# ── Asegurar hoja Plan_MRP ─────────────────────────────────────────────────
_ensure_plan_mrp_sheet(INVENTARIO_MAESTRO)

# ── Cargar plan vivo ───────────────────────────────────────────────────────
df_plan_mrp = _cargar_plan_mrp(INVENTARIO_MAESTRO)

# ── Calcular consumos semanales (todas las semanas) ────────────────────────
# Dos vistas: sin traslados (default) y con traslados (modo "incluir todo")
consumos_sem      = _calcular_consumos_semanales(df_hist, incluir_traslados=False)
consumos_sem_full = _calcular_consumos_semanales(df_hist, incluir_traslados=True)

# ── Snapshots pendientes de congelar ───────────────────────────────────────
hist_pmrp_det = _load_hist_plan_mrp_det()
hist_pmrp_det.setdefault('snapshots', {})
filas_congeladas_marcar = []

if not df_plan_mrp.empty:
    flag_congelar = (df_plan_mrp['Congelar'].astype(str).str.upper().str.strip()
                     .isin(['X', 'TRUE', '1', 'SI', 'SÍ', 'YES']))
    flag_no_fecha = df_plan_mrp['Fecha_Congelado'].isna() | (
                    df_plan_mrp['Fecha_Congelado'].astype(str).str.strip() == '')
    pendientes = df_plan_mrp[flag_congelar & flag_no_fecha]
    if not pendientes.empty:
        _ahora = pd.Timestamp.now().replace(microsecond=0)
        for sem_id, grp in pendientes.groupby('Semana_ID'):
            filas_snap = []
            for _, r in grp.iterrows():
                cod = r['Codigo']; alm = r['Almacen']
                plan_v = float(r['Cantidad_Plan'])
                real_v = consumos_sem.get((cod, alm, sem_id))
                d_abs = (real_v - plan_v) if real_v is not None else None
                d_pct = (d_abs / plan_v * 100.0) if (d_abs is not None and plan_v) else None
                filas_snap.append({
                    'codigo': cod, 'almacen': alm,
                    'plan': plan_v,
                    'consumo_real': real_v,
                    'desv_abs': d_abs,
                    'desv_pct': d_pct,
                })
                filas_congeladas_marcar.append((cod, alm, sem_id))
            sem_ini_str = ''
            try:
                _gi = grp['Semana_Inicio'].iloc[0]
                if pd.notna(_gi):
                    sem_ini_str = pd.Timestamp(_gi).strftime('%Y-%m-%d')
            except Exception:
                pass
            hist_pmrp_det['snapshots'][sem_id] = {
                'fecha_congelado': _ahora.isoformat(),
                'semana_inicio':   sem_ini_str or sem_id,
                'filas':           filas_snap,
            }
        _save_hist_plan_mrp_det(hist_pmrp_det)
        _marcar_fecha_congelado_excel(INVENTARIO_MAESTRO, filas_congeladas_marcar, _ahora)
        ok(f"Plan_MRP: congelados {len(filas_congeladas_marcar)} renglones en {len(pendientes['Semana_ID'].unique())} semana(s)")

# ── Refrescar consumo_real de snapshots previos cuya semana ya cerró ────────
# (si el snapshot se congeló con consumo_real=null porque la semana estaba abierta)
_refrescados = 0
_today_ts = pd.Timestamp(TODAY).normalize()
for sem_id, snap in hist_pmrp_det.get('snapshots', {}).items():
    try:
        sem_ini = pd.Timestamp(snap.get('semana_inicio') or sem_id)
    except Exception:
        continue
    sem_fin = sem_ini + pd.Timedelta(days=6)
    if sem_fin >= _today_ts:
        continue  # semana aún abierta, no recalcular para no falsear cierre
    for fila in snap.get('filas', []):
        if fila.get('consumo_real') is None:
            real_v = consumos_sem.get((fila['codigo'], fila['almacen'], sem_id))
            if real_v is not None:
                fila['consumo_real'] = real_v
                plan_v = fila.get('plan') or 0
                fila['desv_abs'] = real_v - plan_v
                fila['desv_pct'] = (fila['desv_abs'] / plan_v * 100.0) if plan_v else None
                _refrescados += 1
if _refrescados:
    _save_hist_plan_mrp_det(hist_pmrp_det)
    ok(f"Plan_MRP: refrescado consumo_real en {_refrescados} filas de snapshots cerrados")

# ── Construir dataset combinado plan_vs_cons ────────────────────────────────
# Para cada (Codigo, Almacen, Semana_ID) presente en plan vivo o en snapshots:
#   - plan: del snapshot si congelado, si no del plan vivo
#   - consumo_real: del snapshot (si existe) o calculado del histórico
#   - congelado: bool + fecha
_plan_vivo_map = {}
if not df_plan_mrp.empty:
    for _, r in df_plan_mrp.iterrows():
        _plan_vivo_map[(r['Codigo'], r['Almacen'], r['Semana_ID'])] = {
            'plan_vivo': float(r['Cantidad_Plan']) if pd.notna(r['Cantidad_Plan']) else None,
            'notas':     str(r['Notas']) if pd.notna(r['Notas']) else '',
            'congelado_flag': bool((str(r['Congelar']).strip().upper() in ['X','TRUE','1','SI','SÍ','YES'])
                                   and pd.notna(r['Fecha_Congelado']) and str(r['Fecha_Congelado']).strip() != ''),
        }
_snap_map = {}
for sem_id, snap in hist_pmrp_det.get('snapshots', {}).items():
    fc = snap.get('fecha_congelado', '')
    for fila in snap.get('filas', []):
        _snap_map[(fila['codigo'], fila['almacen'], sem_id)] = {
            'plan_snap':    fila.get('plan'),
            'real_snap':    fila.get('consumo_real'),
            'fecha_congelado': fc,
        }

# Mapa de descripción por código (probar Master, luego Costo)
_desc_map = {}
try:
    _mas_cod = col(df_mas, 'SKU', 'Código', 'Codigo', 'Número de artículo', 'Numero de articulo')
    _mas_des = col(df_mas, 'Descripcion', 'Descripción', 'Descripción del artículo', 'desc')
    for _, r in df_mas.iterrows():
        _c = str(r[_mas_cod]).strip()
        _d = str(r[_mas_des]).strip() if pd.notna(r[_mas_des]) else ''
        if _c and _c != 'nan' and _d:
            _desc_map[_c] = _d
except Exception:
    pass
# Fallback / complemento desde Costo
try:
    _cos_cod = col(df_cos, 'Número de artículo', 'Numero de articulo', 'Código', 'Codigo', 'SKU')
    _cos_des = col(df_cos, 'Descripción del artículo', 'Descripcion', 'Descripción', 'desc')
    for _, r in df_cos.iterrows():
        _c = str(r[_cos_cod]).strip()
        _d = str(r[_cos_des]).strip() if pd.notna(r[_cos_des]) else ''
        if _c and _c != 'nan' and _d and _c not in _desc_map:
            _desc_map[_c] = _d
except Exception:
    pass

# Mapa de grupo (desde Costo si disponible)
_grupo_map = {}
try:
    _cos_cod2 = col(df_cos, 'Número de artículo', 'Numero de articulo', 'Código', 'Codigo')
    _cos_grp  = col(df_cos, 'Nombre de grupo', 'Grupo', 'grupo')
    for _, r in df_cos.iterrows():
        _c = str(r[_cos_cod2]).strip()
        _g = str(r[_cos_grp]).strip() if pd.notna(r[_cos_grp]) else ''
        if _c and _c != 'nan':
            _grupo_map[_c] = _g
except Exception:
    pass

# Las filas mostradas vienen estrictamente del archivo fuente:
#   - Plan vivo capturado en hoja Plan_MRP (df_plan_mrp)
#   - Snapshots históricos congelados (_historico_plan_mrp_detalle.json)
# Para cada (cod, alm, sem_id) capturada, el consumo real se calcula del
# Historico Movs:
#   * Si la semana ya cerró → consumo total de esa semana
#   * Si la semana está abierta (en curso) → consumo parcial sábado→hoy
# (consumos_sem ya devuelve esto automáticamente por SKU×Alm×Sem)
_keys = set(_plan_vivo_map.keys()) | set(_snap_map.keys())
plan_vs_cons_rows = []
for (cod, alm, sem_id) in _keys:
    snap = _snap_map.get((cod, alm, sem_id))
    vivo = _plan_vivo_map.get((cod, alm, sem_id))
    plan_v = snap['plan_snap'] if snap else (vivo['plan_vivo'] if vivo else None)
    # Consumo SIN traslados (default). Si la semana está congelada,
    # respeta el snapshot histórico. Si la semana sigue abierta,
    # consumos_sem ya devuelve el parcial (suma desde sábado hasta hoy).
    real_v = snap['real_snap'] if snap and snap['real_snap'] is not None \
             else consumos_sem.get((cod, alm, sem_id))
    # Consumo INCLUYENDO traslados internos (vista alternativa, no se guarda
    # en snapshots para no romper histórico congelado)
    real_full_v = consumos_sem_full.get((cod, alm, sem_id))
    d_abs = (real_v - plan_v) if (real_v is not None and plan_v is not None) else None
    d_pct = (d_abs / plan_v * 100.0) if (d_abs is not None and plan_v) else None
    # Estado
    estado = 'on'
    if d_pct is None:        estado = 'na'
    elif abs(d_pct) > 25:    estado = 'red'
    elif abs(d_pct) > 10:    estado = 'yel'
    # Etiqueta WK y bandera de semana abierta
    _wk = _wk_label(sem_id)
    _abierta = _semana_abierta(sem_id, _today_ts)
    try:
        _sem_fin_str = (pd.Timestamp(sem_id) + pd.Timedelta(days=6)).strftime('%Y-%m-%d') if sem_id else ''
    except Exception:
        _sem_fin_str = ''
    plan_vs_cons_rows.append({
        'cod':       cod,
        'desc':      _desc_map.get(cod, ''),
        'grupo':     _grupo_map.get(cod, ''),
        'alm':       alm,
        'sem':       sem_id,            # fecha sábado (YYYY-MM-DD) – mantiene clave
        'wk':        _wk,               # 'WK20'
        'sem_ini':   sem_id,            # inicio (sábado)
        'sem_fin':   _sem_fin_str,      # cierre (viernes)
        'abierta':   _abierta,          # True si la semana aún no ha cerrado
        'plan':      round(plan_v, 2) if plan_v is not None else None,
        'real':      round(real_v, 2) if real_v is not None else None,
        'real_full': round(real_full_v, 2) if real_full_v is not None else None,
        'd_abs':     round(d_abs, 2) if d_abs is not None else None,
        'd_pct':     round(d_pct, 1) if d_pct is not None else None,
        'est':       estado,
        'cong':      snap['fecha_congelado'] if snap else '',
        'notas':     vivo['notas'] if vivo else '',
    })

# Ordenar por semana DESC, luego SKU
plan_vs_cons_rows.sort(key=lambda r: (r['sem'] or '', r['cod'] or '', r['alm'] or ''), reverse=True)

# KPIs globales (sobre semanas con plan + real disponibles)
_rows_valid = [r for r in plan_vs_cons_rows if r['plan'] is not None and r['real'] is not None]
_sum_plan = sum(r['plan'] for r in _rows_valid) or 1
_sum_min  = sum(min(r['plan'], r['real']) for r in _rows_valid)
_pct_cumpl = (_sum_min / _sum_plan * 100.0) if _sum_plan else 0
_n_on  = sum(1 for r in _rows_valid if r['est'] == 'on')
_n_yel = sum(1 for r in _rows_valid if r['est'] == 'yel')
_n_red = sum(1 for r in _rows_valid if r['est'] == 'red')
_sum_dabs = sum(r['d_abs'] for r in _rows_valid if r['d_abs'] is not None)

# Lista de semanas únicas (DESC) para filtros
_semanas_pvc = sorted(set(r['sem'] for r in plan_vs_cons_rows if r['sem']), reverse=True)
_almacenes_pvc = sorted(set(r['alm'] for r in plan_vs_cons_rows if r['alm']))

plan_vs_cons_data = {
    'rows':       plan_vs_cons_rows,
    'semanas':    _semanas_pvc,
    'almacenes':  _almacenes_pvc,
    'kpis': {
        'pct_cumplimiento': round(_pct_cumpl, 1),
        'n_total_eval':     len(_rows_valid),
        'n_on':             _n_on,
        'n_yel':            _n_yel,
        'n_red':            _n_red,
        'sum_dabs':         round(_sum_dabs, 2),
        'n_plan_total':     len(plan_vs_cons_rows),
        'n_snapshots':      len(hist_pmrp_det.get('snapshots', {})),
    },
}

ok(f"Plan vs Consumo: {len(plan_vs_cons_rows)} filas | Cumplimiento {round(_pct_cumpl,1)}% | "
   f"on={_n_on} yel={_n_yel} red={_n_red} | snapshots={len(hist_pmrp_det.get('snapshots', {}))}")

# ══════════════════════════════════════════════════════════════════════════════
#  EMPACAR JSON UNIFICADO
# ══════════════════════════════════════════════════════════════════════════════

step("Leyendo tracking desde _Tracking_BP.xlsx")

TRACKING_FILE = BASE_DIR / '_Tracking_BP.xlsx' if (BASE_DIR / '_Tracking_BP.xlsx').exists() else REPORTES_DIR / '_Tracking_BP.xlsx'

_TRACK_DEFS = {
    '_Track_Compras':   ['Codigo', 'Descripcion', 'Proveedor', 'Qty_pedida',
                         'fecha_pedido', 'fecha_llegada', 'estatus', 'comentarios'],
    '_Track_GLM':       ['Codigo', 'Descripcion', 'folio', 'Qty_disponible',
                         'estatus', 'fecha_carga', 'comentarios'],
    '_Track_Traslados': ['clave', 'Codigo', 'Descripcion', 'Alm_origen',
                         'Alm_destino', 'Qty', 'owner', 'estatus',
                         'fecha_traslado', 'comentarios'],
    '_Track_Capacidad': ['almacen', 'new_cap', 'prev_cap',
                         'fecha_cambio', 'responsable', 'comentarios'],
}

_TRACK_COL_WIDTHS = {
    '_Track_Compras':   [10, 30, 20, 12, 14, 14, 14, 35],
    '_Track_GLM':       [10, 30, 14, 14, 14, 14, 35],
    '_Track_Traslados': [12, 10, 30, 12, 12, 12, 16, 14, 14, 35],
    '_Track_Capacidad': [10, 12, 12, 14, 18, 35],
}

def _apply_sheet_style(ws, hdrs, col_widths, tab_color=None):
    try:
        from openpyxl.styles import PatternFill, Font, Alignment
        _hdr_fill = PatternFill('solid', fgColor='1E3A5F')
        _hdr_font = Font(bold=True, color='FFFFFF', size=10)
        for _ci, (_h, _w) in enumerate(zip(hdrs, col_widths + [20]*(len(hdrs)-len(col_widths))), 1):
            _cell = ws.cell(row=1, column=_ci, value=_h)
            _cell.fill = _hdr_fill
            _cell.font = _hdr_font
            _cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[
                __import__('openpyxl').utils.get_column_letter(_ci)
            ].width = _w
        ws.row_dimensions[1].height = 18
        ws.freeze_panes = 'A2'
        if tab_color:
            ws.sheet_properties.tabColor = tab_color
    except Exception:
        for _ci, _h in enumerate(hdrs, 1):
            ws.cell(row=1, column=_ci, value=_h)

def _ensure_tracking_file(trk_path):
    try:
        import openpyxl as _xl
        _is_new = not trk_path.exists()
        if _is_new:
            _wb = _xl.Workbook()
            if 'Sheet' in _wb.sheetnames:
                del _wb['Sheet']
        else:
            try:
                _wb = _xl.load_workbook(str(trk_path))
            except Exception:
                _wb = _xl.Workbook()
                if 'Sheet' in _wb.sheetnames:
                    del _wb['Sheet']
                _is_new = True

        _changed = _is_new
        _tab_colors = {
            '_Track_Compras':   '2563EB',
            '_Track_GLM':       '16A34A',
            '_Track_Traslados': 'D97706',
            '_Track_Capacidad': '7C3AED',
        }
        for _sn, _hdrs in _TRACK_DEFS.items():
            _needs_rebuild = False
            if _sn not in _wb.sheetnames:
                _needs_rebuild = True
            else:
                _ws_chk = _wb[_sn]
                _rows_chk = list(_ws_chk.values)
                if not _rows_chk or len([c for c in _rows_chk[0] if c]) < len(_hdrs):
                    _needs_rebuild = True

            if _needs_rebuild:
                _saved_rows = []
                if _sn in _wb.sheetnames:
                    _ws_old = _wb[_sn]
                    _old_rows = list(_ws_old.values)
                    if len(_old_rows) > 1:
                        _old_hdrs = [str(h) if h else '' for h in _old_rows[0]]
                        for _row in _old_rows[1:]:
                            if _row and _row[0] is not None:
                                _saved_rows.append(dict(zip(_old_hdrs, _row)))
                    del _wb[_sn]

                _ws = _wb.create_sheet(_sn)
                _apply_sheet_style(_ws, _hdrs, _TRACK_COL_WIDTHS.get(_sn, []), _tab_colors.get(_sn))

                if _sn == '_Track_Capacidad':
                    _saved_by_alm = {str(r.get('almacen','')).strip(): r for r in _saved_rows}
                    for _ri, _alm in enumerate(ALMACENES_OPERATIVOS, 2):
                        _prev = _saved_by_alm.get(_alm, {})
                        _ws.cell(row=_ri, column=1, value=_alm)
                        _ws.cell(row=_ri, column=2, value=_prev.get('new_cap') or None)
                        _ws.cell(row=_ri, column=3, value=_prev.get('prev_cap') or CAP_RESPALDO.get(_alm, ''))
                        _ws.cell(row=_ri, column=4, value=_prev.get('fecha_cambio') or None)
                        _ws.cell(row=_ri, column=5, value=_prev.get('responsable') or None)
                        _ws.cell(row=_ri, column=6, value=_prev.get('comentarios') or None)
                else:
                    for _ri, _row_d in enumerate(_saved_rows, 2):
                        for _ci, _h in enumerate(_hdrs, 1):
                            _ws.cell(row=_ri, column=_ci, value=_row_d.get(_h))
                _changed = True

        if _changed:
            _wb.save(str(trk_path))
            ok(f"_Tracking_BP.xlsx {'creado' if _is_new else 'actualizado'} — pestañas con formato completo")
        _wb.close()
    except Exception as _e:
        warn(f"No se pudo crear/actualizar _Tracking_BP.xlsx: {_e}")

_ensure_tracking_file(TRACKING_FILE)

def _read_tracking_tab(trk_path, sheet_name):
    try:
        import openpyxl as _xl
        if not trk_path.exists():
            return {}
        _wb = _xl.load_workbook(str(trk_path), read_only=True, data_only=True)
        if sheet_name not in _wb.sheetnames:
            _wb.close()
            return {}
        _ws = _wb[sheet_name]
        _rows = list(_ws.values)
        _wb.close()
        if not _rows:
            return {}
        _headers = [str(h).strip() if h else None for h in _rows[0]]
        result = {}
        for _row in _rows[1:]:
            if not _row or _row[0] is None:
                continue
            _key = str(_row[0]).strip()
            _fields = {}
            for _i, _h in enumerate(_headers[1:], 1):
                if _h and _i < len(_row) and _row[_i] is not None:
                    _fields[_h] = _row[_i]
            result[_key] = _fields
        return result
    except Exception as _e:
        warn(f"Error leyendo pestaña {sheet_name} de tracking: {_e}")
        return {}

tracking_compras   = _read_tracking_tab(TRACKING_FILE, '_Track_Compras')
tracking_glm       = _read_tracking_tab(TRACKING_FILE, '_Track_GLM')
tracking_traslados = _read_tracking_tab(TRACKING_FILE, '_Track_Traslados')
tracking_capacidad = _read_tracking_tab(TRACKING_FILE, '_Track_Capacidad')

def _load_json_fallback(json_path, existing_dict):
    try:
        if json_path.exists():
            with open(str(json_path), 'r', encoding='utf-8') as _f:
                _jdata = json.load(_f)
            if _jdata:
                _merged = {**_jdata, **existing_dict}
                return _merged
    except Exception:
        pass
    return existing_dict

tracking_compras   = _load_json_fallback(SISTEMA_DIR / '_tracking_compras.json',   tracking_compras)
tracking_glm       = _load_json_fallback(SISTEMA_DIR / '_tracking_glm.json',       tracking_glm)
tracking_traslados = _load_json_fallback(SISTEMA_DIR / '_tracking_traslados.json', tracking_traslados)
tracking_capacidad = _load_json_fallback(SISTEMA_DIR / '_tracking_capacidad.json', tracking_capacidad)

def _persist_tracking_to_excel(trk_path, all_tracking):
    try:
        import openpyxl as _xl
        _sheet_map = {
            '_Track_Compras':   ('Codigo',  tracking_compras),
            '_Track_GLM':       ('Codigo',  tracking_glm),
            '_Track_Traslados': ('clave',   tracking_traslados),
            '_Track_Capacidad': ('almacen', tracking_capacidad),
        }
        if trk_path.exists():
            _wb = _xl.load_workbook(str(trk_path))
        else:
            _wb = _xl.Workbook()
            if 'Sheet' in _wb.sheetnames:
                del _wb['Sheet']
        for _sn, (_key_col, _data_dict) in _sheet_map.items():
            if not _data_dict:
                continue
            if _sn in _wb.sheetnames:
                _ws = _wb[_sn]
            else:
                _ws = _wb.create_sheet(_sn)
                _apply_sheet_style(_ws, _TRACK_DEFS.get(_sn, [_key_col]),
                                   _TRACK_COL_WIDTHS.get(_sn, []))
            _hdrs = _TRACK_DEFS.get(_sn, [_key_col])
            for _ri in range(2, _ws.max_row + 2):
                for _ci in range(1, len(_hdrs) + 1):
                    _ws.cell(row=_ri, column=_ci, value=None)
            for _ri, (_key, _fields) in enumerate(_data_dict.items(), 2):
                _ws.cell(row=_ri, column=1, value=_key)
                for _ci, _h in enumerate(_hdrs[1:], 2):
                    _val = _fields.get(_h)
                    if _val is not None:
                        _ws.cell(row=_ri, column=_ci, value=_val)
        _wb.save(str(trk_path))
        _wb.close()
        ok(f"Tracking persistido en _Tracking_BP.xlsx")
    except Exception as _e:
        warn(f"No se pudo persistir tracking en Excel: {_e}")

_persist_tracking_to_excel(TRACKING_FILE, {
    '_Track_Compras':   tracking_compras,
    '_Track_GLM':       tracking_glm,
    '_Track_Traslados': tracking_traslados,
    '_Track_Capacidad': tracking_capacidad,
})

ok(f"Tracking compras: {len(tracking_compras)} | GLM: {len(tracking_glm)} | Traslados: {len(tracking_traslados)} | Capacidad: {len(tracking_capacidad)}")

# ── Inyectar campos editables (Folio_OC, Estatus, Fecha_Carga, Owner) de los Tracking xlsx a mrp_plan ──
try:
    # Usar _master_track (cargado ANTES de regenerar los Excels) para inyectar al JSON.
    # Esto garantiza que los datos editados por el usuario nunca se pierdan entre refreshes.
    for _r in mrp_plan:
        _k = (_r.get('Codigo',''), _r.get('Almacen',''))
        _entry = _master_track.get(_k, {})
        _r['Folio_OC']    = _entry.get('Folio_OC','')
        _r['Estatus_trk'] = _entry.get('Estatus','')
        _r['Fecha_Carga'] = _entry.get('Fecha_Carga','')
        _r['Owner']       = _entry.get('Owner','')
        _r['Comentarios'] = _entry.get('Comentarios','')
    ok(f"Campos de tracking inyectados en mrp_plan ({_n_con_datos} registros con datos del usuario)")
except Exception as _ei:
    warn(f"No se pudo inyectar tracking en mrp_plan: {_ei}")

step("Serializando JSON unificado")

# ── SIX Planeación ─────────────────────────────────────────────────────────
six_list   = []
six_semana = 0
six_fechas = []
if 'Six' in sheets_available:
    try:
        # Leer con header=None para acceder al raw completo (Six tiene 3 filas de metadata)
        _raw6 = pd.read_excel(str(INVENTARIO_MAESTRO), sheet_name='Six',
                              engine='openpyxl', header=None)

        # Semana: fila 0, col 3
        try:
            six_semana = int(_raw6.iloc[0, 3]) if pd.notna(_raw6.iloc[0, 3]) else 0
        except Exception:
            six_semana = 0

        # Fechas: fila 2, cualquier columna con datetime
        try:
            _frow6 = _raw6.iloc[2]
            for _ci6 in range(len(_frow6)):
                _v6 = _frow6[_ci6]
                if pd.notna(_v6) and hasattr(_v6, 'strftime'):
                    six_fechas.append(_v6.strftime('%Y-%m-%d'))
            six_fechas = six_fechas[:7]
        except Exception:
            six_fechas = []

        # Header real: fila 3; datos: fila 4 en adelante
        _hdr6 = [str(v).strip() for v in _raw6.iloc[3]]
        _data6 = _raw6.iloc[4:].copy()
        _data6.columns = _hdr6
        # Quitar primeras 2 cols (nan nan antes de Cliente)
        _data6 = _data6.iloc[:, 2:].copy()

        # Filtrar filas sin Cliente
        if 'Cliente' in _data6.columns:
            _data6 = _data6[_data6['Cliente'].notna() & (_data6['Cliente'].astype(str).str.strip() != '')]

        # Normalizar cols numéricas
        for _dc6 in ['Lunes','Martes','Miércoles','Jueves','Viernes','Sábado','Domingo','Totales',
                     'Cajas','Pallets','Total cajas fisicas','Total cajas equivalentes',
                     'Prioridad','Factor']:
            if _dc6 in _data6.columns:
                _data6[_dc6] = pd.to_numeric(_data6[_dc6], errors='coerce').fillna(0)

        _data6 = _data6.fillna('')
        six_list = _data6.to_dict(orient='records')
        ok(f"Six (Sem {six_semana}): {len(six_list)} registros | fechas: {six_fechas}")
    except Exception as _ex6:
        warn(f"Error procesando hoja Six: {_ex6}")
else:
    warn("Hoja Six no encontrada — pestaña Six estará vacía")

D = {
    'ver_piezas': ver_piezas, 'total_conteos': total_conteos,
    'total_piezas_dif': total_piezas_dif, 'n_pendientes': n_pendientes,
    'total_faltante': total_faltante, 'total_sobrante': total_sobrante,
    'total_neto': total_neto, 'total_dif_count': total_dif_count,
    'total_dif_pzs_neg': total_dif_pzs_neg, 'total_dif_pzs_pos': total_dif_pzs_pos,
    'gen_date': TODAY.strftime('%d/%m/%Y'),
    'ver_alm':   ver_alm.to_dict(orient='records'),
    'imp_neg':   imp_neg.to_dict(orient='records'),
    'imp_pos':   imp_pos.to_dict(orient='records'),
    'diffs_all': diffs_out.to_dict(orient='records'),
    'neteos':    neteos_list,
    'pendientes': pendientes_list,
    'cov_30':    cov_merged.to_dict(orient='records'),
    'cap_changes': cap_changes,
    'total_cap_pct': total_cap_pct,
    'cap_df':    cap_df.to_dict(orient='records'),
    'total_costo_inv': round(total_costo),
    'costo_alm': costo_alm.to_dict(orient='records'),
    'sku_cost_top': sku_cost[['Código','desc','En stock','costo_unit','valor']
                             ].rename(columns={'En stock':'stock'}).to_dict(orient='records'),
    'movs_by_alm_date': movs_by_alm_date.to_dict(orient='records'),
    'sal_by_date_alm': sal_by_date_alm.to_dict(orient='records'),
    'hist_date_min': hist_date_min, 'hist_date_max': hist_date_max,
    'total_cob_pct': total_cob_pct,
    'm_alms':    M_ALMS, 'total_inv_codes': total_inv_codes,
    'abc_summary': abc_summary.to_dict(orient='records'),
    'abc_full':  abc_full.to_dict(orient='records'),
    'no_mov':    no_mov_list,
    'top10':     top10,
    'meses':     MESES_VALIDOS,
    'alms_hist': alms_hist,
    'trend':     trend_pivot,

    'mrp_gen_date': datetime.datetime.now().strftime('%d/%m/%Y %H:%M'),
    'mrp_resumen':  mrp_resumen,
    'mrp_plan':     mrp_plan,
    'cambios_plan': cambios_plan_list,
    'cortos_resumen': cortos_resumen,
    'inv_nac':      inv_nac_list,
    'mrp_por_alm':  mrp_por_alm,
    'glm_prov':     glm_prov,
    'inv_sin_demanda': inv_sin_demanda_list,
    'planeacion':      planeacion_list,

    'six_planeacion':     six_list,
    'six_semana':         six_semana,
    'six_fechas':         six_fechas,

    'mate2122':           mate2122_list,
    'mate2122_total_mxn': total_mate2122_mxn,
    'mate2122_total_pzs': total_mate2122_pzs,

    'cap_detail': cap_detail_list,
    'zona_map':   ZONA_MAP,

    'tracking_compras':   tracking_compras,
    'tracking_glm':       tracking_glm,
    'tracking_traslados': tracking_traslados,
    'tracking_capacidad': tracking_capacidad,

    'mrp_sal_diario': mrp_sal_diario,
    'var_movs': var_movs,

    'hist_summary':  hist_summary,
    'hist_full':     hist_full,
    'semana_actual': semana_key,
    'semana_label':  semana_label,

    'plan_vs_cons':  plan_vs_cons_data,
}

data_json = json.dumps(D, ensure_ascii=False, default=str)
ok(f"JSON generado: {len(data_json):,} chars")

# ══════════════════════════════════════════════════════════════════════════════
#  GENERAR HTML UNIFICADO
# ══════════════════════════════════════════════════════════════════════════════

def generar_pdf_resumen(D, output_path):
    step("Generando PDF Ejecutivo")
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        
        doc = SimpleDocTemplate(
            str(output_path),
            pagesize=letter,
            rightMargin=40, leftMargin=40,
            topMargin=40, bottomMargin=40
        )
        story = []
        styles = getSampleStyleSheet()
        
        # Colores corporativos BP
        PURPLE_BP = colors.HexColor("#6B1FA2")
        RED_BP = colors.HexColor("#C8102E")
        DARK_TEXT = colors.HexColor("#2D3748")
        LIGHT_BG = colors.HexColor("#F5F0FA")
        
        title_style = ParagraphStyle(
            'TitleStyle',
            parent=styles['Heading1'],
            fontSize=20,
            textColor=PURPLE_BP,
            spaceAfter=5
        )
        subtitle_style = ParagraphStyle(
            'SubtitleStyle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor("#718096"),
            spaceAfter=15
        )
        section_style = ParagraphStyle(
            'SectionStyle',
            parent=styles['Heading2'],
            fontSize=12,
            textColor=PURPLE_BP,
            spaceBefore=10,
            spaceAfter=5
        )
        cell_style = ParagraphStyle(
            'CellStyle',
            parent=styles['Normal'],
            fontSize=8,
            textColor=DARK_TEXT
        )
        cell_bold_style = ParagraphStyle(
            'CellBoldStyle',
            parent=styles['Normal'],
            fontSize=8,
            textColor=DARK_TEXT,
            fontName='Helvetica-Bold'
        )
        cell_header_style = ParagraphStyle(
            'CellHeaderStyle',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.white,
            fontName='Helvetica-Bold',
            alignment=1 # Center
        )
        
        # PAGE 1: INDICADORES CLAVE Y ALMACENES
        # 1. Header
        story.append(Paragraph("🍓 Berries Paradise — Resumen Ejecutivo MRP", title_style))
        story.append(Paragraph(f"Generado automáticamente el {D.get('mrp_gen_date', 'N/A')} | Inteligencia de Materiales", subtitle_style))
        story.append(Spacer(1, 8))
        
        # 2. KPIs Principales
        story.append(Paragraph("📊 INDICADORES CLAVE DE CONTROL (KPIS)", section_style))
        
        pct_cumpl = D.get('plan_vs_cons', {}).get('kpis', {}).get('pct_cumplimiento', 0.0)
        
        kpis_data = [
            [
                Paragraph("Indicador", cell_header_style), 
                Paragraph("Valor", cell_header_style), 
                Paragraph("Estatus", cell_header_style)
            ],
            [
                Paragraph("Veracidad de Inventario (Conteos)", cell_bold_style),
                Paragraph(f"{D.get('ver_piezas', 0.0):.2f}%", cell_style),
                Paragraph("Excelente (>=95%)" if D.get('ver_piezas', 0.0) >= 95 else "Atención Requerida (<95%)", cell_style)
            ],
            [
                Paragraph("Cumplimiento Plan vs Consumo", cell_bold_style),
                Paragraph(f"{pct_cumpl:.1f}%", cell_style),
                Paragraph("Estable (>=80%)" if pct_cumpl >= 80 else "Desviación (<80%)", cell_style)
            ],
            [
                Paragraph("Ocupación de Almacenes (Tarimas)", cell_bold_style),
                Paragraph(f"{D.get('total_cap_pct', 0.0):.1f}%", cell_style),
                Paragraph("Normal (<85%)" if D.get('total_cap_pct', 0.0) < 85 else "Saturación (>=85%)", cell_style)
            ],
            [
                Paragraph("Cobertura de Inventario (>= 2 semanas)", cell_bold_style),
                Paragraph(f"{D.get('total_cob_pct', 0.0):.1f}%", cell_style),
                Paragraph("Criterio OK (>=60%)" if D.get('total_cob_pct', 0.0) >= 60 else "Alerta Cobertura (<60%)", cell_style)
            ],
            [
                Paragraph("Valor del Déficit MRP (Cortos)", cell_bold_style),
                Paragraph(f"${D.get('total_faltante', 0.0):,.0f} MXN", cell_style),
                Paragraph("Bajo Control" if D.get('total_faltante', 0.0) > -50000 else "Alerta Déficit", cell_style)
            ]
        ]
        
        t_kpis = Table(kpis_data, colWidths=[200, 140, 160])
        t_kpis.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), PURPLE_BP),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('TOPPADDING', (0,0), (-1,-1), 5),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, LIGHT_BG]),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#C4A0DC")),
        ]))
        story.append(t_kpis)
        story.append(Spacer(1, 15))
        
        # 3. Almacenes más Críticos por Ocupación
        story.append(Paragraph("📦 OCUPACIÓN FÍSICA DE ALMACENES (TOP CRÍTICOS)", section_style))
        alm_headers = [
            Paragraph("Almacén", cell_header_style),
            Paragraph("Tarimas Usadas", cell_header_style),
            Paragraph("Capacidad Total", cell_header_style),
            Paragraph("% Ocupación", cell_header_style)
        ]
        alm_rows = [alm_headers]
        
        sorted_alms = sorted(D.get('cap_df', []), key=lambda x: x.get('pct_ocup', 0.0), reverse=True)[:5]
        for a in sorted_alms:
            alm_rows.append([
                Paragraph(str(a.get('alm')), cell_bold_style),
                Paragraph(f"{a.get('pallets_used', 0.0):,.1f}", cell_style),
                Paragraph(f"{a.get('cap', 0):,}", cell_style),
                Paragraph(f"{a.get('pct_ocup', 0.0):.1f}%", cell_style)
            ])
            
        t_alms = Table(alm_rows, colWidths=[120, 120, 120, 142])
        t_alms.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), PURPLE_BP),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('TOPPADDING', (0,0), (-1,-1), 5),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, LIGHT_BG]),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#C4A0DC")),
        ]))
        story.append(t_alms)
        
        # PAGE BREAK
        story.append(PageBreak())
        
        # PAGE 2: CORTOS, REFILL Y GLM
        # 4. SKUs Cortos Críticos (Sin Stock / Desabasto)
        story.append(Paragraph("🚨 ARTÍCULOS EN CORTO CRÍTICO (SIN STOCK / DESABASTO)", section_style))
        short_headers = [
            Paragraph("SKU", cell_header_style),
            Paragraph("Descripción", cell_header_style),
            Paragraph("Almacén", cell_header_style),
            Paragraph("Stock Actual", cell_header_style),
            Paragraph("Mínimo", cell_header_style),
            Paragraph("Déficit (Pzs)", cell_header_style),
            Paragraph("Valor Déficit", cell_header_style)
        ]
        short_rows = [short_headers]
        
        cortos = [r for r in D.get('mrp_plan', []) if r.get('Dec_externa') == 'Corto' and not r.get('Is_refill')]
        # fall back to any cortos if list is empty
        if not cortos:
            cortos = [r for r in D.get('mrp_plan', []) if r.get('Dec_externa') == 'Corto']
        sorted_cortos = sorted(cortos, key=lambda x: float(x.get('Valor_deficit', 0.0) or 0.0), reverse=True)[:5]
        
        if sorted_cortos:
            for c in sorted_cortos:
                short_rows.append([
                    Paragraph(str(c.get('Codigo')), cell_bold_style),
                    Paragraph(str(c.get('Descripcion', 'N/D'))[:24], cell_style),
                    Paragraph(str(c.get('Almacen')), cell_style),
                    Paragraph(f"{int(float(c.get('Stock', 0.0) or 0.0)):,}", cell_style),
                    Paragraph(f"{int(float(c.get('Min', 0.0) or 0.0)):,}", cell_style),
                    Paragraph(f"{int(float(c.get('Deficit', 0.0) or 0.0)):,}", cell_style),
                    Paragraph(f"${float(c.get('Valor_deficit', 0.0) or 0.0):,.0f} MXN", cell_style)
                ])
        else:
            short_rows.append([Paragraph("Sin cortos críticos en el sistema", cell_style), "", "", "", "", "", ""])
            
        t_shorts = Table(short_rows, colWidths=[55, 117, 55, 75, 70, 75, 85])
        t_shorts.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), PURPLE_BP),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('TOPPADDING', (0,0), (-1,-1), 5),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, LIGHT_BG]),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#C4A0DC")),
        ]))
        story.append(t_shorts)
        story.append(Spacer(1, 10))
        
        # 5. Refill Crítico
        story.append(Paragraph("🔄 REFILL CRÍTICO REQUERIDO (REABASTOS URGENTES)", section_style))
        refill_headers = [
            Paragraph("SKU", cell_header_style),
            Paragraph("Descripción", cell_header_style),
            Paragraph("Almacén", cell_header_style),
            Paragraph("Stock Actual", cell_header_style),
            Paragraph("Mínimo", cell_header_style),
            Paragraph("Déficit (Pzs)", cell_header_style),
            Paragraph("Consumo/Sem", cell_header_style),
            Paragraph("Valor Déficit", cell_header_style)
        ]
        refill_rows = [refill_headers]
        
        refills = [r for r in D.get('mrp_plan', []) if r.get('Is_refill')]
        sorted_refills = sorted(refills, key=lambda x: float(x.get('Valor_deficit', 0.0) or 0.0), reverse=True)[:5]
        
        if sorted_refills:
            for r in sorted_refills:
                refill_rows.append([
                    Paragraph(str(r.get('Codigo')), cell_bold_style),
                    Paragraph(str(r.get('Descripcion', 'N/D'))[:24], cell_style),
                    Paragraph(str(r.get('Almacen')), cell_style),
                    Paragraph(f"{int(float(r.get('Stock', 0.0) or 0.0)):,}", cell_style),
                    Paragraph(f"{int(float(r.get('Min', 0.0) or 0.0)):,}", cell_style),
                    Paragraph(f"{int(float(r.get('Deficit', 0.0) or 0.0)):,}", cell_style),
                    Paragraph(f"{round(float(r.get('Dem_wk_hist', 0.0) or 0.0)):,}", cell_style),
                    Paragraph(f"${float(r.get('Valor_deficit', 0.0) or 0.0):,.0f} MXN", cell_style)
                ])
        else:
            refill_rows.append([Paragraph("Sin reabastos/refills pendientes", cell_style), "", "", "", "", "", "", ""])
            
        t_refills = Table(refill_rows, colWidths=[50, 112, 50, 60, 55, 60, 65, 80])
        t_refills.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), PURPLE_BP),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('TOPPADDING', (0,0), (-1,-1), 5),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, LIGHT_BG]),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#C4A0DC")),
        ]))
        story.append(t_refills)
        story.append(Spacer(1, 10))
        
        # 6. GLM Crítico
        story.append(Paragraph("📦 GLM CRÍTICO (DISPONIBILIDAD PROVEEDORES)", section_style))
        glm_headers = [
            Paragraph("SKU", cell_header_style),
            Paragraph("Descripción", cell_header_style),
            Paragraph("Almacén", cell_header_style),
            Paragraph("Stock Local", cell_header_style),
            Paragraph("Mínimo", cell_header_style),
            Paragraph("Déficit (Pzs)", cell_header_style),
            Paragraph("Inventario GLM", cell_header_style),
            Paragraph("Valor Déficit", cell_header_style)
        ]
        glm_rows = [glm_headers]
        
        glms = [r for r in D.get('mrp_plan', []) if r.get('Dec_externa') == 'GLM']
        sorted_glms = sorted(glms, key=lambda x: float(x.get('Valor_deficit', 0.0) or 0.0), reverse=True)[:5]
        
        if sorted_glms:
            for g in sorted_glms:
                glm_rows.append([
                    Paragraph(str(g.get('Codigo')), cell_bold_style),
                    Paragraph(str(g.get('Descripcion', 'N/D'))[:24], cell_style),
                    Paragraph(str(g.get('Almacen')), cell_style),
                    Paragraph(f"{int(float(g.get('Stock', 0.0) or 0.0)):,}", cell_style),
                    Paragraph(f"{int(float(g.get('Min', 0.0) or 0.0)):,}", cell_style),
                    Paragraph(f"{int(float(g.get('Deficit', 0.0) or 0.0)):,}", cell_style),
                    Paragraph(f"{int(float(g.get('Inv_glm', 0.0) or 0.0)):,}", cell_style),
                    Paragraph(f"${float(g.get('Valor_deficit', 0.0) or 0.0):,.0f} MXN", cell_style)
                ])
        else:
            glm_rows.append([Paragraph("Sin órdenes GLM pendientes", cell_style), "", "", "", "", "", "", ""])
            
        t_glms = Table(glm_rows, colWidths=[50, 112, 50, 60, 55, 60, 65, 80])
        t_glms.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), PURPLE_BP),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('TOPPADDING', (0,0), (-1,-1), 5),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, LIGHT_BG]),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#C4A0DC")),
        ]))
        story.append(t_glms)
        
        doc.build(story)
        ok(f"PDF ejecutivo generado con éxito: {output_path.name}")
    except Exception as pdf_err:
        warn(f"Error generando PDF ejecutivo: {pdf_err}")

step("Generando HTML unificado")

with open(str(TEMPLATE), 'r', encoding='utf-8') as f:
    html = f.read()

if '/*D_PLACEHOLDER*/' not in html:
    print("\nERROR: No se encontró '/*D_PLACEHOLDER*/' en el template.")
    sys.exit(1)

html_final = html.replace('/*D_PLACEHOLDER*/', data_json)

with open(str(DASHBOARD), 'w', encoding='utf-8') as f:
    f.write(html_final)

sz = DASHBOARD.stat().st_size
ok(f"Dashboard guardado: {DASHBOARD.name} ({sz/1024:.0f} KB)")

# ── Subir JSON consolidado a Supabase ──
if SUPABASE_URL and SUPABASE_KEY:
    try:
        step("Subiendo JSON consolidado a Supabase...")
        _payload_dict = json.loads(data_json)
        _sup_data = {
            "id": 1,
            "payload": _payload_dict,
            "updated_at": datetime.datetime.now(datetime.timezone.utc).isoformat()
        }
        _success = supabase_upsert("dashboard_payload", _sup_data)
        if _success:
            ok("JSON consolidado subido exitosamente a Supabase Cloud!")
        else:
            warn("No se pudo subir el JSON consolidado a Supabase.")
    except Exception as _esup_up:
        warn(f"Error subiendo JSON a Supabase: {_esup_up}")

# Generar PDF Resumen Ejecutivo
pdf_resumen_path = REPORTES_DIR / "Resumen_Ejecutivo_BP.pdf"
generar_pdf_resumen(D, pdf_resumen_path)

print(f"\n{'='*62}")
print(f"  LISTO — {DASHBOARD.name} ({sz/1024:.0f} KB)")
print(f"  Inventarios: {ver_piezas}% veracidad | {total_inv_codes} SKUs")
print(f"  MRP Plan: {mrp_resumen['n_total']} SKU×Alm | {mrp_resumen['n_corto']} cortos")
print(f"{'='*62}\n")

# ══════════════════════════════════════════════════════════════════════════════
#  PUBLICACIÓN AUTOMÁTICA EN GITHUB PAGES
#  Copia el HTML al repo y hace git commit + push
# ══════════════════════════════════════════════════════════════════════════════
import subprocess, shutil, time, pathlib as _pathlib

GITHUB_REPO_DIR = BASE_DIR
GITHUB_REMOTE_URL = "https://github.com/hectorrenteriabp/MaterialesBP.git"

GIT_EXE = shutil.which("git")

def _git(*args, cwd):
    """Ejecuta un comando git y devuelve (returncode, stdout+stderr)."""
    if not GIT_EXE:
        raise RuntimeError("Git no encontrado en el sistema.")
    
    result = subprocess.run(
        [GIT_EXE] + list(args),
        cwd=str(cwd),
        capture_output=True, text=True, encoding='utf-8', errors='replace'
    )
    return result.returncode, (result.stdout + result.stderr).strip()

print(f"\n{'-'*62}")
print("  Publicando en GitHub Pages...")

try:
    if not GIT_EXE:
        raise RuntimeError("git no encontrado en PATH. Instala Git Desktop o Git para Windows.")

    # NUEVO: Pausamos 2 segundos para darle tiempo a Windows de mostrar el archivo generado
    time.sleep(2)
    
    # NUEVO: Verificamos que el archivo se haya creado físicamente antes de intentar subirlo
    if not DASHBOARD.exists():
         raise RuntimeError(f"¡El archivo {DASHBOARD.name} no se generó! Revisa si OneDrive lo está bloqueando.")

    # ── Inicializar repo si la carpeta existe pero no tiene .git ─────────────
    _git_dir = GITHUB_REPO_DIR / ".git"
    if not _git_dir.exists():
        print("  Repo git no inicializado — configurando por primera vez...")
        rc, out = _git("init", cwd=GITHUB_REPO_DIR)
        if rc != 0: raise RuntimeError(f"git init falló: {out}")
        print(f"    git init OK")

        rc, out = _git("remote", "add", "origin", GITHUB_REMOTE_URL, cwd=GITHUB_REPO_DIR)
        if rc != 0: raise RuntimeError(f"git remote add falló: {out}")
        print(f"    Remote configurado: {GITHUB_REMOTE_URL}")

        print("    Descargando historial remoto (git fetch)...")
        rc, out = _git("fetch", "origin", cwd=GITHUB_REPO_DIR)
        if rc != 0: raise RuntimeError(f"git fetch falló: {out}\n    Verifica conexión a internet y credenciales.")

        rc, out = _git("checkout", "-B", "main", "--track", "origin/main", cwd=GITHUB_REPO_DIR)
        if rc != 0:
            rc, out = _git("checkout", "-b", "main", cwd=GITHUB_REPO_DIR)
            if rc != 0: raise RuntimeError(f"git checkout main falló: {out}")
        print("    Rama main configurada OK")

    print(f"  Preparando archivo para Git: {DASHBOARD.name}")

    # ── Staging ───────────────────────────────────────────────────────────────
    rc, out = _git("add", DASHBOARD.name, cwd=GITHUB_REPO_DIR)
    if rc != 0:
        raise RuntimeError(f"git add falló: {out}")

    # ── Asegurar identidad git local (por si el usuario nunca configuró git) ──
    _git("config", "user.email", "dashboard@berriesparadise.com", cwd=GITHUB_REPO_DIR)
    _git("config", "user.name",  "BP Dashboard", cwd=GITHUB_REPO_DIR)

    # ── Verificar si hay cambios staged ──────────────────────────────────────
    rc_diff, _ = _git("diff", "--cached", "--quiet", cwd=GITHUB_REPO_DIR)
    if rc_diff == 0:
        print("  Sin cambios nuevos que publicar en GitHub (ya estaba al día).")
    else:
        fecha_commit = TODAY.strftime('%d/%m/%Y %H:%M')
        rc, out = _git("commit", "-m", f"Dashboard BP actualizado {fecha_commit}", cwd=GITHUB_REPO_DIR)
        if rc != 0: raise RuntimeError(f"git commit falló: {out}")
        print(f"  Commit OK: {out.splitlines()[0] if out else ''}")

        # Push — forzamos la subida (-f) para sobrescribir cualquier historial previo en GitHub
        rc, out = _git("push", "-f", "origin", "main", cwd=GITHUB_REPO_DIR)
        if rc != 0:
            rc, out = _git("push", "-f", "--set-upstream", "origin", "main", cwd=GITHUB_REPO_DIR)
        if rc != 0:
            raise RuntimeError(f"git push falló:\n{out}")
        print(f"  [OK] GitHub Pages actualizado correctamente.")
        if out:
            for _ln in out.splitlines()[-3:]:
                if _ln.strip():
                    print(f"     {_ln.strip()}")

except Exception as _ge:
    print(f"\n  [!]  No se pudo publicar en GitHub:")
    print(f"      {_ge}")
    print(f"")
    print(f"  Posibles causas:")
    print(f"    1. La carpeta del repo no existe en esa ruta")
    print(f"    2. Git no está instalado (descarga en https://git-scm.com)")
    print(f"    3. Credenciales de GitHub no configuradas")
    print(f"    4. Sin conexión a internet")

print(f"{'-'*62}")

# ══════════════════════════════════════════════════════════════════════════════
#  ENVÍO DE CORREO AUTOMÁTICO VÍA OUTLOOK LOCAL
# ══════════════════════════════════════════════════════════════════════════════
print("  Enviando Resumen Ejecutivo por Correo (Outlook)...")

try:
    import win32com.client as win32
    
    # Iniciamos sesión de Outlook
    print("    Inicializando conexión con Outlook local...")
    outlook = win32.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)
    
    # Configuración de destinatarios y asunto
    mail.To = "hector.renteria@berriesparadise.com"
    mail.Subject = "📋 Resumen Ejecutivo Diario - Materiales & Planeación Berries Paradise"
    
    # Plantilla HTML estilizada
    mail.HTMLBody = """
    <div style="font-family: 'Segoe UI', Arial, sans-serif; max-width: 600px; margin: 0 auto; border: 1px solid #e2e8f0; border-radius: 8px; overflow: hidden; box-shadow: 0 4px 6px rgba(0,0,0,0.05);">
       <!-- Encabezado -->
       <div style="background: linear-gradient(135deg, #6B1FA2 0%, #C8102E 100%); padding: 24px; text-align: center; color: white;">
         <h1 style="margin: 0; font-size: 22px; font-weight: 600; letter-spacing: 0.5px;">Materiales & Planeación</h1>
         <p style="margin: 4px 0 0 0; font-size: 14px; opacity: 0.9;">Berries Paradise ®</p>
       </div>
       
       <!-- Contenido -->
       <div style="padding: 24px; background-color: #ffffff; color: #334155; line-height: 1.6;">
         <p style="margin-top: 0; font-size: 16px;">Hola <strong>Hector</strong>,</p>
         <p style="font-size: 14px;">Se ha generado y actualizado con éxito el <strong>Resumen Ejecutivo Diario</strong> correspondiente al día de hoy.</p>
         
         <div style="background-color: #f8fafc; border-left: 4px solid #6B1FA2; padding: 12px 16px; margin: 20px 0; border-radius: 0 4px 4px 0;">
           <span style="font-weight: 600; color: #6B1FA2; display: block; margin-bottom: 4px;">📈 Estado del Reporte:</span>
           <ul style="margin: 0; padding-left: 20px; font-size: 13px;">
             <li><strong>Origen:</strong> Actualización en el Servidor/PC de Planeación.</li>
             <li><strong>Entregable:</strong> PDF de 2 páginas (Adjunto en este correo).</li>
             <li><strong>Canal Web:</strong> Dashboard Unificado interactivo disponible en la nube.</li>
           </ul>
         </div>

         <p style="font-size: 14px;">El PDF adjunto contiene el análisis detallado de:</p>
         <table style="width: 100%; border-collapse: collapse; font-size: 13px; margin: 15px 0;">
           <tr style="border-bottom: 1px solid #f1f5f9;">
             <td style="padding: 8px 0; font-weight: 600; color: #C8102E;">⚠️ Cortos Críticos</td>
             <td style="padding: 8px 0; text-align: right; color: #64748b;">Artículos con stock por debajo del mínimo</td>
           </tr>
           <tr style="border-bottom: 1px solid #f1f5f9;">
             <td style="padding: 8px 0; font-weight: 600; color: #6B1FA2;">🔄 Refill Requerido</td>
             <td style="padding: 8px 0; text-align: right; color: #64748b;">Propuestas de abastecimiento MRP</td>
           </tr>
           <tr style="border-bottom: 1px solid #f1f5f9;">
             <td style="padding: 8px 0; font-weight: 600; color: #d97706;">📦 Capacidad GLM</td>
             <td style="padding: 8px 0; text-align: right; color: #64748b;">Nivel de ocupación en almacenes clave</td>
           </tr>
         </table>

         <div style="text-align: center; margin: 25px 0 10px 0;">
           <a href="https://hectorrenteriabp.github.io/MaterialesBP/BP_Dashboard_Unificado.html" style="background-color: #6B1FA2; color: white; padding: 12px 24px; text-decoration: none; border-radius: 6px; font-weight: 600; font-size: 14px; display: inline-block; box-shadow: 0 4px 6px rgba(107, 31, 162, 0.2);">💻 Abrir Dashboard Interactivo</a>
         </div>
       </div>

       <!-- Pie de página -->
       <div style="background-color: #f1f5f9; padding: 16px; text-align: center; font-size: 11px; color: #94a3b8; border-top: 1px solid #e2e8f0;">
         <p style="margin: 0;">Este es un servicio automatizado de Berries Paradise. No responda a este correo electrónico.</p>
         <p style="margin: 4px 0 0 0;">© 2026 Berries Paradise S.A. de C.V.</p>
       </div>
     </div>
    """
    
    # Adjuntar el archivo PDF
    pdf_path = REPORTES_DIR / "Resumen_Ejecutivo_BP.pdf"
    if pdf_path.exists():
        mail.Attachments.Add(str(pdf_path.resolve()))
        mail.Send()
        print("  [OK] Correo enviado correctamente a Hector Rentería.")
    else:
        print("  [!] Advertencia: No se encontró el PDF en la ruta de Reportes, no se envió el correo.")
        
except Exception as _ee:
    print(f"  [!] No se pudo enviar el correo vía Outlook:")
    print(f"      {_ee}")
    print("      Verifica que pywin32 esté instalado y Outlook esté configurado localmente.")

print(f"{'-'*62}")

# ── PAUSA FINAL — para que la ventana no se cierre ─────────────────────────
import sys as _sys, os as _os
if _sys.platform == 'win32' and _sys.stdout.isatty():
    print()
    print("  Presiona ENTER para cerrar esta ventana...")
    try:
        input()
    except Exception:
        pass
